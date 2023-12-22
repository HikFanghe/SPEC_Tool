import os, sys
import numpy as np
import pandas as pd
import warnings
import openpyxl
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import Filters, FilterColumn
import ctypes
import multiprocessing
from UI.spec import Ui_MainWindow
from UI.Help_win import Ui_HelpWindow
from PyQt5.QtCore import pyqtSignal, QThread, QCoreApplication
from PyQt5.QtGui import QIcon
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMainWindow, QMessageBox, QFileDialog
import qdarkstyle
from qdarkstyle.light.palette import LightPalette
from pythonping import ping
from itertools import chain

from lib.Integ_Isapi import *
from Case.specCase import SpecTestCase
from lib.mtd_com_fun import *
from win32com.client import Dispatch
import os
import json
import requests
import traceback
from Config import SpecConfig
from lib.defaultFun import *
from lib.Integ_Isapi import Integ_Isapi


class AutoVivification(dict):
    """定义多级字典"""

    def __getitem__(self, item):
        try:
            return dict.__getitem__(self, item)
        except KeyError:
            value = self[item] = type(self)()
            return value

warnings.filterwarnings('ignore')
# 忽视警告

BASE_PATH = os.getcwd()


def my_print(msg, level='INFO'):
    """
        自定义打印函数
    """
    # 获取被调用函数在被调用时所处代码行数
    line = sys._getframe().f_back.f_lineno
    # 获取被调用函数所在模块文件名
    file_name = sys._getframe(1).f_code.co_filename
    file_name = file_name.replace("\\", "/").split("/")[-1]
    try:
        testLogObj = open(testLogFile, "a+")
        print(f"[{str(datetime.datetime.now())[:19]} {file_name} {line}] [{level}]{msg}")
        print(f"[{str(datetime.datetime.now())[:19]} {file_name} {line}] [{level}]{msg}", file=testLogObj)
        testLogObj.close()
    except Exception as e:
        testLogObj = open(testLogFile, "a+")
        print(traceback.format_exc(), file=testLogFile)
        print(f"[ERROR]将执行打印写入日志文件出现异常:{str(e)}!", file=testLogFile)
        testLogObj.close()


def get_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")

    return os.path.normpath(os.path.join(base_path, relative_path))

def service_cfg(service_type="SPEC_VERSION"):
    """查询服务配置信息"""
    url = 'http://hice.hikvision.com.cn/gw1/frontends_group_qd/CATTDataManage/CloudParameter/Operation'
    data = {
        "name": service_type,
        "userName": "linfanghe",
        "operation": "view"
    }

    try:
        searchReq = requests.post(url, data=json.dumps(data), timeout=30)
        searchData = searchReq.json()
        searchReq.close()
        if not searchData['Success']:
            return False, f'连接云空间获取{service_type}云服务配置失败,请联系集成测试工具开发组排查!'
        CFG = searchData["msg"]["currentVersion"]
        return True, CFG
    except Exception as ex:
        print(traceback.print_exc())
        return False, f'连接云空间获取{service_type}云服务配置出现异常:{str(ex)}!请联系集成测试工具开发组排查!'

def excepthook(exc_type, exc_value, exc_traceback):
    # 将异常信息转化为字符串
    exception_str = ''.join(traceback.format_exception(exc_type, exc_value, exc_traceback))

    # my_print(f"[崩溃]程序发生异常崩溃:{exception_str}", "ERROR")

    # 显示崩溃提醒框
    QMessageBox.critical(None, "异常", f"SPEC工具出现异常了，请联系[林芳禾]排查，请勿关闭此弹窗或截图保留！！！\n异常信息：{exception_str}")

    # 退出应用程序
    # QApplication.quit()

class spec_test(QThread, SpecTestCase):
    # 信号槽返回
    trans_result = pyqtSignal(int, str, str)

    def __init__(self, ip, username, password, local_path, device_type, spec_excel_path, productOpt):
        super().__init__()
        self.ip = ip
        self.username = username
        self.password = password
        self.local_path = local_path

        # 重置父类中的属性用于HTTPDigestAuth
        self.integIsapi_setInit(username, password)
        self.mtdComFun_setInit(username, password)

        # 测试报告路径
        self.spec_excel_path = spec_excel_path
        # sepc文档路径
        self.device_type = device_type
        self.stopFlag = False
        # 产品/软件参数
        self.productOpt = productOpt

        ##获取设备基本信息
        chanFlag, chanMsg = self.get_channel()
        if chanFlag:
            self.chans_num = chanMsg
        else:
            print(chanMsg)

        # 设备是否支持AIOP#
        self.AIOP_cap = False
        self.AIOP_list = None
        self.AIOP_Version = None
        AIOP_Flag, AIOP_Msg, AIOP_Version = self.is_heop()
        if AIOP_Flag:
            self.AIOP_cap = AIOP_Flag
            self.AIOP_list = AIOP_Msg
            self.AIOP_Version = AIOP_Version

        # 设备类型(IPD/IPC)
        dev_Flag, dev_Msg = self.get_deviceInfo()
        if dev_Flag:
            self.devType = dev_Msg
        else:
            self.devType = None

        self.deviceInfo = {
            "IP": self.ip,
            "U_NAME": self.username,
            "P_WORD": self.password,
            "chalNum": self.chans_num,
            "dev_type": self.devType
        }

        # 获取智能资源
        self.IntegSupportList = self.Integ_support_list(self.chans_num, self.ip, self.AIOP_cap, self.AIOP_list)

        # 判断设备是否是海外设备
        self.isOverseas = False
        isOverseasFlag, isOverseasMsg = self.device_is_overseas_judge()
        if isOverseasFlag:
            self.isOverseas = isOverseasMsg

        # 获取设备全部通道码流分辨率和帧率
        self.all_channel_stream_flag, self.all_channel_stream_msg = self.get_all_channel_stream()
        if not self.all_channel_stream_flag:
            self.trans_result.emit(False, f"设备{self.ip}获取设备全部通道码流分辨率和帧率失败，原因为{self.all_channel_stream_msg}", self.ip)

        # 用例：[结果处理方法，结果比对方法]
        self.resultHandle = SpecConfig.resultHandle
        # 产品参数字典
        self.ProductParamDict = SpecConfig.ProductParamDict
        # 固定写法的规格项
        self.FixedOutputDict = SpecConfig.FixedOutputDict
        # 获取规格项字典
        self.SpecItemDict = SpecConfig.SpecItemDict

    def excel_creat(self, save_path):
        # 动态创建excel测试报告表
        try:
            # 获取模板中所有的子项
            all_content = self.get_all_content()
            if self.productOpt != '产品参数 (包含软件参数)':
                all_content = {key: value for key, value in all_content.items() if value == self.productOpt}
            print(all_content)

            file_local = self.spec_excel_path
            df = pd.read_excel(file_local)
            max_rows = df.shape[0] + 1
            print(f"最大行数为{max_rows}")

            # 打开源文件
            source_file = openpyxl.load_workbook(file_local, data_only=True)
            source_sheet = source_file.active
            # 创建目标文件
            target_file = openpyxl.Workbook()
            target_sheet = target_file.active

            # 拆分所有的合并单元格，并赋予合并之前的值。
            # 由于openpyxl并没有提供拆分并填充的方法，所以使用该方法进行完成

            all_merged_cell_ranges = list(source_sheet.merged_cells.ranges)

            for merged_cell_range in all_merged_cell_ranges:
                merged_cell = merged_cell_range.start_cell
                source_sheet.unmerge_cells(range_string=merged_cell_range.coord)
                for row_index, col_index in merged_cell_range.cells:
                    cell = source_sheet.cell(row=row_index, column=col_index)
                    cell.value = merged_cell.value

            ###### 根据flag信息确定是否生成规格项校验报告
            checkItemNameFlag = True
            if checkItemNameFlag:
                checkCount = 3
                custom_thrid_font = Font(name='Calibri', size=12)
                custom_third_align = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                               wrap_text=True)
                self.checkItemNameDict = SpecConfig.itemNameCheck
                checkItemfilePath = save_path + f'\\{os.path.basename(file_local).split(".")[0]}_{self.ip}_itemCheck.xlsx'
                checkItemWb = openpyxl.Workbook()
                checkItemSheet = checkItemWb.active
                self.checkItemNameExcel(checkItemSheet)

            # 循环复制数据和格式-在目标文件中输入spec真值
            pre = ''
            i = 4
            excludeFlag = False
            for row in range(5, max_rows + 1):
                for col in [2, 1, 4]:
                    cell = source_sheet.cell(column=col, row=row)
                    if col == 2:
                        item = re.sub('\s', '', cell.value)

                        ############## 名称校验 ##############
                        newItem = ""
                        if item in self.checkItemNameDict:
                            newItem = self.checkItemNameDict.get(item, -1)
                            if newItem != item:
                                checkItemSheet[f'A{checkCount}'].value = item
                                checkItemSheet[f'B{checkCount}'].value = newItem
                                checkItemSheet[f'A{checkCount}'].font = custom_thrid_font
                                checkItemSheet[f'A{checkCount}'].alignment = custom_third_align
                                checkItemSheet[f'B{checkCount}'].font = custom_thrid_font
                                checkItemSheet[f'B{checkCount}'].alignment = custom_third_align

                                # item = newItem if newItem != -1 else item
                                checkCount += 1

                        if item in all_content or newItem in all_content:
                            i += 1
                            excludeFlag = True

                            ####### 若在名称校验字典里的，报告名仍以spec文件的规格项名称呈现，索引采用兼容后的名称
                            if (newItem != "") and (newItem != -1):
                                target_sheet.cell(row=i, column=cell.column + 1, value=item)
                                target_sheet.cell(row=i, column=1, value=all_content[newItem])
                            else:
                                target_sheet.cell(row=i, column=cell.column + 1, value=item)
                                target_sheet.cell(row=i, column=1, value=all_content[item])
                    elif col == 1:
                        # if excludeFlag:
                        #     if cell.value is None:
                        #         target_sheet.cell(row=i, column=cell.column + 1, value=pre)
                        #     else:
                        #         pre = cell.value
                        #         target_sheet.cell(row=i, column=cell.column + 1, value=cell.value)
                        if excludeFlag:
                            target_sheet.cell(row=i, column=cell.column + 1, value=cell.value)
                    elif col == 4:
                        if excludeFlag:
                            target_sheet.cell(row=i, column=cell.column, value=cell.value)
                            excludeFlag = False

            # 保存测试报告目标文件
            file_name = os.path.basename(file_local).split(".")[0]
            save_local = save_path + f'\\{file_name}_{self.ip}_report.xlsx'
            target_file.save(save_local)

            # 保存规格项校验报告
            if checkItemNameFlag:
                checkItemWb.save(checkItemfilePath)

            df = pd.read_excel(file_local)
            hang = int(df.shape[0])
            # 产品名称
            data1 = df.iloc[0].values
            data1 = list(data1)
            data2 = df.iloc[1].values
            data2 = list(data2)

            workbook = openpyxl.load_workbook(save_local)
            worksheet = workbook.active
            custom_first_font = Font(name='Arial', size=16, bold=True, color=Color(rgb='000000'))
            custom_second_font = Font(name='微软雅黑', size=13, bold=True)
            custom_thrid_font = Font(name='微软雅黑', size=15, bold=True)
            custom_fourth_font = Font(name='Calibri', size=14)

            for cell in worksheet['A']:
                cell.font = Font(name='Calibri', size=14, bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            for row_num in ['B', 'C', 'D', 'E', 'F', 'G']:
                for cell in worksheet[row_num]:
                    cell.font = custom_fourth_font
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)


            cell = worksheet['A1']
            cell.value = "SPEC测试报告"
            cell.font = custom_first_font
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill('solid',fgColor='483D8B')
            cell = worksheet['A2']
            cell.value = data1[0]
            cell = worksheet['A3']
            cell.value = data2[0]
            cell = worksheet['A4']
            cell.value = "规格项类型"
            cell = worksheet['B4']
            cell.value = "模块名"
            cell = worksheet['C4']
            cell.value = "子项"
            cell = worksheet['D4']
            cell.value = "SPEC值"
            cell = worksheet['E4']
            cell.value = "实际设备值"
            cell = worksheet['F4']
            cell.value = "比对结果"
            cell = worksheet['G4']
            cell.value = '''备注：\n1.当设备实际值与SPEC值比对成功时比对结果为P，比对失败时比对结果为F；\n2.当规格项不支持自动化时，设备实际值为NT，比对结果为check，请手动确认。'''

            worksheet.merge_cells('A1:G1')
            worksheet.merge_cells('A2:G2')
            worksheet.merge_cells('A3:G3')
            worksheet['A2'].font = custom_second_font
            worksheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet['A2'].fill = PatternFill('solid', fgColor='D3D3D3')
            worksheet['A3'].font = custom_second_font
            worksheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet['A3'].fill = PatternFill('solid', fgColor='D3D3D3')
            for cell in worksheet[4]:
                cell.font = custom_thrid_font
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

            worksheet.column_dimensions['A'].width = 18
            worksheet.column_dimensions['B'].width = 18
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['F'].width = 12
            for column_letter in ['D', 'E']:
                worksheet.column_dimensions[column_letter].width = 50
            worksheet.column_dimensions['G'].width = 51

            worksheet.row_dimensions[1].height = 40
            for row_num in range(2, 4):
                worksheet.row_dimensions[row_num].height = 30
            worksheet.row_dimensions[4].height = 105

            start_row = 5
            end_row = hang + 1
            # # 遍历指定的行范围，并设置文本自动换行
            # for row_num in range(start_row, end_row + 1):
            #     worksheet.row_dimensions[row_num].height = 25
            #     for cell in worksheet[row_num]:
            #     # worksheet[row_num].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

            workbook.save(save_local)
            workbook.close()
            return True, save_local
        except Exception as e:
            print(traceback.print_exc())
            return False, ""

    def get_channel(self):
        # 获取设备的通道数
        try:
            channel_url = f"/ISAPI/Streaming/channels"
            ret = self.get(channel_url)
            channel = re.findall("<channelName>(.*?)</channelName>", ret, re.S)
            if len(channel) < 1:
                return False, "读取通道数小于1，请排查原因"
            return True, int(len(list(set(channel))))
        except Exception as e:
            print(traceback.print_exc())
            return False, f'设备获取通道数量过程中出现异常:{e}'

    def get_excel_content(self, file, col=3):
        # 获取spec的excel中所有的子项内容
        df = pd.read_excel(file, usecols=[col])
        df = np.array(df)
        df_list = df.tolist()[3:]
        return df_list

    def get_all_content(self):
        # 获取所有需要测试的子项内容
        # file_local = BASE_PATH + '\\SpecTemplate\\SPEC_data.xlsx'
        file_local = get_path('SpecTemplate/SPEC_data.xlsx')
        df = pd.read_excel(file_local)
        item = list(df.iloc[:, 1].values)
        param = list(df.loc[:, '参数类别'].values)
        # data_list = list(chain.from_iterable(data1))
        data_dict = dict(zip(item, param))
        return data_dict

    def stop(self):
        # my_print(f"设备{self.ip}的测试线程被强制停止", level="INFO")
        msg = '设备的测试线程被强制停止'
        self.VD_LOG_DEBUG(self.ip, msg, level="INFO")
        self.stopFlag = True

    def run(self):
        try:
            # 判断设备类型
            if self.device_type == '低功耗设备':
                PowerConsFlag = self.fullPowerConsumption(self.ip)
                if not PowerConsFlag:
                    msg = f"设备{self.ip} 低功耗设备切换为全功耗模式执行失败!"
                    self.VD_LOG_DEBUG(self.ip, msg, level="ERROR")
                    self.trans_result.emit(2, msg, self.ip)
                    msg = f"设备{self.ip} 请确认设备是否为低功耗设备!"
                    self.VD_LOG_DEBUG(self.ip, msg, level="ERROR")
                    self.trans_result.emit(2, msg, self.ip)
                    msg = f"设备{self.ip} 任务基于设备当前模式执行，spec结果可能有误，请手动确认!"
                    self.VD_LOG_DEBUG(self.ip, msg, level="ERROR")
                    self.trans_result.emit(2, msg, self.ip)
                else:
                    msg = f"设备{self.ip} 该低功耗设备已切换为【全功耗模式】!"
                    self.VD_LOG_DEBUG(self.ip, msg, level="INFO")
                    self.trans_result.emit(2, msg, self.ip)


            if self.device_type == '低时延设备':
                lowLatFlag, lowLatMsg = self.lowLatencytoIntelVCA()
                if lowLatFlag:
                    msg = f"设备{self.ip} 该低时延设备已切换为【{lowLatMsg}智能模式】!"
                    self.VD_LOG_DEBUG(self.ip, msg, level="INFO")
                    self.trans_result.emit(2, msg, self.ip)
                else:
                    msg = f"设备{self.ip} 低时延设备切换为【非普通监控】模式执行失败，失败原因：{lowLatMsg}！"
                    self.VD_LOG_DEBUG(self.ip, msg, level="ERROR")
                    self.trans_result.emit(2, msg, self.ip)
                    self.stop()
                    msg = f"设备{self.ip} 请确认设备是否为低时延设备!"
                    self.VD_LOG_DEBUG(self.ip, msg, level="ERROR")
                    self.trans_result.emit(2, msg, self.ip)
                    msg = f"设备{self.ip} 任务基于设备当前模式执行，spec结果可能有误，请手动确认!"
                    self.VD_LOG_DEBUG(self.ip, msg, level="ERROR")
                    self.trans_result.emit(2, msg, self.ip)


            if not os.path.exists(self.local_path):
                os.makedirs(self.local_path)
            Data = datetime.datetime.now().strftime('%Y%m%dT%H%M%S')
            excel_path = self.local_path + f"\\{Data}\\" + f"{self.ip}"
            if not os.path.exists(excel_path):
                os.makedirs(excel_path)
            print(excel_path)
            print(self.spec_excel_path)

            # 测试前动态创建excel测试报告
            save_flag, save_local = self.excel_creat(excel_path)
            if not save_flag:
                msg = f"测试报告动态创建失败!"
                # my_print(msg, level="ERROR")
                self.VD_LOG_DEBUG(self.ip, msg)
                self.trans_result.emit(1, msg, self.ip)
                return

            if self.stopFlag == True:
                self.trans_result.emit(0, f"设备{self.ip}测试正常，手动停止测试", self.ip)
                return

            # # 获取模板中所有的子项的字典
            # all_content = self.get_all_content()
            # print(all_content)

            # 获取spec测试文件中的全部子项（下列代码表示将二维列表展开为一维）
            finally_test_project = [item for sublist in self.get_excel_content(save_local, 2) for item in sublist]
            print(finally_test_project)

            # 当按下停止时停止测试
            if self.stopFlag == True:
                self.trans_result.emit(0, f"设备{self.ip}测试正常，手动停止测试", self.ip)
                return

            # 打开测试报告文件
            wb = openpyxl.load_workbook(save_local)
            sheet = wb.active

            # 【名称校验flag】
            checkItemNameFlag = True

            # 这里添加比对模块和结果输出模块

            for i in range(len(finally_test_project)):
                try:
                    item = re.sub('\s', '', str(finally_test_project[i]))
                    # 规格项名称转换
                    if checkItemNameFlag:
                        newItem = ""
                        if item in self.checkItemNameDict:
                            newItem = self.checkItemNameDict.get(item, -1)
                            item = newItem if (newItem != -1) and (newItem != "") else item

                    if item in self.SpecItemDict:
                        actualData = eval(self.SpecItemDict[item])
                        actualData = str(actualData).strip('\n')
                        if item in self.ProductParamDict:
                            FLAG = self.ProductParamDict.get(item, -1)
                            if FLAG == 0:
                                devCap = "NT"
                            elif FLAG == 1:
                                devCap = actualData
                            elif FLAG == -1:
                                devCap = f'获取【{item}】能力出现异常，请联系工具组排查！'
                            sheet[f'E{i + 5}'].value = devCap
                            sheet[f'F{i + 5}'].value = "Check"
                            sheet[f'F{i + 5}'].font = Font(name='Calibri', size=14, bold=True, color=Color(rgb='CD2626'))
                            sheet[f'G{i + 5}'].value = actualData
                        else:
                            sheet[f'E{i + 5}'].value = actualData
                            trueData = str(sheet[f'D{i + 5}'].value).strip('\n')
                            if item in self.resultHandle:
                                itemFun = self.resultHandle[item]
                                res = self.resHandleCallFun(eval(itemFun[0]), eval(itemFun[1]), actualData, trueData, delStr=itemFun[2], sep = itemFun[3])
                                sheet[f'F{i + 5}'].value = res
                                # if res == 'F':
                                #     sheet[f'F{i + 5}'].font = Font(name='Calibri', size=14, bold=True, color=Color(rgb='CD2626'))
                            else:
                                if actualData == trueData:
                                    sheet[f'F{i + 5}'].value = "P"
                                else:
                                    sheet[f'F{i + 5}'].value = "F"
                                    # sheet[f'F{i + 5}'].font = Font(name='Calibri', size=14, bold=True, color=Color(rgb='CD2626'))
                    else:
                        sheet[f'E{i + 5}'].value = "NT"
                        sheet[f'F{i + 5}'].value = "Check"
                        # sheet[f'F{i + 5}'].font = Font(name='Calibri', size=14, bold=True, color=Color(rgb='CD2626'))
                    # 输出备注信息
                    if newItem in self.FixedOutputDict:
                        sheet[f'G{i + 5}'].value = self.FixedOutputDict.get(newItem, 'Error')
                except Exception as e:
                    print(e)
                    MtdComFun.VD_LOG_DEBUG('', 'e', level="INFO")
            # 保存测试报告文件
            wb.save(save_local)

            # 按规格项类型排序
            # 读取 Excel 文件并指定引擎为 openpyxl
            # df = pd.read_excel(save_local, engine='openpyxl', skiprows=3)
            # # d代表对 B 列的数据进行排序，默认排序方式为升序，Pandas排序时默认不对第一行（通常是标题行或列名行）进行排序
            # df.sort_values(by=['规格项类型', '模块名'], inplace=True)

            ################### 测试
            df = MtdComFun.sortExcelReport(save_local)

            # 读取同一个文件
            wb = openpyxl.load_workbook(save_local)
            sheet = wb.active
            # 清除工作表中的数据
            for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None

            # 将排序后的数据写回工作表
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=5):
                for c_idx, value in enumerate(row, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    if c_idx == 6 and (value == 'F' or value == 'Check'):
                        cell.font = Font(name='Calibri', size=14, bold=True, color=Color(rgb='CD2626'))

            wb.save(save_local)
            self.trans_result.emit(0, f"设备{self.ip}测试过程成功", self.ip)
            return
        except Exception as e:
            print(traceback.print_exc())
            self.trans_result.emit(1, f"设备{self.ip}测试过程中出现异常:{e}", self.ip)
            return

# 暂时未用这个类，因为【判断设备】是否在线必须要阻塞
# class pingDevice(QThread):
#     '''设备网络在线状态检查 线程'''
#
#     # 定义【设备网络检查】信号函数
#     ping_result = pyqtSignal(int, str, str)
#
#     def __init__(self, ip):
#         super().__init__()
#         # 设备ip
#         self.ip = ip
#         # 设备网络在线标志位
#         self.onlineFlag = False
#
#     def run(self):
#         try:
#             msg = ping(self.ip)
#             if "Reply" in str(msg):
#                 self.ping_result.emit(2, f"[IP:{self.ip}]设备在线，开始创建测试任务!", self.ip)
#                 self.onlineFlag = True
#             else:
#                 self.ping_result.emit(2, f"[IP:{self.ip}]设备不在线，请检查设备或网络!", self.ip)
#         except Exception as e:
#             self.ping_result.emit(2, f"[IP:{self.ip}]检查网络过程中出错，请联系工具组同事排查!", self.ip)

class CameraTest(QMainWindow, Ui_MainWindow, Integ_Isapi):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 禁止最大化
        self.setFixedSize(self.width(), self.height())
        # 初始化按钮的使能状态
        self.run_start_test.setEnabled(True)
        self.run_stop_test.setEnabled(False)

        # 创建设备网络检查线程
        # self.pingThreadTaskList = []

    def excelDirSelect_1(self):
        file_name = QFileDialog.getOpenFileName(self, "打开excel", "C:/", "*.xlsx;;*.xls;;*.csv")
        self.pushButton_excel_1.setText(file_name[0])

    def excelDirSelect_2(self):
        file_name = QFileDialog.getOpenFileName(self, "打开excel", "C:/", "*.xlsx;;*.xls;;*.csv")
        self.pushButton_excel_2.setText(file_name[0])

    def excelDirSelect_3(self):
        file_name = QFileDialog.getOpenFileName(self, "打开excel", "C:/", "*.xlsx;;*.xls;;*.csv")
        self.pushButton_excel_3.setText(file_name[0])

    def excelDirSelect_4(self):
        file_name = QFileDialog.getOpenFileName(self, "打开excel", "C:/", "*.xlsx;;*.xls;;*.csv")
        self.pushButton_excel_4.setText(file_name[0])

    def excelDirSelect_5(self):
        file_name = QFileDialog.getOpenFileName(self, "打开excel", "C:/", "*.xlsx;;*.xls;;*.csv")
        self.pushButton_excel_5.setText(file_name[0])

    def closeEvent(self, event):
        """
            重写关闭窗口事件函数
        """
        checkFlag = self.messageBox("确定退出软件吗?")
        if not checkFlag:
            event.ignore()

    def messageBox(self, msg):
        """
            弹窗函数
        """
        reply = QMessageBox.question(self,
                                     'PlayerDemo',
                                     msg,
                                     QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)
        if reply == QMessageBox.Yes:
            # my_print("用户选择：YES!")
            MtdComFun.VD_LOG_DEBUG('', '用户选择：YES', level="INFO")
            return True
        else:
            # my_print("用户选择：NO或关闭弹窗!")
            MtdComFun.VD_LOG_DEBUG('', '"用户选择：NO或关闭弹窗!"', level="INFO")
            # 清空提示栏字符
            self.textEdit.clear()
            return False

    def is_ipv4_address(self, address):
        # 判断字符串是否符合IPV4的格式
        pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
        if re.match(pattern, address):
            parts = address.split('.')
            if all(0 <= int(part) <= 255 for part in parts):
                return True
        return False

    def open_log_file(self):
        dir_name = os.path.join(BASE_PATH, "Log")
        os.startfile(dir_name)

    def open_report_file(self):
        dir_name = os.path.join(BASE_PATH, "Report")
        os.startfile(dir_name)

    def check_online(self, ip: str,username = 'admin', password = 'abcd1234'):
        onlineFlag = False
        try:
            start_time = time.time()
            # 检查设备是否在线
            message = ping(ip)
            if "Reply" in str(message):
                onlineFlag = True

        # # 基于self.isapi.get看是否能接通设备，以此来判断设备是否在线
        # try:
        #     Session = requests.Session()
        #     Session.auth = HTTPDigestAuth(username, password)
        #     url = f'http://{ip}/ISAPI/System/deviceInfo'
        #     for i in range(3):
        #         print('ISAPI [GET %s]' % (url))
        #         devInfo = Session.get(url)
        #         devInfo = devInfo.text
        #         if "notSupport" in devInfo or "invalidOperation" in devInfo or "404 -- Not Found" in devInfo:
        #             continue
        #
        #         if "</DeviceInfo>" not in devInfo:
        #             continue
        #
        #         onlineFlag = True
        #         break
        except Exception as e:
            print(traceback.print_exc())
        finally:
            end_time = time.time()
            time_interval = end_time - start_time
            print(time_interval)
            self.textEdit.append(f"[{ip} 网络检查时间]：{time_interval} s!")
            return onlineFlag

    # def pingNetwork(self, ip):
    #     # 创建网络检查线程
    #     pingThread = pingDevice(ip)
    #     # 将网络线程添加进线程列表中
    #     self.pingThreadTaskList.append(pingThread)
    #     # 将网络检查线程【信号】函数与play【槽】函数链接
    #     pingThread.ping_result.connect(self.play)
    #     # 将网络检查线程finished【信号】函数与remove_pingThread_task【槽】函数链接
    #     pingThread.finished.connect(self.remove_pingThread_task)
    #     # 启动运行登录线程
    #     pingThread.start()
    #
    #     # 等待线程结束
    #     pingThread.join()

        # # 获取线程返回值
        # if pingThread.onlineFlag:
        #     return True
        # else:
        #     return False


    def get_ip_list(self):
        # 获取IP,设备类型，测试文件
        ip_list = []
        devices_list = []
        excel_file_list = []
        productParam_list = []

        ip_1 = self.lineEdit_1.text()
        devices_1 = self.comboBox_1.currentText()
        excel_path_1 = self.pushButton_excel_1.text()
        paramOpt_1 = self.comboBox.currentText()
        if ip_1 != "" and not self.is_ipv4_address(ip_1):
            self.textEdit.append("设备1的IP非IPV4格式，请确认是否输入错误")
            return False, ip_list, devices_list, excel_file_list
        if ip_1 != "" and '/' not in excel_path_1:
            self.textEdit.append("已填入IP，请选择设备1的对应excel测试文件")
            return False, ip_list, devices_list, excel_file_list
        if ip_1 == "" and '/' in excel_path_1:
            self.textEdit.append("已选择文件，请填入设备1的IP")
            return False, ip_list, devices_list, excel_file_list
        if ip_1 != "":
            ip_list.append(ip_1)
            devices_list.append(devices_1)
            excel_file_list.append(excel_path_1)
            productParam_list.append(paramOpt_1)

        ip_2 = self.lineEdit_2.text()
        devices_2 = self.comboBox_2.currentText()
        excel_path_2 = self.pushButton_excel_2.text()
        paramOpt_2 = self.comboBox_6.currentText()
        if ip_2 != "" and not self.is_ipv4_address(ip_2):
            self.textEdit.append("设备2的IP非IPV4格式，请确认是否输入错误")
            return False, ip_list, devices_list, excel_file_list
        if ip_2 != "" and '/' not in excel_path_2:
            self.textEdit.append("已填入IP，请选择设备2的对应excel测试文件")
            return False, ip_list, devices_list, excel_file_list
        if ip_2 == "" and '/' in excel_path_2:
            self.textEdit.append("已选择文件，请填入设备2的IP")
            return False, ip_list, devices_list, excel_file_list
        if ip_2 != "":
            ip_list.append(ip_2)
            devices_list.append(devices_2)
            excel_file_list.append(excel_path_2)
            productParam_list.append(paramOpt_2)

        ip_3 = self.lineEdit_3.text()
        devices_3 = self.comboBox_3.currentText()
        excel_path_3 = self.pushButton_excel_3.text()
        paramOpt_3 = self.comboBox_7.currentText()
        if ip_3 != "" and not self.is_ipv4_address(ip_3):
            self.textEdit.append("设备3的IP非IPV4格式，请确认是否输入错误")
            return False, ip_list, devices_list, excel_file_list
        if ip_3 != "" and '/' not in excel_path_3:
            self.textEdit.append("已填入IP，请选择设备3的对应excel测试文件")
            return False, ip_list, devices_list, excel_file_list
        if ip_3 == "" and '/' in excel_path_3:
            self.textEdit.append("已选择文件，请填入设备3的IP")
            return False, ip_list, devices_list, excel_file_list
        if ip_3 != "":
            ip_list.append(ip_3)
            devices_list.append(devices_3)
            excel_file_list.append(excel_path_3)
            productParam_list.append(paramOpt_3)

        ip_4 = self.lineEdit_4.text()
        devices_4 = self.comboBox_4.currentText()
        excel_path_4 = self.pushButton_excel_4.text()
        paramOpt_4 = self.comboBox_8.currentText()
        if ip_4 != "" and not self.is_ipv4_address(ip_4):
            self.textEdit.append("设备4的IP非IPV4格式，请确认是否输入错误")
            return False, ip_list, devices_list, excel_file_list
        if ip_4 != "" and '/' not in excel_path_4:
            self.textEdit.append("已填入IP，请选择设备4的对应excel测试文件")
            return False, ip_list, devices_list, excel_file_list
        if ip_4 == "" and '/' in excel_path_4:
            self.textEdit.append("已选择文件，请填入设备4的IP")
            return False, ip_list, devices_list, excel_file_list
        if ip_4 != "":
            ip_list.append(ip_4)
            devices_list.append(devices_4)
            excel_file_list.append(excel_path_4)
            productParam_list.append(paramOpt_4)

        ip_5 = self.lineEdit_5.text()
        devices_5 = self.comboBox_5.currentText()
        excel_path_5 = self.pushButton_excel_5.text()
        paramOpt_5 = self.comboBox_9.currentText()
        if ip_5 != "" and not self.is_ipv4_address(ip_5):
            self.textEdit.append("设备5的IP非IPV4格式，请确认是否输入错误")
            return False, ip_list, devices_list, excel_file_list
        if ip_5 != "" and '/' not in excel_path_5:
            self.textEdit.append("已填入IP，请选择设备5的对应excel测试文件")
            return False, ip_list, devices_list, excel_file_list
        if ip_5 == "" and '/' in excel_path_5:
            self.textEdit.append("已选择文件，请填入设备5的IP")
            return False, ip_list, devices_list, excel_file_list
        if ip_5 != "":
            ip_list.append(ip_5)
            devices_list.append(devices_5)
            excel_file_list.append(excel_path_5)
            productParam_list.append(paramOpt_5)

        return True, ip_list, devices_list, excel_file_list, productParam_list

    def Stop(self):
        self.textEdit.append("即将停止所有设备的测试")
        # k对应ip，v是ip所对应的线程
        for k in list(self.thread_dict.keys()):
            self.thread_dict[k].stop()
            self.thread_dict[k].quit()
            self.thread_dict[k].wait()
            self.thread_dict.pop(k)

        self.run_start_test.setEnabled(True)
        self.run_stop_test.setEnabled(False)

    def remove_thread_task(self):
        """
            当线程执行结束后移除线程任务对象操作
        """
        sender = self.sender()
        for k in list(self.thread_dict.keys()):
            if sender == self.thread_dict[k]:
                try:
                    MtdComFun.VD_LOG_DEBUG(self.thread_dict[k].ip, f'设备{k}测试已完成', level="INFO")
                    self.thread_dict.pop(k)

                except Exception as e:
                    # my_print(f"线程终止时将线程对象移出线程列表出现异常:{str(e)}!", level="ERROR")
                    # print(f"线程终止时将线程对象移出线程列表出现异常:{str(e)}!")
                    MtdComFun.VD_LOG_DEBUG('', f"线程终止时将线程对象移出线程列表出现异常:{str(e)}!", level="ERROR")
        if self.thread_dict == {}:
            self.run_start_test.setEnabled(True)
            self.run_stop_test.setEnabled(False)

    # self.trans_result.emit(True, f"设备{self.ip}测试{self.testnum}全部成功", self.ip, self.testnum)
    # def remove_pingThread_task(self):
    #     """
    #         当线程执行结束后移除线程任务对象操作
    #     """
    #     sender = self.sender()
    #     if sender in self.pingThreadTaskList:
    #         try:
    #             self.pingThreadTaskList.remove(sender)
    #         except Exception as e:
    #             self.play(2, f"网络检查线程终止时将线程对象移出线程列表出现异常:{str(e)}!", '')

    def play(self, ok_Index, msg, ip):
        if ok_Index == 1:
            self.textEdit.append(f"设备{ip} spec测试失败，失败原因:{msg}")
        elif ok_Index == 0:
            self.textEdit.append(f"设备{ip} spec测试结束")
        elif ok_Index == 2:
            self.textEdit.append(msg)

    def Run(self):
        # 清除打印
        self.textEdit.clear()
        # 初始化按钮的使能状态
        self.run_start_test.setEnabled(False)
        self.run_stop_test.setEnabled(True)
        self.thread_dict = {}
        self.textEdit.append("工具开始执行,请稍等片刻......")
        QCoreApplication.processEvents()

        try:
            ip_flag, self.ip_list, self.devices_list, self.excel_file_list, self.paramOpt = self.get_ip_list()
            if not ip_flag:
                return

            # 获取工具界面的账号和密码
            username = self.usernameEdit.text()
            password = self.passwordEdit.text()
            if username == "":
                self.textEdit.append("账号不能为空，请勿删除默认账号，或填入新账号")
                return
            if password == "":
                self.textEdit.append("密码不能为空，请勿删除默认密码，或填入新密码")
                return

            not_online_ip = []
            for ip in self.ip_list:
                flag = self.check_online(ip, username, password)
                if not flag:
                    not_online_ip.append(ip)
            if not_online_ip != []:
                self.textEdit.append("IP为" + str(not_online_ip) + "的设备不在线，请检查设备或网络")
                return

            self.textEdit.append("设备列表：" + str(self.ip_list) + "设备全部在线，即将开始测试")

            self.textEdit.append(f"即将开始测试,当前设备数量为{len(self.ip_list)}台，请确认所有设备账号密码均为默认账号密码")

            # 创建一个包含n个线程的线程池,n为实际ip个数
            for i in range(len(self.ip_list)):
                Path = os.path.join(BASE_PATH, "Report")
                self.thread_dict[i] = spec_test(self.ip_list[i], username, password, Path, self.devices_list[i],
                                                self.excel_file_list[i].replace("/", "\\"), productOpt=self.paramOpt[i])
                self.textEdit.append(f"设备{self.ip_list[i]} 开始测试")

            for k, v in self.thread_dict.items():
                v.trans_result.connect(self.play)
                v.finished.connect(self.remove_thread_task)
                v.start()

        except Exception as e:
            # """设备测试打印展示"""
            # # 获取被调用函数在被调用时所处代码行数
            # line = sys._getframe().f_back.f_lineno
            # # 获取被调用函数所在模块文件名
            # file_name = sys._getframe(1).f_code.co_filename
            # file_name = file_name.replace("\\", "/").split("/")[-1]
            # typeDict = {0: "INFO", 1: "ERROR"}
            # logMsg = f"[{str(datetime.datetime.now())[:19]} {file_name} {line}] <font color=\"#FF0000\">{traceback.print_exc()}</font>"

            print(traceback.print_exc())
            # my_print(f"测试过程中出现异常:{e}", level="ERROR")
            MtdComFun.VD_LOG_DEBUG('', f"测试过程中出现异常:{e}", level="ERROR")
            self.textEdit.append(f"测试过程中出现异常:{e}")
            return

    # 版本校验
    def checkCattVersion(self):
        """
        检查软件版本
        :return:
        """
        path = os.getcwd()
        parser = Dispatch("Scripting.FileSystemObject")
        current_version = parser.GetFileVersion(path + '\\SPEC.exe')
        current_version_list = current_version.split('.')
        # print current_version

        flag, newest_version = service_cfg('SPEC_VERSION')
        if not flag:
            return False, newest_version

        newest_version_list = newest_version.split('.')
        # print newest_version

        ##对照年月日
        for i in range(4):
            if current_version_list[i] == newest_version_list[i]:
                continue
            else:
                return False, f"当前SPEC版本为{current_version},最新版本为{newest_version},请使用最新版SPEC！"
        return True, "当前SPEC已是最新版本，请开始运行！"
class My_helpWindow(QMainWindow, Ui_HelpWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 禁止最大化
        self.setFixedSize(self.width(), self.height())

def main():
    app = QtWidgets.QApplication(sys.argv)
    # 修改app样式，白色样式
    app.setStyleSheet(qdarkstyle.load_stylesheet(qt_api='pyqt5', palette=LightPalette()))
    # 修改app样式，黑色样式
    # app.setStyleSheet(qdarkstyle.load_stylesheet(qt_api='pyqt5'))
    window = CameraTest()
    window.setWindowIcon(QIcon(get_path("SpecTemplate/app.jpg")))
    # window.setWindowIcon(QIcon("SpecTemplate/app.jpg"))
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")

    # 设置自定义的异常处理函数
    sys.excepthook = excepthook

    # 版本校验弹窗
    flag, msg = window.checkCattVersion()
    if not flag:
        QMessageBox.warning(window, "Warning", msg, QMessageBox.Ok)
        window.close()
        return

    window.show()
    # 弹出使用说明文档
    mh = My_helpWindow()
    mh.show()

    sys.exit(app.exec_())

if __name__ == '__main__':
    # 如果不存在日志和抓图目录则先创建
    LogPath = os.path.join(BASE_PATH, "Log")
    if not os.path.exists(LogPath):
        os.makedirs(LogPath)
    ReportPath = os.path.join(BASE_PATH, "Report")
    if not os.path.exists(ReportPath):
        os.makedirs(ReportPath)
    current_time = time.strftime('%Y-%m-%d_%H_%M_%S', time.localtime(time.time()))
    # 执行日志文件
    testLogFile = os.path.join(LogPath, f"testLog_{current_time}.txt")

    multiprocessing.freeze_support()
    main()
    sys.exit(0)












