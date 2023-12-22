# -*- coding: utf-8 -*-
"""
@Auth ： Fanghe.Lin
@File ：mtd_com_fun.py
@IDE ：PyCharm
@Time ： 2023/10/10 10:49
@Institution: Hikvision, China
"""

# encoding: utf-8
'''
@author: qianjin
@license: (C) Copyright 2013-2019, Node Supply Chain Manager Corporation Limited.
@contact: deamoncao100@gmail.com
@software: garner
@file: mtd_com_fun.py
@time: 2022/11/23 21:06
@desc:常用的公共函数

'''
from openpyxl import load_workbook
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import MergedCell
# from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.cell import MergedCell
from lxml import etree
import os, sys, re
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
import json
import urllib
import paramiko
import time
import ssl
from functools import wraps
from http import cookiejar as cookielib
import traceback
import requests
from lib.Integ_Isapi import Integ_Isapi

PY_RUN_CORRECT = "P"
PY_RUN_NA = "NA"
PY_RUN_WRONG = "F"

time_seq = time.strftime('[%Y-%m-%d-%H-%M-%S]', time.localtime(time.time()))
# 中文设备智能包名
HEOP_APP_ZH = {
    "facesnap": ["人脸抓拍"],
    "faceContrast": ["人脸抓拍", "比对建模"],
    "mixedTargetDetection": ["全结构化"],
    "faceHumanModelingContrast": ["全结构化", "比对建模"],
    "roadDetection": ["道路监控"],
    "smart": ["smart"],
    "intelligentTraffic": ["智能交通"],
    "mixedTargetFaceHuman": ["全结构化"],
    "AIOpenPlatform": ["AI开放平台"],
    "personDensity": ["人员密度"],
    "SafeHelmet": ["安全帽检测"],
    'Dismission/Sleeping': ['岗位值守检测'],
    'KitchenHygieneDetection': ['明厨亮灶'],
    'GarbageDetection': ['垃圾检测'],
    'WorkingClothesDetection': ['工服检测'],
    'IndoorFirePassageDetection': ['室内消防通道占用'],
    'riverwayGaugeReading': ["水尺检测"],
    'smokeDetection': ["烟雾检测"],
    'personArming': ["人员布控"],
    'ShipAndFloatDetection': ["船只和漂浮物检测"],
    'multidimensionalPeopleCounting': ["多维客流"],
    'objectsThrownDetection': ["高空抛物检测"],
    'personDensityQueueLeavePosition': ["人数统计"]
}

# 英文设备智能包名
HEOP_APP_EN = {
    "facesnap": ["Face Capture"],
    "faceContrast": ["Face Capture", "Comparison and Modeling"],
    "mixedTargetDetection": ["Capture Target With Feature"],
    "faceHumanModelingContrast":
        ["Capture Target With Feature", "Comparison and Modeling"],
    "roadDetection": ["Road Traffic"],
    "smart": ["Smart Event"],
    "intelligentTraffic": ["Intell Traffic Event"],
    "mixedTargetFaceHuman": ["Capture Target With Feature"],
    "AIOpenPlatform": ["AI Open Plat"],
    "personDensity": ["basePersonDensity"],
    "SafeHelmet": ["Hard Hat Detection"],
    'Dismission/Sleeping': ['Leave Position Detection'],
    'KitchenHygieneDetection': ['KitchenHygieneDetection'],
    'GarbageDetection': ['Garbage Detection'],
    'WorkingClothesDetection': ['Working Clothes Detection'],
    'IndoorFirePassageDetection': ['Indoor Fire Passage Detection'],
    'smokeDetection': ["SMOG DETECT"],
    'ShipAndFloatDetection': ["Ship and Float Detection"],
    'personArming': ["Person Arming"],
    'multidimensionalPeopleCounting': ["multidimensional pdc"],
    'objectsThrownDetection': ["Thrown Detect Event"],
    'personDensityQueueLeavePosition': ["personCounting"]
}


def sslwrap(func):
    @wraps(func)
    def bar(*args, **kw):
        kw['ssl_version'] = ssl.PROTOCOL_TLSv1
        return func(*args, **kw)

    return bar


class AutoVivification(dict):
    """定义多级字典"""

    def __getitem__(self, item):
        try:
            return dict.__getitem__(self, item)
        except KeyError:
            value = self[item] = type(self)()
            return value


ssl.wrap_socket = sslwrap(ssl.wrap_socket)
ssl._create_default_https_context = ssl._create_unverified_context


class Zhimakaimen:
    """class Zhimakaimen
    """
    m_prefix = 'http://10.1.32.22:8080/'
    m_site = 'http://psh.hikvision.com.cn:8080/sso/page/command?pageName=homeCN'
    m_site_action = 'http://psh.hikvision.com.cn:8080/sso/rsa/encrypt'
    m_host = '10.1.32.22:8080'
    m_origin = m_prefix
    m_refer = 'http://psh.hikvision.com.cn:8080/sso/page/command?pageName=homeCN'
    m_loginAddr = 'https://sso.hikvision.com/login?service=http%3A%2F%2Fpsh.hikvision.com.cn%3A8080%2Fsso%2Fpage%2Fcommand%3FpageName%3Dhome'

    # m_loginAddr='https://sso.hikvision.com/login'

    def __init__(self, userName, password):
        # print (userName, password)
        ssl.wrap_socket = sslwrap(ssl.wrap_socket)
        self.m_cookieJar = cookielib.LWPCookieJar("cookie")
        self.m_opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.m_cookieJar))
        self.m_opener.addheaders = [
            ('User-agent',
             'Mozillaebug():/4.0 (compatible; MSIE 5.5; Windows NT)')
        ]
        urllib.request.install_opener(self.m_opener)
        self.m_userName = userName
        self.m_password = password

    def Login(self):
        """login.

        send request with username, password. create cookie store these for service using
        """
        userInfo = {
            'username': self.m_userName,
            'password': self.m_password,
            'lt': '',
            'execution': '',
            '_eventId': '',
        }
        try:
            respond = self.m_opener.open(self.m_loginAddr).read().decode()
            if re.search(r'<div id="msg" class="success">',
                         respond) is not None:
                return True
        except Exception as why:
            print("error:can't connect to internet" + str(why))
            return False

        userInfo['username'] = self.m_userName
        userInfo['password'] = self.m_password
        userInfo['lt'] = re.findall(
            r'name="lt" value="(?P<str1>[0-9a-zA-Z\- ]+)"', respond)[0]
        userInfo['execution'] = re.findall(
            r'name="execution" value="(?P<str1>[0-9a-zA-Z\- ]+)"', respond)[0]
        userInfo['_eventId'] = re.findall(
            r'name="_eventId" value="(?P<str1>[0-9a-zA-Z\- ]+)"', respond)[0]
        # print(userInfo)
        requestStr = urllib.request.Request(
            self.m_loginAddr,
            urllib.parse.urlencode(userInfo).encode("utf-8"))
        # print requestStr
        try:
            respond = self.m_opener.open(requestStr).read().decode()
            # print respond
            # if re.search(r'<div id="msg" class="success">', respond) is not None:
            if re.search(r'PSH Shell', respond) is not None:
                return True
            else:
                print('login faild, login msg is wrong')
                print(respond)
                return False
        except Exception as why:
            print('error: ' + str(why))

    def GetZhima(self, inStr):
        """get zhimakaimen's secret
        """
        requestMsg = {
            'source': inStr,
            'customMsg': "catt",
            # 'customMsg' : self.m_userName
        }
        try:
            self.m_opener.open(self.m_site).read().decode()
        except Exception as why:
            print('open site url faild: ' + str(why))
            return None

        # print urllib.urlencode(requestMsg)
        request = urllib.request.Request(
            self.m_site_action,
            urllib.parse.urlencode(requestMsg).encode("utf-8"))
        # print request
        request.add_header('X-Requested-With', 'XMLHttpRequest')
        request.add_header(
            'Referer',
            'http://psh.hikvision.com.cn:8080/sso/page/command?pageName=homeCN'
        )
        request.add_header('Accept',
                           "application/json, text/javascript, */*; q=0.01")
        request.add_header('Content-Type',
                           'application/x-www-form-urlencoded; charset=utf-8')
        request.add_header('Accept-Encoding', "gzip, deflate")
        request.add_header(
            'User-agent',
            'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/7.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.3; .NET4.0C; .NET4.0E)'
        )
        request.add_header('Host', 'psh.hikvision.com.cn:8080')
        # request.add_header('Origin', self.m_origin)
        try:
            jsonOriginal = self.m_opener.open(request).read()
        except Exception as why:
            print('open site_action url faild: ' + str(why))
            return None
        jsonObject = json.loads(jsonOriginal)
        reStr = jsonObject['result']
        reg = re.compile(r'[\n\r]+')
        return reg.sub("", reStr)


class MySShConnect():
    def __init__(self, ip, username, password):
        self.ip = ip
        self.username = username
        self.password = password

    def SSH_client(self):
        recv_buffer = 99999

        try:
            ssh = paramiko.SSHClient()  ##1.创建一个ssh对象
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy(
            ))  # 2.解决问题:如果之前没有，连接过的ip，会出现选择yes或者no的操作，自动选择yes
            ssh.connect(hostname=self.ip,
                        port=22,
                        username=self.username,
                        password=self.password,
                        timeout=30,
                        allow_agent=False,
                        look_for_keys=False)  # 3.连接服务器
            connect_result = True
        except Exception as e:
            # 不能正常连接的server打印信息到控制台
            print(str(e))
            connect_result = False
        return ssh, recv_buffer, connect_result

    def psh_koulin(self, ssh):
        kaimenCmds = ["debug\n", "zhimakaimen\n"]
        koulin = ''
        for cmd in kaimenCmds:
            ssh.send(cmd)
            koulin = ssh.recv(115200).decode("utf-8", "ignore")
            while (not koulin.endswith(': ')) and (
                    not koulin.endswith('# ')) and (len(koulin) < 1024 * 10):
                koulin = koulin + ssh.recv(115200).decode("utf-8", "ignore")
            if (len(koulin) != 0) and ("help" not in koulin):
                break
        return str(koulin)

    def send_cmd(self, ssh, cmd, limit=1024, rate=10, packageNums=100):
        recv_buffer = 65536
        recv_buffer_limit = limit * rate
        try:
            package = 1
            ssh.send(cmd + '\n')
            out = ssh.recv(recv_buffer).decode("utf-8", "ignore")  # recv_buffer
            time.sleep(1)
            while (not out.endswith('# ')) and (
                    len(out) < recv_buffer_limit) and (package < packageNums):
                out = out + ssh.recv(recv_buffer).decode("utf-8", "ignore")
                package += 1
        except Exception as why:
            info = 'Failed to execute SSH command %s: %s .' % (cmd, repr(why))
            return info

        string = str(out)
        return string

    def send_cmd_4G(self, ssh, cmd, limit=1024, rate=10, packageNums=100):
        recv_buffer = 65536
        # recv_buffer_limit = limit * rate
        try:
            # package = 1
            ssh.send(cmd + '\r\n')
            time.sleep(2)
            out = ssh.recv(recv_buffer).decode("utf-8", "ignore")  # recv_buffer
        except Exception as why:
            info = 'Failed to execute SSH command %s: %s .' % (cmd, repr(why))
            return info

        string = str(out)
        return string

    def psh_change_to_ash(self, ssh):
        koulin = self.psh_koulin(ssh)
        print(koulin)
        zhimakaimen = Zhimakaimen("pubyfcs055", "DFp03yg7]")
        if zhimakaimen.Login():
            if len(koulin) > 1000:
                if "enc_string:" in koulin:
                    final_koulin = koulin.split("\n")[-2][11:]
                else:
                    final_koulin = koulin.split("\n")[-2]
            else:
                final_koulin = koulin.split("\n")[1]
            reStr = zhimakaimen.GetZhima(final_koulin)
            if reStr is not None:
                try:
                    reStr = reStr.encode('utf-8').decode("utf-8", "ignore")
                    if reStr is not None:
                        self.send_cmd(ssh, reStr)
                        time.sleep(1)
                        return True, "进入debug模式成功！"
                except Exception as why:
                    return False, repr(why)
        else:
            print('error', "Login failed, please check <config.ini>.")
            return False, "Login failed, please check <config.ini>."

    def zhimakaimeng_debug(self):
        try:
            client, recv_buffer, result = self.SSH_client()
            if not result:
                return False, ("SSH连接失败，请确认是否真的支持SSH或功能是否正常!", "")
            ssh = client.invoke_shell()
            ssh.settimeout(60)
            out = ssh.recv(recv_buffer).decode("utf-8", "ignore")
            while (not out.endswith('# ')):
                out = out + ssh.recv(1024).decode("utf-8", "ignore")
            # print(out)
            if 'psh' in out:
                koulin = self.psh_koulin(ssh)
                print(koulin)
                zhimakaimen = Zhimakaimen("pubyfcs055", "DFp03yg7]")
                if zhimakaimen.Login():
                    if len(koulin) > 1000:
                        if "enc_string:" in koulin:
                            final_koulin = koulin.split("\n")[-2][11:]
                        else:
                            final_koulin = koulin.split("\n")[-2]
                    else:
                        final_koulin = koulin.split("\n")[1]
                    reStr = zhimakaimen.GetZhima(final_koulin)
                    if reStr is not None:
                        try:
                            reStr = reStr.encode('ascii').decode(
                                "utf-8", "ignore")
                            if reStr is not None:
                                self.send_cmd(ssh, reStr)
                                time.sleep(1)
                                return True, (ssh, client)
                            # instring = raw_input("Pass anykey to exit!!!")
                        except Exception as why:
                            return False, (str(why), "")
                else:
                    print('error', "Login failed, please check <config.ini>.")
                    # if os.path.exists('%s\config.ini'%path):
                    #     os.remove('%s\config.ini'%path)
                    return False, ("Login failed, please check <config.ini>.",
                                   "")

            elif 'ash' in out:
                return True, (ssh, client)
        except Exception as why:
            return False, (str(why), "")


class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        self.merged_cell_ranges=self.ws.merged_cells.ranges  # 表示获取合并单元格范围列表

    def set_merged_cell(self, start_row, start_column, end_row, end_column):
        self.ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

    def parser_merged_cell(self, merged_cell_ranges, row, col):
        """
        检查是否为合并单元格并获取对应行列单元格的值。
        如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
        :param sheet: 当前工作表对象
        :param row: 需要获取的单元格所在行
        :param col: 需要获取的单元格所在列
        :return:
        """
        cell = self.ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
            for merged_range in merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                if cell.coordinate in merged_range:
                    # 获取合并区域左上角的单元格作为该单元格的值返回
                    cell = self.ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        return cell

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.parser_merged_cell(self.merged_cell_ranges, row, column)

        return cell_value.value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.parser_merged_cell(self.merged_cell_ranges, i, column)
            # print("第%s行第%s列：%s" % (index_row, column, cell_.value))
            column_data.append(cell_value.value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.parser_merged_cell(self.merged_cell_ranges, row, i)
            # print("第%s行第%s列：%s" % (index_row, column, cell_.value))
            row_data.append(cell_value.value)
        return row_data

    def openxls_create(self):
        '''
        创建xlsx表格，并且写入数据

        颜色
        COLOR_INDEX = (
            '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
            '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
            '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
            '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
            '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
            '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
            '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
            '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
            '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
            '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
            '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
            '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
            '00993300', '00993366', '00333399', '00333333',  #60-63
        )
        BLACK = COLOR_INDEX[0]
        WHITE = COLOR_INDEX[1]
        RED = COLOR_INDEX[2]
        DARKRED = COLOR_INDEX[8]
        BLUE = COLOR_INDEX[4]
        DARKBLUE = COLOR_INDEX[12]
        GREEN = COLOR_INDEX[3]
        DARKGREEN = COLOR_INDEX[9]
        YELLOW = COLOR_INDEX[5]
        DARKYELLOW = COLOR_INDEX[19]

        字体
        ws.cell(5,3).value='哈哈哈'
        ws.cell(5,3).font = Font(name='仿宋',size=12,color=Color(index=0),b=True,i=True)

        # size   sz  字体大小
        # b bold  是否粗体
        # i italic  是否斜体
        # name family  字体样式

        # 边框
        Side(style='thin',color=Color(index=0))

        # style可选项
        style = ('dashDot','dashDotDot', 'dashed','dotted',
        'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
        'mediumDashed', 'slantDashDot', 'thick', 'thin')
        #  'medium' 中粗
        #  'thin'  细
        #  'thick'  粗
        #  'dashed'  虚线
        #  'dotted'  点线

        # 填充
        PatternFill(patternType='solid',fgColor=Color(), bgColor=Color())
        # fgColor   前景色
        # bgColor   后景色
        # 参数可选项
        patternType = {'darkDown', 'darkUp', 'lightDown', 'darkGrid', 'lightVertical',
                       'solid', 'gray0625', 'darkHorizontal', 'lightGrid', 'lightTrellis',
                       'mediumGray', 'gray125', 'darkGray', 'lightGray', 'lightUp',
                       'lightHorizontal', 'darkTrellis', 'darkVertical'}ws.cell(3,3).fill = PatternFill()

        # 对齐
        Alignment(horizontal='fill',vertical='center')

        # 参数可选项
        horizontal = {'fill', 'distributed', 'centerContinuous', 'right',
                      'justify', 'center', 'left', 'general'}

        vertical = {'distributed', 'justify', 'center', 'bottom', 'top'}

        ws.cell(3,3).alignment= Alignment()

        # 行高，列宽
        row = ws.row_dimensions[1]
        row.height = 15
        col = ws.column_dimensions[1]
        col.width = 10
        '''

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            # cell = self.ws.cell(row=row, column=colunm)
            # if isinstance(cell, MergedCell):
            #     cell.value=cellvalue
            #
            # else:
            # self.ws.merged_cell_ranges
            self.ws.cell(row=row, column=colunm).alignment = styles.Alignment(horizontal='center', wrapText=True,
                                                                              vertical='center')
            self.ws.cell(row=row, column=colunm).font = styles.Font(name='仿宋', size=12, color=styles.Color(index=0),
                                                                    b=True)
            # self.ws.cell(row=row, column=colunm).border=styles.Border(left=styles.Side(border_style='thin', color='000000'),
            #     right=styles.Side(border_style='mediumDashed', color='000000'),
            #     top=styles.Side(border_style='double', color='000000'),
            #     bottom=styles.Side(border_style='dashed', color='000000'))
            self.ws.cell(row=row, column=colunm).border = styles.Border(
                left=styles.Side(border_style='thin', color='000000'),
                right=styles.Side(border_style='thin', color='000000'),
                top=styles.Side(border_style='thin', color='000000'),
                bottom=styles.Side(border_style='thin', color='000000'))
            self.ws.cell(row=row, column=colunm).value = cellvalue
            # self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).alignment = styles.Alignment(wrapText=True, vertical='center')
            self.ws.cell(row=row, column=colunm).font = styles.Font(name='仿宋', size=12, color=styles.Color(index=0),
                                                                    b=True)
            self.ws.cell(row=row, column=colunm).border = styles.Border(
                left=styles.Side(border_style='thin', color='000000'),
                right=styles.Side(border_style='thin', color='000000'),
                top=styles.Side(border_style='thin', color='000000'),
                bottom=styles.Side(border_style='thin', color='000000'))
            self.ws.cell(row=row, column=colunm).value = "writefail"
            # self.wb.save(self.file)

    def saveExc(self):
        self.wb.save(self.file)
        self.wb.close()


def getDeviceInfo(integ_isapi):
    '''获取设备基本信息'''
    try:
        dev_info = {}
        page = integ_isapi.get(f'/ISAPI/System/Video/inputs/channels')
        root = etree.XML(str(page).encode('utf-8'))
        if '</name>' in str(page):
            channel_name = root.xpath("//*[local-name()='name']")
            dev_info["chalNum"] = len(channel_name)
            # dev_info[f'channel1_name']= channel_name
            # dev_info[f'channel1_name'] = channel_name
            # dev_info[f'channel{channel}_name'] = channel_name

        page = integ_isapi.get(f'/ISAPI/System/deviceInfo')
        root = etree.XML(str(page).encode('utf-8'))

        serialNumber = root.xpath("//*[local-name()='serialNumber']")[0].text if root.xpath(
            "//*[local-name()='serialNumber']") else ""
        # dev_info['serialNumber']=serialNumber
        dev_info.update({"serialNumber": serialNumber})
        if '</model>' in str(page):
            dev_model = root.xpath("//*[local-name()='model']")[0].text
            if not dev_model:
                dev_model = ''
            # dev_info['dev_model']=self.dev_model
            dev_info['dev_model'] = dev_model

        if '</deviceName>' in str(page):
            device_Name = root.xpath("//*[local-name()='deviceName']")[0].text
            if not device_Name:
                device_Name = ''
            # dev_info['deviceName']=self.dev_model
            dev_info['deviceName'] = device_Name
        if '</deviceDescription>' in str(page):
            deviceDescription = root.xpath("//*[local-name()='deviceDescription']")[0].text
            if 'IPD' in deviceDescription:
                # dev_info['dev_type'] = 'IPD'
                dev_type = 'IPD'

            elif 'IPC' in deviceDescription:
                # dev_info['dev_type'] = 'IPC'
                dev_type = 'IPC'
            elif 'IPZ' in deviceDescription:
                # dev_info['dev_type'] = 'IPZ'
                dev_type = 'IPZ'
            else:
                # dev_info['dev_type'] = deviceDescription
                dev_type = deviceDescription
            dev_info['dev_type'] = dev_type
            dev_info['deviceDescription'] = deviceDescription
        elif '</deviceType>' in str(page):
            dev_type = root.xpath("//*[local-name()='deviceType']")[0].text
            # dev_info['dev_type']=self.dev_type
            dev_info['dev_type'] = dev_type
            dev_info['deviceType'] = dev_type

        if '</firmwareVersionInfo>' in str(page):
            firmwareVersionInfo = root.xpath("//*[local-name()='firmwareVersionInfo']")[0].text
            # print(firmwareVersionInfo)
            # print(firmwareVersionInfo.upper().split('-'))
            dev_platform = firmwareVersionInfo.upper().split('-')[2]
            # dev_info['dev_platform']=self.dev_platform
            dev_info['dev_platform'] = dev_platform
            dev_info.update({"firmwareVersionInfo": firmwareVersionInfo})
            # dev_info['firmwareVersionInfo']=firmwareVersionInfo
        else:
            dev_info['dev_platform'] = ''
        if '</firmwareVersion>' in str(page):
            firmwareVersion = root.xpath("//*[local-name()='firmwareVersion']")[0].text
            # dev_info['firmwareVersion'] = firmwareVersion
            dev_info['firmwareVersion'] = firmwareVersion
        else:
            dev_info['firmwareVersion'] = ''
        if '</firmwareReleasedDate>' in str(page):
            firmwareReleasedDate = root.xpath("//*[local-name()='firmwareReleasedDate']")[0].text
            dev_info['firmwareReleasedDate'] = firmwareReleasedDate
            # dev_info['firmwareReleasedDate'] = firmwareReleasedDate
        else:
            dev_info['firmwareReleasedDate'] = ''
        if dev_info['firmwareReleasedDate'] != '' and dev_info['firmwareVersion'] != '':
            dev_info['dev_softwareVersion'] = f"{dev_info['firmwareVersion']} {dev_info['firmwareReleasedDate']}"
            # dev_info['softwareVersion'] = dev_info['softwareVersion']
        # print(dev_info)
        return True, dev_info
    except Exception as e:
        return False, str(e)


def VD_LOG(*args, sep=' ', end='\n', file=None):
    '''公共打印log'''
    init_path = os.getcwd()
    # 获取被调用函数在被调用时所处代码行数
    line = sys._getframe().f_back.f_lineno
    # 获取被调用函数所在模块文件名
    file_name = sys._getframe(1).f_code.co_filename
    file_name = file_name.replace("\\", "/").split("/")[-1]
    # sys.stdout.write(f'"{__file__}:{sys._getframe().f_lineno}"  {x}\n')
    args = (str(arg) for arg in args)  # REMIND 防止是数字不能被join
    args_str = " ".join(args)
    # 打印到标准输出，并设置文字和背景颜色
    # sys.stdout.write(f'"{file_name}:{line}" {time.strftime("%H:%M:%S")} \033[0;94m{"".join(args)}\033[0m\n') # 36 93 96 94

    # time_seq = time.strftime('[%Y-%m-%d %H:%M:%S]', time.localtime(time.time())) time.strftime("%Y%m%d%H")
    # init_path=os.path.abspath(os.path.dirname(__file__))
    # init_path=os.getcwd()
    log_path_name = "%s\\log\\log_%s_device.log" % (init_path, time_seq)
    info = f'[{file_name}:{line}]{" " + args_str}\n'
    info = info.replace("\r\n", "\n")
    info_log = time_seq + info
    info_log = info_log.encode("gbk", errors="ignore").decode("gbk")

    with open(log_path_name, "a+") as f:
        f.write(info_log)
        sys.stdout.write(info)
        f.close()


def VD_LOG_DEBUG(ip, *args, sep=' ', end='\n', file=None):
    '''公共打印log'''
    init_path = os.getcwd()
    # 获取被调用函数在被调用时所处代码行数
    line = sys._getframe().f_back.f_lineno
    # 获取被调用函数所在模块文件名
    file_name = sys._getframe(1).f_code.co_filename
    file_name = file_name.replace("\\", "/").split("/")[-1]
    # sys.stdout.write(f'"{__file__}:{sys._getframe().f_lineno}"  {x}\n')
    args = (str(arg) for arg in args)  # REMIND 防止是数字不能被join
    args_str = " ".join(args)
    # 打印到标准输出，并设置文字和背景颜色 sys.stdout.write(f'"{file_name}:{line}" {time.strftime("%H:%M:%S")} \033[0;94m{"".join(
    # args)}\033[0m\n') # 36 93 96 94

    # time_seq = time.strftime('[%Y-%m-%d %H:%M:%S]', time.localtime(time.time()))
    # init_path=os.path.abspath(os.path.dirname(__file__))
    # init_path=os.getcwd()
    try:
        os.makedirs(f"log\\{ip}")
    except FileExistsError:
        pass
    log_path_name = f"%s\\log\\{ip}\\log_%s_%s.txt" % (init_path, ip, time_seq)
    info = f'[{file_name}:{line}]{" " + args_str}\n'
    info = info.replace("\r\n", "\n")
    info_log = time_seq + info
    info_log = info_log.encode("gbk", errors="ignore").decode("gbk")

    with open(log_path_name, "a+") as f:
        f.write(info_log)
        sys.stdout.write(info)
        f.close()

VD_LOG_DEBUG("10.20.100.12","123123231q2w4e")

def getMsgResult_Test(dics):
    '''格式化信息和结果用于判断执行用例是否成功'''
    msg = ''
    result = PY_RUN_CORRECT
    list_dic = []
    islistFlag = 0
    for key, value in dics.items():
        list_dic.append(value["result"])
        if isinstance(value["msg"], list):
            islistFlag = 1
    # list_dic=[value for value in dics.values()]
    if not islistFlag:
        print(list_dic)
        if len(set(list_dic)) == 1:
            if list_dic[0] == "PY_RUN_CORRECT":

                msg = ";".join([value["msg"] for key, value in dics.items()])
                result = PY_RUN_CORRECT
                return msg, result
            elif list_dic[0] == 'PY_RUN_NA':
                msg = "设备所有通道功能不支持"
                result = PY_RUN_NA
                return msg, result
        for key, value in dics.items():
            msg += value["msg"] + ";"
            if value['result'] == 'PY_RUN_WRONG':
                result = PY_RUN_WRONG
    return msg, result


def get_stream_supportChannelNum(a_list):
    '''获取码流ID'''
    one_list = 0
    two_list = 0
    three_list = 0
    four_list = 0
    five_list = 0
    # a_list = ["101", "102", "103", "104", "105", "201", "202", "203", "204"]
    for item in a_list:
        if "1" in item.split("0")[1]:
            one_list += 1
        if "2" in item.split("0")[1]:
            two_list += 1
        if "3" in item.split("0")[1]:
            three_list += 1
        if "4" in item.split("0")[1]:
            four_list += 1
        if "5" in item.split("0")[1]:
            five_list += 1
    return {
        "1": one_list,
        "2": two_list,
        "3": three_list,
        "4": four_list,
        "5": five_list
    }


def getReceiveDicParas(evType, relationResuleList):
    """获取关联智能最多的组合
        找到每组中最多的指定智能的组合
    """
    a = {}
    for index, item in enumerate(relationResuleList):
        c = 0
        for b in item.values():
            if evType in b:
                c += 1
        a[index] = c
    d = [int(i) for i in a.values()]
    maxInTypeValue = sorted(d)[-1]
    # print(sorted(d)[-1])
    for c, db in a.items():
        if int(db) == maxInTypeValue:
            return True, relationResuleList[int(c)]
        else:
            pass
    return False, "获取存在最多指定智能关联组合失败"


def device_is_overseas_judge(integ_isapi):
    '''设备是否为海外设备判断'''
    try:
        languageUrl = '/SDK/language'
        ret = integ_isapi.isapi_get(languageUrl, path=[".//type"], type=["value"])
        if not ret['success']:
            return False, f"判断设备是否为海外设备协议执行失败:{ret['msg']}!"

        language = ret['data'][0][0]

        if language not in ['english', 'English']:
            return True, False

        return True, True
    except Exception as e:
        print(traceback.print_exc())
        return False, f"判断设备是否为海外设备出现异常:{str(e)}!"


def channel_VCAResourse_OpenHEOP_cap_judge(integ_isapi, channel=1, vcaType='facesnap'):
    """
        开放HEOP智能资源能力判断接口

        参数:
            channel:通道号,从1开始
            vcaType:智能资源名称
        返回:
            return 接口执行成功，[是否是开放HEOP架构，是否支持该智能，智能对应的id号]
            return 接口执行失败，详细报错信息

            成功:
                return True, [True, True, 1]
            失败
                return False, “详细失败报错信息”
    """
    try:
        # 开放HEOP能力判断
        openHeopCapUrl = "/ISAPI/Custom/OpenPlatform/capabilities"
        openHeopCap = integ_isapi.get(openHeopCapUrl)
        print(openHeopCap)

        if "notSupport" in openHeopCap or "invalidOperation" in openHeopCap:
            return True, [False, False, 0]

        if "404 -- Not Found" in openHeopCap:
            return True, [False, False, 0]

        if "</OpenPlatformCap>" not in openHeopCap:
            return False, f"通道/ISAPI/Custom/OpenPlatform/capabilities接口判断设备是否为开放HEOP平台设备协议执行失败:{integ_isapi.get_subStatusCode(openHeopCap)}!"

        if "</isSupportOpenPlatform>" not in openHeopCap:
            return True, [False, False, 0]

        isSupportOpenPlatform = re.findall(
            "<isSupportOpenPlatform>(.*?)</isSupportOpenPlatform>",
            openHeopCap, re.S)[0]

        if isSupportOpenPlatform == "false":
            return True, [False, False, 0]

        # 人脸比对和混合目标比对暂不测试
        if vcaType == "faceContrast" or vcaType == "faceHumanModelingContrast":
            return True, [True, False, 0]

        # 判断是否是多通道的开放HEOP架构，目前暂不测试多通道的设备
        if "</AppChannels>" in openHeopCap:
            return True, [True, False, 0]

        # 判断设备是否为海外设备
        isOverseasFlag, isOverseasMsg = device_is_overseas_judge(integ_isapi)
        if not isOverseasFlag:
            return False, isOverseasMsg

        # 普通监控在开放HEIOP架构中不存在
        if vcaType == "close":
            return True, [True, False, 0]

        nameList = HEOP_APP_EN[
            vcaType] if isOverseasMsg else HEOP_APP_ZH[vcaType]

        # 获取设备当前已加载所有智能资源
        openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
        openHeopData = integ_isapi.get(openHeopUrl)
        print(openHeopData)
        if "</AppList>" not in openHeopData:
            return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{integ_isapi.get_subStatusCode(openHeopData)}!"

        if "</App>" not in openHeopData:
            return True, [True, False, 0]

        idList = re.findall("<id>(.*?)</id>", openHeopData, re.S)
        packageNameList = re.findall("<packageName>(.*?)</packageName>",
                                     openHeopData, re.S)

        print(f"packageNameList={packageNameList}")

        if len(idList) == 0 or len(packageNameList) == 0:
            return False, "通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表失败!"

        for i in nameList:
            if i not in packageNameList:
                return True, [True, False, 0]

        packageName = nameList[0]
        packageName_id = {}
        for i in range(len(idList)):
            packageName_id[packageNameList[i]] = idList[i]

        print(packageName_id)

        index = packageName_id[packageName]

        return True, [True, True, index]

    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备通道{channel}判断智能资源能力出现异常:{str(e)}!"


def channel_VCAResourse_cap_judge(integ_isapi, channel=1, vcaType='facesnap'):
    """
        智能资源能力判断接口

        参数:
            channel:通道号,从1开始
            vcaType:智能资源名称
        返回:
            return 接口执行成功，[是否有智能资源概念(针对smart事件)，是否支持该智能，是否是关联智能关系, 是否是开放HEOP架构]
            return 接口执行失败，详细报错信息

            成功:
                return True, [True, True, isRelative, True]
            失败
                return False, “详细失败报错信息”
    """
    try:
        # 是否为关联智能资源标志
        isRelative = False

        # 是否为开放HEOP标准
        openHeopFlag, openHeopMsg = channel_VCAResourse_OpenHEOP_cap_judge(integ_isapi,
                                                                           channel, vcaType)
        if not openHeopFlag:
            return False, openHeopMsg

        if openHeopMsg[0]:
            if not openHeopMsg[1]:
                return True, [True, False, isRelative, True, 0]
            else:
                return True, [True, True, isRelative, True, openHeopMsg[2]]

        # 关联智能资源能力
        relatedVCAResourceCapUrl = "/ISAPI/System/RelatedVCAResource/capabilities?format=json"
        relatedVCAResourceCap = integ_isapi.get(relatedVCAResourceCapUrl)
        print(relatedVCAResourceCap)

        # 非关联智能资源能力
        vcaResourceCapUrl = f"/ISAPI/System/Video/inputs/channels/{channel}/VCAResource/capabilities"
        vcaResourceCap = integ_isapi.get(vcaResourceCapUrl)
        print(vcaResourceCap)

        # 所有支持的智能资源能力列表
        resultCapList = []

        # # 设备为开放HEOP设备
        # if "<isSupportOpenPlatform>true</isSupportOpenPlatform>" in openHeopCap:
        #     return False, "当前设备为开放HEOP设备，目前工具暂不支持开放HEOP设备的APP有效性测试，请手动测试!"
        # 不存在智能应用概念
        if ("notSupport" in relatedVCAResourceCap
            or "invalidOperation" in relatedVCAResourceCap) and (
                "notSupport" in vcaResourceCap
                or "invalidOperation" in vcaResourceCap):
            if vcaType == "smart":
                smartUrl = "/ISAPI/Smart/capabilities"
                ret = integ_isapi.get(smartUrl)
                if "notSupport" in ret or "invalidOperation" in ret:
                    return True, [False, False, isRelative]
                if "</SmartCap>" not in ret:
                    return False, f"设备通道{channel}获取Smart智能资源能力失败:{integ_isapi.get_subStatusCode(ret)}!"
                return True, [False, True, isRelative, False]
            else:
                return True, [False, False, isRelative, False]
        # 关联智能模式
        elif ("notSupport" not in relatedVCAResourceCap
              and "invalidOperation" not in relatedVCAResourceCap
        ) and "RelatedVCAResourceCap" in relatedVCAResourceCap:
            isRelative = True
            relationResuleList = []
            capList = json.loads(relatedVCAResourceCap
                                 )["RelatedVCAResourceCap"]["ResourceList"]
            for item in capList:
                tempDict = {}
                for key, value in item.items():
                    if key == "MainChannelResource":
                        MainChannelId = value['channelID']
                        MainChannelType = value['resourceType']
                        tempDict[str(MainChannelId)] = MainChannelType
                    elif key == "RelatedChannelsResource":
                        for sonChannel in value:
                            sonChannelId = sonChannel['channelID']
                            sonChannelType = sonChannel['resourceType']
                            tempDict[str(sonChannelId)] = sonChannelType
                relationResuleList.append(tempDict)
            print(relationResuleList)
            for item in relationResuleList:
                temp = item[str(channel)]
                if isinstance(temp, list):
                    resultCapList += temp
                elif temp not in capList:
                    resultCapList.append(temp)
            resultCapList = list(set(resultCapList))
        # 非关联智能模式
        elif ("notSupport" not in vcaResourceCap
              and "invalidOperation" not in vcaResourceCap
        ) and "</VCAResource>" in vcaResourceCap:
            resultCapList = re.findall('<type opt="(.*?)">',
                                       vcaResourceCap, re.S)[0].split(',')
        else:
            return False, f"设备通道{channel}获取智能资源能力失败!"

        print(resultCapList)

        if len(resultCapList) != 0 and vcaType in resultCapList:
            return True, [True, True, isRelative, False]
        else:
            return True, [True, False, isRelative, False]
    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备通道{channel}判断智能资源能力出现异常:{str(e)}!"


def channel_Intelligent_VCAResource_change_isRebootRequired(
        integ_isapi, channel=1, vcaType='facesnap'):
    '''
        判断切换智能是不是需要重启操作接口

        参数:
            channel:通道号,从1开始
            vcaType:智能资源名称
        返回:
            return 接口执行成功，是否需要重启
            return 接口执行失败，详细报错信息

            成功:
                return True, True
                return True, False
            失败:
                return False, “出现异常”
    '''
    try:
        rebootRequiredUrl = f"/ISAPI/System/Video/inputs/channels/{channel}/VCAResource/test"

        DATA = f'''
            <?xml version="1.0" encoding="UTF-8"?>
            <VCAResource version="2.0" xmlns="http://www.hikvision.com/ver20/XMLSchema">
                <type>{vcaType}</type>
            </VCAResource>
            '''
        ret = integ_isapi.post(rebootRequiredUrl, DATA)
        if "</VCAResourceResult>" not in ret:
            return False, f'设备通道{channel}判断切换智能{vcaType}是否需要重启协议执行失败:{integ_isapi.get_subStatusCode(ret)}!'

        getReboot = re.findall('<reboot>(.*?)</reboot>', ret, re.S)[0]

        if getReboot == "false":
            return True, False

        return True, True
    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备通道{channel}判断切换智能{vcaType}是否需要重启出现异常:{str(e)}!"


def integ_reboot_device(integ_isapi, deviceInfo):
    '''判断设备重启和自检是否成功'''
    try:
        # # 重启设备
        ip = deviceInfo["ip"]
        username = deviceInfo["username"]
        password = deviceInfo["password"]
        dev_type = deviceInfo["dev_type"]
        url = "http://" + deviceInfo["ip"] + "/ISAPI/System/reboot/"  # 需要发起请求的URL
        url_1 = "http://" + deviceInfo["ip"] + "/ISAPI/System/deviceInfo/"
        auth = HTTPDigestAuth(username=username, password=password)
        res = requests.put(url=url, data=None, auth=auth, timeout=30)
        isRebootFlag = False
        if res.status_code == 200:
            time.sleep(10)
            # 判断是否重启开始
            t = 0
            while (t < 300):
                try:
                    ret = requests.get(url=url_1, data=None, auth=auth, timeout=10)
                    ##保证设备进入重启状态，当下发重启协议的时候设备需要响应一定时间，否则协议是通的
                    if "</DeviceInfo>" not in ret.text:
                        VD_LOG("设备进入重启状态，下发重启成功，重启开始=======================")
                        isRebootFlag = True
                        break
                    else:
                        VD_LOG("设备还未进入重启状态，等待开始重启=======================")
                        time.sleep(10)
                        t += 10
                        continue
                except:
                    VD_LOG("设备进入重启状态，下发重启成功，重启开始=======================")
                    isRebootFlag = True
                    break
        if isRebootFlag:
            t = 0
            while (t < 300):
                try:
                    ret = requests.get(url=url_1, data=None, auth=auth, timeout=10)
                    print(ret)
                    if "</DeviceInfo>" not in ret.text:
                        VD_LOG("设备进入重启状态，等待重启成功=======================")
                        time.sleep(10)
                        t += 10
                    else:
                        VD_LOG("重启成功=======================")
                        return True, "重启成功"
                except:
                    VD_LOG("设备进入重启状态，等待重启成功=======================")
                    time.sleep(10)
                    t += 10
                    continue
            # return False,"设备进入重启状态，但是等待五分钟没有重启成功，请查看设备是否有问题"


        else:
            return False, f"下发设备重启后五分钟设备依然未进入重启状态，出现异常!"

        return False, "设备进入重启状态，但是等待五分钟没有重启成功，请查看设备是否有问题"
    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备重启出现异常:{str(e)}!"


# def integ_reboot_device1(integ_isapi,deviceInfo):
#     ip = deviceInfo["ip"]
#     username = deviceInfo["username"]
#     password = deviceInfo["password"]
#     dev_type = deviceInfo["dev_type"]
#     url = "http://" + deviceInfo["ip"] + "/ISAPI/System/reboot/"  # 需要发起请求的URL
#     url_1="http://" + deviceInfo["ip"] + "/ISAPI/System/deviceInfo"
#     auth = HTTPDigestAuth(username=username, password=password)
#     res = requests.put(url=url, data=None, auth=auth, timeout=30)
#     isRebootFlag=False
#
#     if res.status_code == 200 :
#         t = 20 #init waiting time
#         time.sleep(20)
#         result = False
#         data = 'No data return'
#         while t <= int(300):
#             data= requests.get(url=url_1, data=None, auth=auth, timeout=30)
#             if (''  in data):  # cmp(data,'')!=0:
#                 result=True
#                 break
#             time.sleep(10)
#             t += 10
#             VD_LOG("等待重启====================")
#
#         time.sleep(10)  # TODO ensure port [8000] alive
#
#         if result or ('No data return' not in data):
#
#             if "DeviceInfo" in data:
#                 ret = ('OK.', "Device reboot successful.")
#             else:
#                 ret = ('NG.', "Abnormal return : %s." % str(data))
#         else:
#             ret = ('NG.', "Device reboot failed within the limitTime.")
#     else:
#         ret = ("NG.",  "Send reboot cmd failed, %s" % ip)
#
#     return ret

def OpenPlatform_App_runStatus_judge(integ_isapi, vcaType='facesnap'):
    """智能是否已经起来判断"""
    App_corresponding = {
        'basicBehavior':
            ['基本行为', 'basicBehavior', "行为分析", 'Behavior Analysis '],
        'fullBehavior':
            ['所有行为', 'fullBehavior', '行为分析', 'Behavior Analysis '],
        'facesnapBehavior': [
            '人脸抓拍行为', 'facesnapBehavior', '行为分析 + 人脸抓拍', "行为分析", "人脸抓拍",
            'Behavior Analysis  + Face Capture ',
            'Behavior Analysis  + Face Recognition', 'Behavior Analysis',
            'Face Capture', 'Face Recognition'
        ],
        'facesnap':
            ['人脸抓拍', 'facesnap', 'Face Capture', 'Face Recognition'],
        'TFS': ['交通取证检测', 'TFS', '违章事件', 'Traffic Violations '],
        'ITS': ['ITS', '智能交通', 'ITS(Intelligent Transportation System)'],
        'smartVehicleDetection': [
            '简易智能&&车辆检测', 'smartVehicleDetection', 'Smart事件 + 车辆检测',
            '车辆检测', 'SMART Event + Vehicle Detection', 'Vehicle Detection'
        ],
        'smartHVTDetection': [
            '简易智能&&混行检测', 'smartHVTDetection', 'Smart事件 + 混行检测',
            'SMART Event + Mixed-traffic Detection',
            'Mixed-traffic Detection', 'SMART Event'
        ],
        'smart': ['简易智能', 'smart', 'Smart Event', 'Smart事件'],
        'judicial': ['司法', 'judicial', '司法事件', 'Judicial Events'],
        'smart264AndRoadDetection': [
            'Smart264事件+道路监控', 'smart264AndRoadDetection',
            'Smart事件 + 道路监控', '道路监控', 'SMART Event + Road Traffic',
            'Road Traffic'
        ],
        'smart264AndFaceDetection': [
            'Smart264事件+人脸侦测', 'smart264AndFaceDetection',
            'Smart事件 + 人脸侦测', '人脸侦测', 'SMART Event + Face Detection',
            'Face Detection'
        ],
        'smart264AndHeatMap': [
            'Smart264事件+热度图', 'smart264AndHeatMap', 'Smart事件 + 热度图', '热度图',
            'SMART Event + Heat Map', 'Heat Map'
        ],
        'smartIntelligentMonitor': [
            '简易智能&&车辆违法停车检测', 'smartIntelligentMonitor', 'Smart事件 + 智慧监控',
            '智慧监控', 'SMART Event + Intelligent Monitoring',
            'Intelligent Monitoring'
        ],
        'smartVehicleIllegalParkingDetection': [
            'Smart事件 + 车辆检测 + 违停检测', 'smartVehicleIllegalParkingDetection',
            '违停检测', '车辆检测',
            'SMART Event + Vehicle Detection + Illegal Parking Detection',
            'Vehicle Detection', 'Illegal Parking Detection'
        ],
        'smartTrafficDataCollection': [
            'Smart事件+智慧监控', 'smartTrafficDataCollection',
            'Smart事件 + 交通数据采集', '交通数据采集',
            'SMART Event + Traffic Data Statistics',
            'Traffic Data Statistics'
        ],
        'roadDetection': ['道路检测', 'roadDetection', '道路监控', 'Road Traffic'],
        'humanRecognition': ['人体目标识别', 'humanRecognition'],
        'perimeterCapture':
            ['周界抓拍', 'perimeterCapture', 'Perimeter Capture'],
        'vehicleDetection': ['车辆检测', 'vehicleDetection'],
        'hvtVehicleDetection':
            ['混行检测', 'HVTDetection', 'hvtVehicleDetection'],
        'mixedTargetDetection': [
            '混合目标检测', 'mixedTargetDetection', '全结构化',
            'Capture Target With Feature'
        ],
        'trackingCaptureMode': ['全景抓拍模式', 'trackingCaptureMode'],
        'nonTrackingCaptureMode': ['单点抓拍模式', 'nonTrackingCaptureMode'],
        'close': ['关闭该通道下所有智能资源', 'close', '普通监控', 'Monitoring'],
        'faceHumanModelingContrast': [
            '混合目标', 'faceHumanModelingContrast', "混合目标比对", "比对模式", "人体",
            "人脸", "人体+人脸", 'Comparison Mode',
            'Multi-Target-Type Comparison', '全结构化', '比对建模'
        ],
        'cityManagement':
            ['智慧城管监测', 'cityManagement', '智慧城管', 'City Management'],
        'teacherBehavior': [
            '教师行为检测', 'teacherBehavior', '教师行为分析',
            "Teacher's Behavior Analysis"
        ],
        '12MPLiveView':
            ['主码流1200W', '12MPLiveView', '1200W预览模式', '12 MP Live Views'],
        'personQueueDetection':
            ['人员排队检测', 'personQueueDetection', '区域关注度', 'Queue Management'],
        'verticalPeopleCounting': ['垂直客流', 'verticalPeopleCounting'],
        'AIOpenPlatform':
            ['AI开放平台', 'AIOpenPlatform', '应用仓库', 'AI Open Platform'],
        'mixedTargetFaceHuman': [
            '混合目标抓拍人脸人体模式',
            'mixedTargetFaceHuman',
            '人脸 + 人体',
            '人脸',
            '人体',
            'Face + Human',
            'Face',
            'Human',
        ],
        'mixedTargetAll':
            ['混合目标抓拍全结构化模式', 'mixedTargetAll', 'Capture Target With Feature'],
        'safetyHelmet': ['安全帽检测', 'safetyHelmet'],
        'faceContrast': [
            '人脸比对', 'faceContrast', 'Face Picture Comparison', '人脸抓拍',
            'Face Capture', '比对建模'
        ],
        'faceMonitor': ['人脸布控', 'faceMonitor'],
        'vehicleMonitor': ['车辆布控', 'vehicleMonitor'],
        'faceCounting': ['人脸客流', 'faceCounting'],
        'heelPDC': ['倾斜客流', 'heelPDC', '倾斜客流统计', 'People Counting'],
        'personDensity':
            ['人员密度', 'personDensity', '人员密度检测', "People Density"],
        'behaviorMixedTargetAll':
            ['行为分析混合目标全结构化', 'behaviorMixedTargetAll', '全结构化'],
        'operationMonitor': ['作业监测', 'operationMonitor'],
        'fielddetection': ['区域入侵', 'fielddetection'],
        'personArming': ['人员布控', 'personArming', 'Person Arming'],
        'smokeDetectAlarm': ['抽烟检测', 'smokeDetectAlarm'],
        'smokeDetection': ['烟雾检测', 'smokeDetection'],
        'personDensityQueueLeavePosition': [
            '横幅检测', 'personDensityQueueLeavePosition'
        ],
        'waterQualityDetection': ['水质检测', 'waterQualityDetection'],
        'ATMSurround': ['ATM环境模式', 'ATMSurround'],
        'ATMPanel': ['ATM面板模式', 'ATMPanel'],
        'ATMSafetyCabin': ['ATM防护舱模式', 'ATMSafetyCabin'],
        'intelligentTraffic': ['智能交通', 'intelligentTraffic'],
        'objectsThrownDetection': [
            '高空抛物检测', '高空抛物', 'objectsThrownDetection', 'objectsThrown'
        ],
        'fullBehaviorAndMixedTargetDetection': [
            '行为分析 + 混合目标检测', 'fullBehaviorAndMixedTargetDetection',
            '混合目标检测', '行为分析', 'Multi-Target-Type Detection',
            'Behavior Analysis '
        ]
    }

    try:
        if vcaType == "close":
            time.sleep(10)
            return True, "普通监控无需判断!"

        # 判断设备是否为海外设备
        judgeFlag, isOverseas = device_is_overseas_judge(integ_isapi)
        if not judgeFlag:
            return False, isOverseas

        url = "/ISAPI/Custom/OpenPlatform/App"
        errMsg = ""
        isUp = False
        for index in range(24):
            ret = integ_isapi.get(url)

            print(ret)
            if "notSupport" in ret or "invalidOperation" in ret:
                time.sleep(15)
                isUp = True
                break

            if "404 -- Not Found" in ret:
                time.sleep(15)
                isUp = True
                break

            if "</packageName>" not in ret:
                errMsg = f"设备等待两分钟依然未获取到{vcaType}智能是否起来的信息!"
                time.sleep(5)
                continue

            idList = re.findall("<id>(.*?)</id>", ret, re.S)
            packageNameList = re.findall(
                "<packageName>(.*?)</packageName>", ret, re.S)
            runStatusList = re.findall("<runStatus>(.*?)</runStatus>", ret,
                                       re.S)

            id_status_dict = {}
            for i in range(len(idList)):
                id_status_dict[idList[i]] = runStatusList[i]
            print(id_status_dict)

            name_status_dict = {}
            for i in range(len(packageNameList)):
                name_status_dict[packageNameList[i]] = runStatusList[i]
            print(name_status_dict)

            if vcaType == "AIOpenPlatform":
                pageItemUrl = "/ISAPI/Custom/OpenPlatform/pageItem"
                ret = integ_isapi.get(pageItemUrl)
                if "notSupport" in ret or "invalidOperation" in ret:
                    time.sleep(15)
                    isUp = True
                    break
                if "</pageItemList>" not in ret:
                    errMsg = "设备等待两分钟依然未获取到智能是否起来的信息!"
                    time.sleep(5)
                    continue

                idListTemp = re.findall("<id>(.*?)</id>", ret, re.S)
                appNameListTemp = re.findall("<appName>(.*?)</appName>",
                                             ret, re.S)

                print(
                    f"appNameListTemp={appNameListTemp}, idListTemp={idListTemp}"
                )

                isFind = False
                for index in range(len(appNameListTemp)):
                    if appNameListTemp[index] == "aiOpenPlat":
                        status = runStatusList[index]
                        if status == "true":
                            isFind = True
                            break
                        else:
                            continue

                if not isFind:
                    errMsg = "设备等待两分钟依然未获取到智能是否起来的信息!"
                    time.sleep(5)
                    continue

                isUp = True
                break

            nameList = HEOP_APP_EN[
                vcaType] if isOverseas else HEOP_APP_ZH[vcaType]

            tempFlag = True
            for name in nameList:
                if name not in packageNameList:
                    errMsg = f"设备等待两分钟依然未获取到{name}是否起来的信息!"
                    tempFlag = False
                    break
                if name_status_dict[name] != "true":
                    errMsg = f"设备等待两分钟后，{name}的是否起来的状态依然是false!"
                    tempFlag = False
                    break

            if not tempFlag:
                time.sleep(5)
                continue

            isUp = True
            break

        if not isUp:
            return False, errMsg

        return True, "智能已经起来!"
    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备判断切换智能{vcaType}后智能是否起来出现异常:{str(e)}!"


def channel_Intelligent_VCAResource_change(integ_isapi, deviceInfo,
                                           channelList,
                                           vcaType='facesnap',
                                           isRelative=False,
                                           isOpenHeop=False,
                                           appId=0,
                                           sleepTime=60):
    """
        智能资源切换接口

        参数:
            channel:通道号,从1开始
            vcaType:智能资源名称
            isRelative:是否为关联智能资源
            isOpenHeop:是否为开放HEOP架构
            appId:开放HEOP架构智能对应id
            sleepTime:掉电记忆等待时间
        返回:
            return 接口执行成功，成功信息
            return 接口执行失败，详细报错信息

            成功:
                return True, “智能资源切换成功”
            失败
                return False, “详细失败报错信息”
    """
    try:
        ip = deviceInfo["ip"]
        username = deviceInfo["username"]
        password = deviceInfo["password"]
        dev_type = deviceInfo["dev_type"]
        needReboot = False
        needRebootFlag, needRebootMsg = channel_Intelligent_VCAResource_change_isRebootRequired(integ_isapi,
                                                                                                1, vcaType)
        if not needRebootFlag:
            print(needRebootMsg)
        else:
            needReboot = needRebootMsg

        if isOpenHeop:
            channel = 1
            # 获取设备当前已加载所有智能资源
            openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
            openHeopData = integ_isapi.get(openHeopUrl)
            print(openHeopData)
            if "</AppList>" not in openHeopData:
                return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{integ_isapi.get_subStatusCode(openHeopData)}!"

            idList = re.findall("<id>(.*?)</id>", openHeopData, re.S)
            packageNameList = re.findall(
                "<packageName>(.*?)</packageName>", openHeopData, re.S)
            runStatusList = re.findall("<runStatus>(.*?)</runStatus>",
                                       openHeopData, re.S)

            # 先停用所有已启用的智能
            for i in range(len(runStatusList)):
                if runStatusList[i] == "true":
                    stopUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{idList[i]}/stop"
                    print(f"[PUT {stopUrl}]")
                    ret = requests.put(stopUrl,
                                       data=None,
                                       auth=HTTPDigestAuth(
                                           username, password),
                                       timeout=180)
                    status_code = ret.status_code
                    content = ret.text
                    ret.close()
                    print(content)
                    if "OK" not in content:
                        return False, f"在正式测试目标智能前关闭所有智能时出现关闭{packageNameList[i]}智能协议执行失败:{integ_isapi.get_subStatusCode(content)}!"
                    time.sleep(5)
                else:
                    continue

            # 掉电记忆
            # if dev_type == 'IPD' and channel == 1:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)
            #     moveFlag, moveMsg = self.put_device_position_cfg()
            #     if not moveFlag:
            #         return False, moveMsg
            #     time.sleep(3)
            # else:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)

            # 配置启用目标智能
            startUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{appId}/start"
            print(f"[PUT {startUrl}]")
            req = requests.put(startUrl,
                               data=None,
                               auth=HTTPDigestAuth(username, password),
                               timeout=300)
            status_code = req.status_code
            content = req.text
            req.close()
            print(content)

            if status_code == 200 and ("rebootRequired" in content
                                       or needReboot):
                rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                if not rebootFlag:
                    return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                # if dev_type== 'IPD' and channel == 1:
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)

                isUpFlag, isUpMsg = OpenPlatform_App_runStatus_judge(integ_isapi,
                                                                     vcaType)
                if not isUpFlag:
                    return False, isUpMsg

                return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
            elif status_code == 200 and "OK" in content:
                isUpFlag, isUpMsg = OpenPlatform_App_runStatus_judge(integ_isapi,
                                                                     vcaType)
                if not isUpFlag:
                    return False, isUpMsg

                return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
            else:
                return False, f"设备通道{channel}切换智能资源{vcaType}协议执行失败:{integ_isapi.get_subStatusCode(content)}!"
        elif not isRelative:
            # 非关联智能
            # 获取当前设备该通道智能资源
            # 可支持的通道
            a_dic = {}
            for channel in channelList:
                vcaResourceUrl = f'/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
                ret = integ_isapi.get(vcaResourceUrl)
                print(ret)
                if "</VCAResource>" not in ret:
                    return False, f'设备通道{channel}获取当前智能资源协议执行失败:{integ_isapi.get_subStatusCode(ret)}!'

                currentType = re.findall("<type>(.*?)</type>", ret, re.S)[0]
                print("currentType=", currentType)

                # 判断当前智能是否为指定智能，如果是跳到下一个通道继续判断
                if currentType == vcaType:
                    continue

                # 当前智能资源与需要切换的智能资源比较
                # if currentType == vcaType:
                #     if dev_type == 'IPD' and channel == 1:
                #         print("等待60s保证掉电记忆配置保存!")
                #         time.sleep(sleepTime)
                #         moveFlag, moveMsg = self.put_device_position_cfg()
                #         if not moveFlag:
                #             return False, moveMsg
                #         time.sleep(3)
                #     else:
                #         print("等待60s保证掉电记忆配置保存!")
                #         time.sleep(sleepTime)
                #     return True, f"设备通道{channel}当前智能资源就是{vcaType},无需切换!"

                # 切换智能资源操作
                # if dev_type == 'IPD' and channel == 1:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                #     # moveFlag, moveMsg = self.put_device_position_cfg()
                #     # if not moveFlag:
                #     #     return False, moveMsg
                #     time.sleep(3)
                # else:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)

                vcaResourceUrl = f'http://{ip}/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
                DATA = f"""
                    <VCAResource xmlns="http://www.hikvision.com/ver20/XMLSchema" version="2.0">
                        <type>{vcaType}</type>
                    </VCAResource>
                    """
                print(DATA)
                print(f"[PUT {vcaResourceUrl}]")
                req = requests.put(vcaResourceUrl,
                                   data=DATA,
                                   auth=HTTPDigestAuth(username, password),
                                   timeout=300)
                status_code = req.status_code
                content = req.text
                req.close()
                print(content)
                if status_code == 200:
                    a_dic[str(channel)] = content
                else:
                    a_dic[str(channel)] = "协议失败"
            if not a_dic:
                return True, f"设备所有可切换的通道当前智能资源就是{vcaType},无需切换!"
            elif "协议失败" not in a_dic.values():
                rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                if not rebootFlag:
                    return False, f"所有支持的通道切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                isUpFlag, isUpMsg = OpenPlatform_App_runStatus_judge(integ_isapi,
                                                                     vcaType)
                if not isUpFlag:
                    return False, isUpMsg

                return True, f"所有支持的通道切换智能资源{vcaType}成功,已完成重启!"
                # if  "rebootRequired" in a_dic.values() or needReboot:
                #     rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                #     if not rebootFlag:
                #         return False, f"所有支持的通道切换智能资源{vcaType}后重启失败:{rebootMsg}!"
                #
                #     isUpFlag, isUpMsg = OpenPlatform_App_runStatus_judge(integ_isapi,
                #         vcaType)
                #     if not isUpFlag:
                #         return False, isUpMsg
                #
                #     return True, f"所有支持的通道切换智能资源{vcaType}成功,已完成重启!"
                # elif "OK" in a_dic.values():
                #     isUpFlag, isUpMsg =OpenPlatform_App_runStatus_judge(integ_isapi,
                #         vcaType)
                #     if not isUpFlag:
                #         return False, isUpMsg
                #
                #     return True, f"所有支持的通道切换智能资源{vcaType}成功,已完成重启!"
                # else:
                #     return False, f"设备通道切换智能资源{vcaType}协议执行失败!"

            else:
                return False, f"设备通道切换智能资源{vcaType}协议执行失败!"
        else:
            # 获取当前设备该通道智能资源
            chan_current_type = {}

            vcaResourceUrl = '/ISAPI/System/RelatedVCAResource?format=json'
            vcaResource = integ_isapi.get(vcaResourceUrl)
            print(vcaResource)
            if "RelatedVCAResource" not in vcaResource:
                return False, f"设备通道获取关联智能资源能力失败:{integ_isapi.get_subStatusCode(vcaResource)}!"

            channelsResource = json.loads(
                vcaResource)['RelatedVCAResource']['ChannelsResource']
            print(channelsResource)

            for item in channelsResource:
                chan_current_type[str(
                    item['channelID'])] = item['resourceType']

            print(chan_current_type)

            # 关联智能资源能力
            relatedVCAResourceCapUrl = "/ISAPI/System/RelatedVCAResource/capabilities?format=json"
            relatedVCAResourceCap = integ_isapi.get(
                relatedVCAResourceCapUrl)
            print(relatedVCAResourceCap)
            if "RelatedVCAResourceCap" not in relatedVCAResourceCap:
                return False, f"设备通道获取关联智能资源能力失败:{integ_isapi.get_subStatusCode(relatedVCAResourceCap)}"

            # 主通道号
            MainChannelId = ''
            # 关联关系表
            relationResuleList = []
            # 处理关联关系并生成关联关系表
            capList = json.loads(relatedVCAResourceCap
                                 )['RelatedVCAResourceCap']['ResourceList']
            for item in capList:
                tempDict = {}
                for key, value in item.items():
                    if key == "MainChannelResource":
                        MainChannelId = value['channelID']
                        MainChannelType = value['resourceType']
                        tempDict[str(MainChannelId)] = MainChannelType
                    elif key == "RelatedChannelsResource":
                        for sonChannel in value:
                            sonChannelId = sonChannel['channelID']
                            sonChannelType = sonChannel['resourceType']
                            tempDict[str(sonChannelId)] = sonChannelType
                relationResuleList.append(tempDict)

            print(MainChannelId)
            print(relationResuleList)

            flag, paras = getReceiveDicParas(vcaType, relationResuleList)
            print(paras)

            putData = {"RelatedVCAResource": {"ChannelsResource": []}}

            for key, value in paras.items():
                temp = {"channelID": "1", "resourceType": "smart"}
                temp["channelID"] = int(key)
                if isinstance(value, str):
                    temp["resourceType"] = value
                elif isinstance(value, list):
                    flag_find = 0
                    for i in value:
                        if i == vcaType:
                            temp["resourceType"] = i
                            flag_find = 1
                            break
                    if not flag_find:
                        temp["resourceType"] = value[0]
                else:
                    pass
                putData['RelatedVCAResource']['ChannelsResource'].append(temp)

            DATA = json.dumps(putData)
            print(DATA)

            # 切换智能资源
            # if dev_type == 'IPD' and channel == 1:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)
            #     moveFlag, moveMsg = self.put_device_position_cfg()
            #     if not moveFlag:
            #         return False, moveMsg
            #     time.sleep(3)
            # else:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)

            vcaResourceUrl = f'http://{ip}/ISAPI/System/RelatedVCAResource?format=json'
            print(f"[PUT {vcaResourceUrl}]")
            req = requests.put(vcaResourceUrl,
                               data=DATA,
                               auth=HTTPDigestAuth(username, password),
                               timeout=300)
            status_code = req.status_code
            content = req.text
            req.close()
            print(content)

            if status_code == 200 and ("rebootRequired" in content or needReboot):
                rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                if not rebootFlag:
                    return False, f"设备切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                # if dev_type == 'IPD' and channel == 1:
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)
                #
                # isUpFlag, isUpMsg = self.OpenPlatform_App_runStatus_judge(
                #     vcaType)
                # if not isUpFlag:
                #     return False, isUpMsg

                return True, f"设备切换智能资源{vcaType}成功,已完成重启!"
            elif status_code == 200 and "OK" in content:
                isUpFlag, isUpMsg = OpenPlatform_App_runStatus_judge(integ_isapi,
                                                                     vcaType)
                if not isUpFlag:
                    return False, isUpMsg
                return True, f"设备切换智能资源{vcaType}成功,无需重启!"
            else:
                return False, f"设备配置切换智能资源为{vcaType}协议执行失败:{integ_isapi.get_subStatusCode(content)}!"
    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备切换智能资源为{vcaType}出现异常:{str(e)}!"


def channel_Intelligent_VCAResource_change_without_judge_runStatus(
        integ_isapi,
        deviceInfo,
        channel=1,
        vcaType='facesnap',
        isRelative=False,
        isOpenHeop=False,
        appId=0,
        sleepTime=60):
    """
        智能资源切换接口

        参数:
            channel:通道号,从1开始
            vcaType:智能资源名称
            isRelative:是否为关联智能资源
            isOpenHeop:是否为开放HEOP架构
            appId:开放HEOP架构智能对应id
        返回:
            return 接口执行成功，成功信息
            return 接口执行失败，详细报错信息

            成功:
                return True, “智能资源切换成功”
            失败
                return False, “详细失败报错信息”
    """

    try:
        ip = deviceInfo["ip"]
        username = deviceInfo["username"]
        password = deviceInfo["password"]
        needReboot = False
        needRebootFlag, needRebootMsg = channel_Intelligent_VCAResource_change_isRebootRequired(integ_isapi,
                                                                                                channel, vcaType)
        if not needRebootFlag:
            print(needRebootMsg)
        else:
            needReboot = needRebootMsg

        if isOpenHeop:
            # 获取设备当前已加载所有智能资源
            openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
            openHeopData = integ_isapi.get(openHeopUrl)
            print(openHeopData)
            if "</AppList>" not in openHeopData:
                return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{integ_isapi.get_subStatusCode(openHeopData)}!"

            idList = re.findall("<id>(.*?)</id>", openHeopData, re.S)
            packageNameList = re.findall(
                "<packageName>(.*?)</packageName>", openHeopData, re.S)
            runStatusList = re.findall("<runStatus>(.*?)</runStatus>",
                                       openHeopData, re.S)

            # 先停用所有已启用的智能
            for i in range(len(runStatusList)):
                if runStatusList[i] == "true":
                    stopUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{idList[i]}/stop"
                    print(f"[PUT {stopUrl}]")
                    ret = requests.put(stopUrl,
                                       data=None,
                                       auth=HTTPDigestAuth(
                                           username, password),
                                       timeout=180)
                    status_code = ret.status_code
                    content = ret.text
                    ret.close()
                    print(content)
                    if "OK" not in content:
                        return False, f"在正式测试目标智能前关闭所有智能时出现关闭{packageNameList[i]}智能协议执行失败:{integ_isapi.get_subStatusCode(content)}!"
                    time.sleep(5)
                else:
                    continue

            # # 掉电记忆
            # if self.dict["dev_type"] == 'IPD' and channel == 1:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)
            #     moveFlag, moveMsg = self.put_device_position_cfg()
            #     if not moveFlag:
            #         return False, moveMsg
            #     time.sleep(3)
            # else:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)

            # 配置启用目标智能
            startUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{appId}/start"
            print(f"[PUT {startUrl}]")
            req = requests.put(startUrl,
                               data=None,
                               auth=HTTPDigestAuth(username,
                                                   password),
                               timeout=180)
            status_code = req.status_code
            content = req.text
            req.close()
            print(content)

            if status_code == 200 and ("rebootRequired" in content
                                       or needReboot):
                rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                if not rebootFlag:
                    return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                # if self.dict["dev_type"] == 'IPD' and channel == 1:
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)

                return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
            elif status_code == 200 and "OK" in content:
                return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
            else:
                return False, f"设备通道{channel}切换智能资源{vcaType}协议执行失败:{integ_isapi.get_subStatusCode(content)}!"
        elif not isRelative:
            # 获取当前设备该通道智能资源
            vcaResourceUrl = f'/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
            ret = integ_isapi.get(vcaResourceUrl)
            print(ret)
            if "</VCAResource>" not in ret:
                return False, f'设备通道{channel}获取当前智能资源协议执行失败:{integ_isapi.get_subStatusCode(ret)}!'

            currentType = re.findall("<type>(.*?)</type>", ret, re.S)[0]
            print("currentType=", currentType)

            # 当前智能资源与需要切换的智能资源比较
            # if currentType == vcaType:
            #     if self.dict["dev_type"] == 'IPD' and channel == 1:
            #         print("等待60s保证掉电记忆配置保存!")
            #         time.sleep(sleepTime)
            #         moveFlag, moveMsg = self.put_device_position_cfg()
            #         if not moveFlag:
            #             return False, moveMsg
            #         time.sleep(3)
            #     else:
            #         print("等待60s保证掉电记忆配置保存!")
            #         time.sleep(sleepTime)
            #     return True, f"设备通道{channel}当前智能资源就是{vcaType},无需切换!"

            # 切换智能资源操作
            # if self.dict["dev_type"] == 'IPD' and channel == 1:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)
            #     moveFlag, moveMsg = self.put_device_position_cfg()
            #     if not moveFlag:
            #         return False, moveMsg
            #     time.sleep(3)
            # else:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)

            vcaResourceUrl = f'http://{ip}/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
            DATA = f"""
                <VCAResource xmlns="http://www.hikvision.com/ver20/XMLSchema" version="2.0">
                    <type>{vcaType}</type>
                </VCAResource>
                """
            print(DATA)
            print(f"[PUT {vcaResourceUrl}]")
            req = requests.put(vcaResourceUrl,
                               data=DATA,
                               auth=HTTPDigestAuth(username,
                                                   password),
                               timeout=200)
            status_code = req.status_code
            content = req.text
            req.close()
            print(content)

            if status_code == 200 and ("rebootRequired" in content
                                       or needReboot):
                rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                if not rebootFlag:
                    return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                # if self.dict["dev_type"] == 'IPD' and channel == 1:
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)

                return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
            elif status_code == 200 and "OK" in content:
                return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
            else:
                return False, f"设备通道{channel}切换智能资源{vcaType}协议执行失败:{integ_isapi.get_subStatusCode(content)}!"
        else:
            # 获取当前设备该通道智能资源
            chan_current_type = {}

            vcaResourceUrl = '/ISAPI/System/RelatedVCAResource?format=json'
            vcaResource = integ_isapi.get(vcaResourceUrl)
            print(vcaResource)
            if "RelatedVCAResource" not in vcaResource:
                return False, f"设备通道{channel}获取关联智能资源能力失败:{integ_isapi.get_subStatusCode(vcaResource)}!"

            channelsResource = json.loads(
                vcaResource)['RelatedVCAResource']['ChannelsResource']
            print(channelsResource)

            for item in channelsResource:
                chan_current_type[str(
                    item['channelID'])] = item['resourceType']

            print(chan_current_type)

            # 当前智能资源与需要切换的智能资源比较
            if chan_current_type[str(channel)] == vcaType:
                # if self.dict["dev_type"] == 'IPD' and channel == 1:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)
                # else:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                return True, f"设备通道{channel}当前智能资源就是{vcaType},无需切换!"

            # 关联智能资源能力
            relatedVCAResourceCapUrl = "/ISAPI/System/RelatedVCAResource/capabilities?format=json"
            relatedVCAResourceCap = integ_isapi.get(
                relatedVCAResourceCapUrl)
            print(relatedVCAResourceCap)
            if "RelatedVCAResourceCap" not in relatedVCAResourceCap:
                return False, f"设备通道{channel}获取关联智能资源能力失败:{integ_isapi.get_subStatusCode(relatedVCAResourceCap)}!"

            # 主通道号
            MainChannelId = ''
            # 关联关系表
            relationResuleList = []
            # 处理关联关系并生成关联关系表
            capList = json.loads(relatedVCAResourceCap
                                 )['RelatedVCAResourceCap']['ResourceList']
            for item in capList:
                tempDict = {}
                for key, value in item.items():
                    if key == "MainChannelResource":
                        MainChannelId = value['channelID']
                        MainChannelType = value['resourceType']
                        tempDict[str(MainChannelId)] = MainChannelType
                    elif key == "RelatedChannelsResource":
                        for sonChannel in value:
                            sonChannelId = sonChannel['channelID']
                            sonChannelType = sonChannel['resourceType']
                            tempDict[str(sonChannelId)] = sonChannelType
                relationResuleList.append(tempDict)

            print(MainChannelId)
            print(relationResuleList)

            paras = {}
            for item in relationResuleList:
                if vcaType in item[str(channel)]:
                    paras = item
                    break
            print(paras)

            putData = {"RelatedVCAResource": {"ChannelsResource": []}}

            for index in range(1, len(paras) + 1):
                temp = {"channelID": "1", "resourceType": "smart"}
                if index == channel:
                    temp["channelID"] = index
                    temp["resourceType"] = vcaType
                else:
                    temp["channelID"] = index
                    if isinstance(paras[str(index)], list):
                        temp["resourceType"] = paras[str(index)][0]
                    else:
                        temp["resourceType"] = paras[str(index)]
                putData['RelatedVCAResource']['ChannelsResource'].append(
                    temp)

            DATA = json.dumps(putData)
            VD_LOG(DATA)

            # 切换智能资源
            # if self.dict["dev_type"] == 'IPD' and channel == 1:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)
            #     moveFlag, moveMsg = self.put_device_position_cfg()
            #     if not moveFlag:
            #         return False, moveMsg
            #     time.sleep(3)
            # else:
            #     print("等待60s保证掉电记忆配置保存!")
            #     time.sleep(sleepTime)

            vcaResourceUrl = f'http://{ip}/ISAPI/System/RelatedVCAResource?format=json'
            print(f"[PUT {vcaResourceUrl}]")
            req = requests.put(vcaResourceUrl,
                               data=DATA,
                               auth=HTTPDigestAuth(username,
                                                   password),
                               timeout=200)
            status_code = req.status_code
            content = req.text
            req.close()
            print(content)

            if status_code == 200 and ("rebootRequired" in content
                                       or needReboot):
                rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                if not rebootFlag:
                    return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                # if self.dict["dev_type"] == 'IPD' and channel == 1:
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)

                return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
            elif status_code == 200 and "OK" in content:

                return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
            else:
                return False, f"设备通道{channel}配置切换智能资源为{vcaType}协议执行失败:{integ_isapi.get_subStatusCode(content)}!"
    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备通道{channel}切换智能资源为{vcaType}出现异常:{str(e)}!"


def Smart_isSupport_judge_with_Type(integ_isapi, smartType=None):
    '''smart智能支持的事件能力判断'''
    try:
        smartCapUrl = "/ISAPI/Smart/capabilities"
        ret = integ_isapi.get(smartCapUrl)
        if "</SmartCap>" not in ret:
            return False, f'设备获取Smart智能能力协议执行失败:{integ_isapi.get_subStatusCode(ret)}!'

        supportCap = []
        for item in smartType:
            if f"<isSupport{item}>true</isSupport{item}>" in ret:
                supportCap.append(item)

        return True, supportCap
    except Exception as e:
        print(str(e))
        return False, "获取smart事件能力失败"


def integ_channel_Intelligent_VCAResource_judeg_and_change(
        integ_isapi,
        deviceInfo,
        channel=1,
        vcaType='facesnap',
        isReturnErr=True,
        sleepTime=60):
    """判断是否支持某智能、切换智能、判断智能是否起来"""
    try:
        # 智能资源切换能力判断
        capFlag, capMsg = channel_VCAResourse_cap_judge(integ_isapi,
                                                        channel, vcaType)
        if not capFlag:
            return False, capMsg

        if not capMsg[1]:
            return True, False

        # 设备智能资源之间是否存在关联
        isRelative = capMsg[2]

        # 设备是否为开放HEOP架构
        isOpenHeop = capMsg[3]

        appId = capMsg[4] if isOpenHeop else 0

        # 延时60s保证掉电记忆配置保存
        # 智能资源切换
        changeFlag, changeMsg = channel_Intelligent_VCAResource_change_without_judge_runStatus(integ_isapi, deviceInfo,
                                                                                               channel,
                                                                                               vcaType,
                                                                                               isRelative,
                                                                                               isOpenHeop,
                                                                                               appId,
                                                                                               sleepTime=sleepTime)
        if not changeFlag:
            return False, changeMsg

        # 判断智能是否起来
        isUpFlag, isUpMsg = OpenPlatform_App_runStatus_judge(integ_isapi, vcaType)
        if not isUpFlag and isReturnErr:
            return False, isUpMsg
        elif not isUpFlag and not isReturnErr:
            if "是否起来的状态依然是false" in isUpMsg:
                return True, isUpMsg
            else:
                return False, isUpMsg
        else:
            return True, True
    except Exception as e:
        print(traceback.print_exc())
        return False, f"设备通道{channel}判断智能{vcaType}是否支持且切换该智能判断智能是否起来的操作出现异常:{str(e)}!"

class HeopClass(Integ_Isapi):
    def __init__(self):

def is_heop(integ_isapi):
    # 创建integ_isapi对象
    Integ_Isapi = integ_isapi()
    # 开放HEOP能力判断
    openHeopCapUrl = "/ISAPI/Custom/OpenPlatform/capabilities"
    openHeopCap = Integ_Isapi.get(openHeopCapUrl)
    print(openHeopCap)

    if "notSupport" in openHeopCap or "invalidOperation" in openHeopCap or "404 -- Not Found" in openHeopCap or "</isSupportOpenPlatform>" not in openHeopCap:
        return False, ""

    if "</OpenPlatformCap>" not in openHeopCap:
        return False, f"通道/ISAPI/Custom/OpenPlatform/capabilities接口判断设备是否为开放HEOP平台设备协议执行失败:{Integ_Isapi.get_xml_subStatusCode(openHeopCap)}!"

    isSupportOpenPlatform = re.findall(
        "<isSupportOpenPlatform>(.*?)</isSupportOpenPlatform>",
        openHeopCap, re.S)[0]

    if isSupportOpenPlatform == "false":
        return False, ""

    openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
    openHeopData = Integ_Isapi.get(openHeopUrl)
    print(openHeopData)
    if "</AppList>" not in openHeopData:
        return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{Integ_Isapi.get_subStatusCode(openHeopData)}!"
    if "</App>" not in openHeopData:
        return False, ""
    packageNameList = re.findall("<packageName>(.*?)</packageName>",
                                 openHeopData, re.S)
    if len(packageNameList) == 0:
        return True, ""
    else:
        return True, packageNameList