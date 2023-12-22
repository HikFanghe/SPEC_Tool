# -*- coding: utf-8 -*-
"""
@Auth ： Fanghe.Lin
@File ：mtd_com_fun.py
@IDE ：PyCharm
@Time ： 2023/10/12 10:37
@Institution: Hikvision, China
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl import styles
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from lxml import etree
import os, sys, re
from requests.auth import HTTPDigestAuth, HTTPBasicAuth
import json
import urllib
import paramiko
import time
import ssl
from functools import wraps
from http import cookiejar as cookielib
import traceback
import requests
from lib.Integ_Isapi import Integ_Isapi
import datetime
from collections import defaultdict
import pandas as pd

PY_RUN_CORRECT = "P"
PY_RUN_NA = "NA"
PY_RUN_WRONG = "F"

time_seq = time.strftime('[%Y-%m-%d-%H-%M-%S]', time.localtime(time.time()))

# 中文设备智能包名
HEOP_APP_ZH = {
    "facesnap": ["人脸抓拍"],
    "faceContrast": ["人脸抓拍", "比对建模"],
    "mixedTargetDetection": ["全结构化"],
    "faceHumanModelingContrast": ["全结构化", "比对建模"],
    "roadDetection": ["道路监控"],
    "smart": ["smart"],
    "intelligentTraffic": ["智能交通"],
    "mixedTargetFaceHuman": ["全结构化"],
    "AIOpenPlatform": ["AI开放平台"],
    "personDensity": ["人员密度"],
    "SafeHelmet": ["安全帽检测"],
    'Dismission/Sleeping': ['岗位值守检测'],
    'KitchenHygieneDetection': ['明厨亮灶'],
    'GarbageDetection': ['垃圾检测'],
    'WorkingClothesDetection': ['工服检测'],
    'IndoorFirePassageDetection': ['室内消防通道占用'],
    'riverwayGaugeReading': ["水尺检测"],
    'smokeDetection': ["烟雾检测"],
    'personArming': ["人员布控"],
    'ShipAndFloatDetection': ["船只和漂浮物检测"],
    'multidimensionalPeopleCounting': ["多维客流"],
    'objectsThrownDetection': ["高空抛物检测"],
    'personDensityQueueLeavePosition': ["人数统计"]
}

# 英文设备智能包名
HEOP_APP_EN = {
    "facesnap": ["Face Capture"],
    "faceContrast": ["Face Capture", "Comparison and Modeling"],
    "mixedTargetDetection": ["Capture Target With Feature"],
    "faceHumanModelingContrast":
        ["Capture Target With Feature", "Comparison and Modeling"],
    "roadDetection": ["Road Traffic"],
    "smart": ["Smart Event"],
    "intelligentTraffic": ["Intell Traffic Event"],
    "mixedTargetFaceHuman": ["Capture Target With Feature"],
    "AIOpenPlatform": ["AI Open Plat"],
    "personDensity": ["basePersonDensity"],
    "SafeHelmet": ["Hard Hat Detection"],
    'Dismission/Sleeping': ['Leave Position Detection'],
    'KitchenHygieneDetection': ['KitchenHygieneDetection'],
    'GarbageDetection': ['Garbage Detection'],
    'WorkingClothesDetection': ['Working Clothes Detection'],
    'IndoorFirePassageDetection': ['Indoor Fire Passage Detection'],
    'smokeDetection': ["SMOG DETECT"],
    'ShipAndFloatDetection': ["Ship and Float Detection"],
    'personArming': ["Person Arming"],
    'multidimensionalPeopleCounting': ["multidimensional pdc"],
    'objectsThrownDetection': ["Thrown Detect Event"],
    'personDensityQueueLeavePosition': ["personCounting"]
}

def sslwrap(func):
    @wraps(func)
    def bar(*args, **kw):
        kw['ssl_version'] = ssl.PROTOCOL_TLSv1
        return func(*args, **kw)

    return bar

ssl.wrap_socket = sslwrap(ssl.wrap_socket)
ssl._create_default_https_context = ssl._create_unverified_context

class AutoVivification(dict):
    """定义多级字典"""

    def __getitem__(self, item):
        try:
            return dict.__getitem__(self, item)
        except KeyError:
            value = self[item] = type(self)()
            return value

class Zhimakaimen:
    """class Zhimakaimen
    """
    m_prefix = 'http://10.1.32.22:8080/'
    m_site = 'http://psh.hikvision.com.cn:8080/sso/page/command?pageName=homeCN'
    m_site_action = 'http://psh.hikvision.com.cn:8080/sso/rsa/encrypt'
    m_host = '10.1.32.22:8080'
    m_origin = m_prefix
    m_refer = 'http://psh.hikvision.com.cn:8080/sso/page/command?pageName=homeCN'
    m_loginAddr = 'https://sso.hikvision.com/login?service=http%3A%2F%2Fpsh.hikvision.com.cn%3A8080%2Fsso%2Fpage%2Fcommand%3FpageName%3Dhome'

    # m_loginAddr='https://sso.hikvision.com/login'

    def __init__(self, userName, password):
        # print (userName, password)
        ssl.wrap_socket = sslwrap(ssl.wrap_socket)
        self.m_cookieJar = cookielib.LWPCookieJar("cookie")
        self.m_opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.m_cookieJar))
        self.m_opener.addheaders = [
            ('User-agent',
             'Mozillaebug():/4.0 (compatible; MSIE 5.5; Windows NT)')
        ]
        urllib.request.install_opener(self.m_opener)
        self.m_userName = userName
        self.m_password = password

    def Login(self):
        """login.

        send request with username, password. create cookie store these for service using
        """
        userInfo = {
            'username': self.m_userName,
            'password': self.m_password,
            'lt': '',
            'execution': '',
            '_eventId': '',
        }
        try:
            respond = self.m_opener.open(self.m_loginAddr).read().decode()
            if re.search(r'<div id="msg" class="success">',
                         respond) is not None:
                return True
        except Exception as why:
            print("error:can't connect to internet" + str(why))
            return False

        userInfo['username'] = self.m_userName
        userInfo['password'] = self.m_password
        userInfo['lt'] = re.findall(
            r'name="lt" value="(?P<str1>[0-9a-zA-Z\- ]+)"', respond)[0]
        userInfo['execution'] = re.findall(
            r'name="execution" value="(?P<str1>[0-9a-zA-Z\- ]+)"', respond)[0]
        userInfo['_eventId'] = re.findall(
            r'name="_eventId" value="(?P<str1>[0-9a-zA-Z\- ]+)"', respond)[0]
        # print(userInfo)
        requestStr = urllib.request.Request(
            self.m_loginAddr,
            urllib.parse.urlencode(userInfo).encode("utf-8"))
        # print requestStr
        try:
            respond = self.m_opener.open(requestStr).read().decode()
            # print respond
            # if re.search(r'<div id="msg" class="success">', respond) is not None:
            if re.search(r'PSH Shell', respond) is not None:
                return True
            else:
                print('login faild, login msg is wrong')
                print(respond)
                return False
        except Exception as why:
            print('error: ' + str(why))

    def GetZhima(self, inStr):
        """get zhimakaimen's secret
        """
        requestMsg = {
            'source': inStr,
            'customMsg': "catt",
            # 'customMsg' : self.m_userName
        }
        try:
            self.m_opener.open(self.m_site).read().decode()
        except Exception as why:
            print('open site url faild: ' + str(why))
            return None

        # print urllib.urlencode(requestMsg)
        request = urllib.request.Request(
            self.m_site_action,
            urllib.parse.urlencode(requestMsg).encode("utf-8"))
        # print request
        request.add_header('X-Requested-With', 'XMLHttpRequest')
        request.add_header(
            'Referer',
            'http://psh.hikvision.com.cn:8080/sso/page/command?pageName=homeCN'
        )
        request.add_header('Accept',
                           "application/json, text/javascript, */*; q=0.01")
        request.add_header('Content-Type',
                           'application/x-www-form-urlencoded; charset=utf-8')
        request.add_header('Accept-Encoding', "gzip, deflate")
        request.add_header(
            'User-agent',
            'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/7.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.3; .NET4.0C; .NET4.0E)'
        )
        request.add_header('Host', 'psh.hikvision.com.cn:8080')
        # request.add_header('Origin', self.m_origin)
        try:
            jsonOriginal = self.m_opener.open(request).read()
        except Exception as why:
            print('open site_action url faild: ' + str(why))
            return None
        jsonObject = json.loads(jsonOriginal)
        reStr = jsonObject['result']
        reg = re.compile(r'[\n\r]+')
        return reg.sub("", reStr)

class MySShConnect():
    def __init__(self, ip, username, password):
        self.ip = ip
        self.username = username
        self.password = password

    def SSH_client(self):
        recv_buffer = 99999

        try:
            ssh = paramiko.SSHClient()  ##1.创建一个ssh对象
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy(
            ))  # 2.解决问题:如果之前没有，连接过的ip，会出现选择yes或者no的操作，自动选择yes
            ssh.connect(hostname=self.ip,
                        port=22,
                        username=self.username,
                        password=self.password,
                        timeout=30,
                        allow_agent=False,
                        look_for_keys=False)  # 3.连接服务器
            connect_result = True
        except Exception as e:
            # 不能正常连接的server打印信息到控制台
            print(str(e))
            connect_result = False
        return ssh, recv_buffer, connect_result

    def psh_koulin(self, ssh):
        kaimenCmds = ["debug\n", "zhimakaimen\n"]
        koulin = ''
        for cmd in kaimenCmds:
            ssh.send(cmd)
            koulin = ssh.recv(115200).decode("utf-8", "ignore")
            while (not koulin.endswith(': ')) and (
                    not koulin.endswith('# ')) and (len(koulin) < 1024 * 10):
                koulin = koulin + ssh.recv(115200).decode("utf-8", "ignore")
            if (len(koulin) != 0) and ("help" not in koulin):
                break
        return str(koulin)

    def send_cmd(self, ssh, cmd, limit=1024, rate=10, packageNums=100):
        recv_buffer = 65536
        recv_buffer_limit = limit * rate
        try:
            package = 1
            ssh.send(cmd + '\n')
            out = ssh.recv(recv_buffer).decode("utf-8", "ignore")  # recv_buffer
            time.sleep(1)
            while (not out.endswith('# ')) and (
                    len(out) < recv_buffer_limit) and (package < packageNums):
                out = out + ssh.recv(recv_buffer).decode("utf-8", "ignore")
                package += 1
        except Exception as why:
            info = 'Failed to execute SSH command %s: %s .' % (cmd, repr(why))
            return info

        string = str(out)
        return string

    def send_cmd_4G(self, ssh, cmd, limit=1024, rate=10, packageNums=100):
        recv_buffer = 65536
        # recv_buffer_limit = limit * rate
        try:
            # package = 1
            ssh.send(cmd + '\r\n')
            time.sleep(2)
            out = ssh.recv(recv_buffer).decode("utf-8", "ignore")  # recv_buffer
        except Exception as why:
            info = 'Failed to execute SSH command %s: %s .' % (cmd, repr(why))
            return info

        string = str(out)
        return string

    def psh_change_to_ash(self, ssh):
        koulin = self.psh_koulin(ssh)
        print(koulin)
        zhimakaimen = Zhimakaimen("pubyfcs055", "DFp03yg7]")
        if zhimakaimen.Login():
            if len(koulin) > 1000:
                if "enc_string:" in koulin:
                    final_koulin = koulin.split("\n")[-2][11:]
                else:
                    final_koulin = koulin.split("\n")[-2]
            else:
                final_koulin = koulin.split("\n")[1]
            reStr = zhimakaimen.GetZhima(final_koulin)
            if reStr is not None:
                try:
                    reStr = reStr.encode('utf-8').decode("utf-8", "ignore")
                    if reStr is not None:
                        self.send_cmd(ssh, reStr)
                        time.sleep(1)
                        return True, "进入debug模式成功！"
                except Exception as why:
                    return False, repr(why)
        else:
            print('error', "Login failed, please check <config.ini>.")
            return False, "Login failed, please check <config.ini>."

    def zhimakaimeng_debug(self):
        try:
            client, recv_buffer, result = self.SSH_client()
            if not result:
                return False, ("SSH连接失败，请确认是否真的支持SSH或功能是否正常!", "")
            ssh = client.invoke_shell()
            ssh.settimeout(60)
            out = ssh.recv(recv_buffer).decode("utf-8", "ignore")
            while (not out.endswith('# ')):
                out = out + ssh.recv(1024).decode("utf-8", "ignore")
            # print(out)
            if 'psh' in out:
                koulin = self.psh_koulin(ssh)
                print(koulin)
                zhimakaimen = Zhimakaimen("pubyfcs055", "DFp03yg7]")
                if zhimakaimen.Login():
                    if len(koulin) > 1000:
                        if "enc_string:" in koulin:
                            final_koulin = koulin.split("\n")[-2][11:]
                        else:
                            final_koulin = koulin.split("\n")[-2]
                    else:
                        final_koulin = koulin.split("\n")[1]
                    reStr = zhimakaimen.GetZhima(final_koulin)
                    if reStr is not None:
                        try:
                            reStr = reStr.encode('ascii').decode(
                                "utf-8", "ignore")
                            if reStr is not None:
                                self.send_cmd(ssh, reStr)
                                time.sleep(1)
                                return True, (ssh, client)
                            # instring = raw_input("Pass anykey to exit!!!")
                        except Exception as why:
                            return False, (str(why), "")
                else:
                    print('error', "Login failed, please check <config.ini>.")
                    # if os.path.exists('%s\config.ini'%path):
                    #     os.remove('%s\config.ini'%path)
                    return False, ("Login failed, please check <config.ini>.",
                                   "")

            elif 'ash' in out:
                return True, (ssh, client)
        except Exception as why:
            return False, (str(why), "")

class ExcelOp(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]
        self.merged_cell_ranges=self.ws.merged_cells.ranges  # 表示获取合并单元格范围列表

    def set_merged_cell(self, start_row, start_column, end_row, end_column):
        self.ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

    def parser_merged_cell(self, merged_cell_ranges, row, col):
        """
        检查是否为合并单元格并获取对应行列单元格的值。
        如果是合并单元格，则取合并区域左上角单元格的值作为当前单元格的值,否则直接返回该单元格的值
        :param sheet: 当前工作表对象
        :param row: 需要获取的单元格所在行
        :param col: 需要获取的单元格所在列
        :return:
        """
        cell = self.ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
            for merged_range in merged_cell_ranges:  # 循环查找该单元格所属的合并区域
                if cell.coordinate in merged_range:
                    # 获取合并区域左上角的单元格作为该单元格的值返回
                    cell = self.ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        return cell

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.parser_merged_cell(self.merged_cell_ranges, row, column)

        return cell_value.value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.parser_merged_cell(self.merged_cell_ranges, i, column)
            # print("第%s行第%s列：%s" % (index_row, column, cell_.value))
            column_data.append(cell_value.value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.parser_merged_cell(self.merged_cell_ranges, row, i)
            # print("第%s行第%s列：%s" % (index_row, column, cell_.value))
            row_data.append(cell_value.value)
        return row_data

    def openxls_create(self):
        '''
        创建xlsx表格，并且写入数据

        颜色
        COLOR_INDEX = (
            '00000000', '00FFFFFF', '00FF0000', '0000FF00', '000000FF', #0-4
            '00FFFF00', '00FF00FF', '0000FFFF', '00000000', '00FFFFFF', #5-9
            '00FF0000', '0000FF00', '000000FF', '00FFFF00', '00FF00FF', #10-14
            '0000FFFF', '00800000', '00008000', '00000080', '00808000', #15-19
            '00800080', '00008080', '00C0C0C0', '00808080', '009999FF', #20-24
            '00993366', '00FFFFCC', '00CCFFFF', '00660066', '00FF8080', #25-29
            '000066CC', '00CCCCFF', '00000080', '00FF00FF', '00FFFF00', #30-34
            '0000FFFF', '00800080', '00800000', '00008080', '000000FF', #35-39
            '0000CCFF', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF', #40-44
            '00FF99CC', '00CC99FF', '00FFCC99', '003366FF', '0033CCCC', #45-49
            '0099CC00', '00FFCC00', '00FF9900', '00FF6600', '00666699', #50-54
            '00969696', '00003366', '00339966', '00003300', '00333300', #55-59
            '00993300', '00993366', '00333399', '00333333',  #60-63
        )
        BLACK = COLOR_INDEX[0]
        WHITE = COLOR_INDEX[1]
        RED = COLOR_INDEX[2]
        DARKRED = COLOR_INDEX[8]
        BLUE = COLOR_INDEX[4]
        DARKBLUE = COLOR_INDEX[12]
        GREEN = COLOR_INDEX[3]
        DARKGREEN = COLOR_INDEX[9]
        YELLOW = COLOR_INDEX[5]
        DARKYELLOW = COLOR_INDEX[19]

        字体
        ws.cell(5,3).value='哈哈哈'
        ws.cell(5,3).font = Font(name='仿宋',size=12,color=Color(index=0),b=True,i=True)

        # size   sz  字体大小
        # b bold  是否粗体
        # i italic  是否斜体
        # name family  字体样式

        # 边框
        Side(style='thin',color=Color(index=0))

        # style可选项
        style = ('dashDot','dashDotDot', 'dashed','dotted',
        'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
        'mediumDashed', 'slantDashDot', 'thick', 'thin')
        #  'medium' 中粗
        #  'thin'  细
        #  'thick'  粗
        #  'dashed'  虚线
        #  'dotted'  点线

        # 填充
        PatternFill(patternType='solid',fgColor=Color(), bgColor=Color())
        # fgColor   前景色
        # bgColor   后景色
        # 参数可选项
        patternType = {'darkDown', 'darkUp', 'lightDown', 'darkGrid', 'lightVertical',
                       'solid', 'gray0625', 'darkHorizontal', 'lightGrid', 'lightTrellis',
                       'mediumGray', 'gray125', 'darkGray', 'lightGray', 'lightUp',
                       'lightHorizontal', 'darkTrellis', 'darkVertical'}ws.cell(3,3).fill = PatternFill()

        # 对齐
        Alignment(horizontal='fill',vertical='center')

        # 参数可选项
        horizontal = {'fill', 'distributed', 'centerContinuous', 'right',
                      'justify', 'center', 'left', 'general'}

        vertical = {'distributed', 'justify', 'center', 'bottom', 'top'}

        ws.cell(3,3).alignment= Alignment()

        # 行高，列宽
        row = ws.row_dimensions[1]
        row.height = 15
        col = ws.column_dimensions[1]
        col.width = 10
        '''

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            # cell = self.ws.cell(row=row, column=colunm)
            # if isinstance(cell, MergedCell):
            #     cell.value=cellvalue
            #
            # else:
            # self.ws.merged_cell_ranges
            self.ws.cell(row=row, column=colunm).alignment = styles.Alignment(horizontal='center', wrapText=True,
                                                                              vertical='center')
            self.ws.cell(row=row, column=colunm).font = styles.Font(name='仿宋', size=12, color=styles.Color(index=0),
                                                                    b=True)
            # self.ws.cell(row=row, column=colunm).border=styles.Border(left=styles.Side(border_style='thin', color='000000'),
            #     right=styles.Side(border_style='mediumDashed', color='000000'),
            #     top=styles.Side(border_style='double', color='000000'),
            #     bottom=styles.Side(border_style='dashed', color='000000'))
            self.ws.cell(row=row, column=colunm).border = styles.Border(
                left=styles.Side(border_style='thin', color='000000'),
                right=styles.Side(border_style='thin', color='000000'),
                top=styles.Side(border_style='thin', color='000000'),
                bottom=styles.Side(border_style='thin', color='000000'))
            self.ws.cell(row=row, column=colunm).value = cellvalue
            # self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).alignment = styles.Alignment(wrapText=True, vertical='center')
            self.ws.cell(row=row, column=colunm).font = styles.Font(name='仿宋', size=12, color=styles.Color(index=0),
                                                                    b=True)
            self.ws.cell(row=row, column=colunm).border = styles.Border(
                left=styles.Side(border_style='thin', color='000000'),
                right=styles.Side(border_style='thin', color='000000'),
                top=styles.Side(border_style='thin', color='000000'),
                bottom=styles.Side(border_style='thin', color='000000'))
            self.ws.cell(row=row, column=colunm).value = "writefail"
            # self.wb.save(self.file)

    def saveExc(self):
        self.wb.save(self.file)
        self.wb.close()

class MtdComFun(Integ_Isapi):
    def __init__(self, ip = None, username = 'admin', password = 'abcd1234'):
        super().__init__()
        self.ip = ip
        self.username = username
        self.password = password

    def mtdComFun_setInit(self, username, password):
        # 设备账号
        self.username = username
        # 设备密码
        self.password = password

    # 调用结果处理函数
    def resHandleCallFun(self, handleFun, compFun, actual_data, true_data, delStr = None, sep = None):
        '''调用【结果处理】
        :param
            handleFun: 结果处理函数
            compFun: 比对函数
            actual_data: 设备获取的实际能力集
            true_data: sepc真值
            delStr: list格式，需要对【spec以及真值中的字符串】进行删除的字符串（有的字符串删除后方便比对）
        '''
        return handleFun(actual_data, true_data, compFun, delStr, sep)

    # 结果处理-通用方法1
    def resHandleGeneral_1(self, actual_data, true_data, compFun, delStr=None, sep = None):
        '''以单一分隔符可分开的字符串, 分隔符类型：[',', '/', '|']
        :param
            actual_dat: 设备获取的实际值
            true_data: sepc真值
            compFun: 比对函数
            delStr: 需要在spec中函数的字符串（有的字符串删除后方便比对）
        :return
            'P'/'F'
        '''

        true_data =self.data2Set(true_data, delStr, sep)
        actual_data = self.data2Set(actual_data, delStr, sep)
        return compFun(true_data, actual_data)


    # 将字符串转为Set
    def data2Set(self, data, delStr=None, sep = None):
        if data is None:
            return ''
        if delStr:
            data = self.delSepcificStr(data, delStr)
        sep_list = sep if sep is not None else [',', '，', '/', '|', '、', '||']
        for sep in sep_list:
            ac_list = data.split(sep)
            if len(ac_list) > 1:
                break
        return set([item.strip() for item in ac_list])

    # 结果处理-通用方法2
    def resHandleGeneral_2(self, actual_data, true_data, compFun, delStr, sep = None):
        '''以单一分隔符可分开的字符串, 分隔符类型：[',', '/', '|']
        :param
            actual_dat: 设备获取的实际值
            true_data: sepc真值
            compFun: 比对函数
            delStr: 需要在spec中函数的字符串（有的字符串删除后方便比对）
        :return
            'P'/'F'
        '''
        if delStr:
            true_data = self.delSepcificStr(true_data, delStr)
            actual_data = self.delSepcificStr(actual_data, delStr)

        actual_data = self.data2NestDict(actual_data)
        true_data = self.data2NestDict(true_data)
        return compFun(actual_data, true_data)

    # 处理data为嵌套字典
    def data2NestDict(self, s):
        dict1 = AutoVivification()
        flag, Msg = self.ifKeyOrRecursion(s)

        # 兼容格式中存在';；'作为代替\n分隔符的书写格式
        res_list = []
        for sub in s.strip('\n').strip('\r').split('\n'):
            for i, sep in enumerate([';', '；']):
                subList = sub.split(sep)
                if len(subList) > 1:
                    res_list += subList
                    break
                else:
                    if i == (len([';', '；']) - 1):
                        res_list.append(sub)

        # 过滤list中为''的元素
        res_list = list(filter(lambda x: x != '', res_list))

        if flag:
            for res in res_list:
                if res != '':
                    res = res.replace('_x000D_', '').replace('\r', '')
                    tmp = self.splitStr(res, sep = [':', '：'])[0]
                    if tmp not in Msg:
                        try:
                            print(key)
                        except NameError:
                            key = ''
                        self.data2NestDictBackFun(dict1, key, res)
                    else:
                        key = re.sub('[:：\s]', '', tmp)
        else:
            for i, res in enumerate(res_list):
                if res != '':
                    res = res.replace('_x000D_', '').replace('\r', '')
                    tmp = self.splitStr(res, sep = [':', '：'])
                    if len(tmp) > 1:
                        self.data2NestDictBackFun(dict1, tmp[0], tmp[1])
                    else:
                        dict1[i] = self.data2Set(res)

        return dict1

    # 处理data为嵌套字典调用的递归函数
    def data2NestDictBackFun(self, dict, key, s):
        '''# 处理data为嵌套字典调用的递归函数'''
        Len = len(re.findall('[:：(（]',s))
        if Len == 1:
            tmp = self.splitStr(s, sep = [':','：','(', '（'])
            k, v = tmp[0].strip(), tmp[1].strip()
            v = re.sub('[)）]*', '', v)
            tmp = self.splitStr(v, sep = [',', '，', '/', '|', '、'])
            tmp = [i.strip() for i in tmp]
            dict[key][k] = set(tmp)
            return dict
        elif Len == 0:
            tmp = self.splitStr(s, sep = [',', '，', '/', '|', '、'])
            tmp = [i.strip() for i in tmp]
            dict[key] = set(tmp)
            return dict
        else:
            tmp = self.splitStr(s, sep = [':','：','(','（'])
            dict[key] = self.data2NestDictBackFun(dict[key], tmp[0], tmp[1])

    # 判断参数中的某一行为单纯key还是需要进行递归处理的行
    def ifKeyOrRecursion(self, line):
        '''# 判断参数中的某一行为单纯key还是需要进行递归处理的行'''
        line = line.replace('_x000D_', '')
        tmp_flag = re.findall('.+[:：][\r\n]', line)
        if tmp_flag:
            startStr = tmp_flag[0][0]
            pattern = "\\" + startStr + ".+[:：][\s]*[\r\n]"
            flag = re.findall(pattern, line)
        else:
            flag = tmp_flag

        if len(flag) == 0:
            return False, 0
        else:
            flag = [re.sub('[:：]', '', i.strip('\n').strip('\r').strip(' ')) for i in flag]
            return True, flag

    # 字符串分割
    def splitStr(self, line, sep = None):
        '''字符串分割'''
        line = re.sub('，', ',', line)
        line = re.sub('×', 'x', line)
        sep = [',', '，', '/', '|', '、'] if not sep else sep
        for i in sep:
            tmp = line.split(i)
            if len(tmp) > 1:
                break
        return tmp

    # 删除特定字符串
    def delSepcificStr(self, raw_str, del_str):
        '''删除特定字符串'''
        # pattern = '[(' + ')('.join(del_str) + ')]'
        # raw_str = re.sub(pattern, '', raw_str)

        for i in del_str:
            raw_str = re.sub(i, '', raw_str)
        return raw_str

    # 结果处理-通用方法3
    def resHandleGeneral_3(self, actual_data, true_data, compFun, delStr, sep=None):
        '''以给定的模式提取字符串'''
        if delStr:
            true_data = self.delSepcificStr(true_data, delStr)
            actual_data = self.delSepcificStr(actual_data, delStr)

        if not sep:
            return compFun(actual_data, true_data)
        else:
            pattern = sep
            ac = re.findall(pattern, str(actual_data))
            tr = re.findall(pattern, str(true_data))
            if len(ac) == 0 or len(tr) == 0:
                return compFun(actual_data, true_data)
            else:
                return compFun(ac, tr)

    # 结果比对-通用方法1
    def resCompGeneral_1(self, actural_data, true_data):
        '''以【对称差集】的方法对结果进行比对'''
        set1 = set(actural_data)
        set2 = set(true_data)
        dif = set1 ^ set2
        return 'F' if len(dif) > 0 else 'P'

    # 结果比对-通用方法2
    def resCompGeneral_2(self, actual_data, true_data):
        '''对嵌套字典的结果进行递归比对'''
        return 'P' if self.NestDictCompBackFun(actual_data, true_data) else 'F'

    # 结果比对-通用方法3
    def resCompGeneral_3(self, actual_data, true_data):
        '''直接对字符串进行比对'''
        actual_data = actual_data if isinstance(actual_data, str) else ''.join(actual_data)
        true_data = true_data if isinstance(true_data, str) else ''.join(true_data)
        return 'P' if actual_data == true_data else 'F'

    # 结果比对-嵌套字典递归比对
    def NestDictCompBackFun(self, data1, data2):
        '''嵌套字典递归比对'''
        if isinstance(data1, set) and isinstance(data2, set):
            dif = data1 ^ data2
            print('dif set:', dif)
            return data1 == data2
        elif isinstance(data1, dict) and isinstance(data2, dict):
            if data1.keys() == data2.keys():
                key_index_list = list(data1.keys())
                # 递归
                for i in range(len(data1)):
                    if not self.NestDictCompBackFun(data1[key_index_list[i]], data2[key_index_list[i]]):
                        return False
                # for i, j in zip(data1, data2):
                #     if not self.NestDictCompBackFun(data1[i], data2[j]):
                #         return False
            else:
                return False
        else:
            return False
        return True


    def getDeviceInfo(self):
        '''获取设备基本信息'''
        try:
            dev_info = {}
            page = self.get(f'/ISAPI/System/Video/inputs/channels')
            root = etree.XML(str(page).encode('utf-8'))
            if '</name>' in str(page):
                channel_name = root.xpath("//*[local-name()='name']")
                dev_info["chalNum"] = len(channel_name)
                # dev_info[f'channel1_name']= channel_name
                # dev_info[f'channel1_name'] = channel_name
                # dev_info[f'channel{channel}_name'] = channel_name

            page = self.get(f'/ISAPI/System/deviceInfo')
            root = etree.XML(str(page).encode('utf-8'))

            serialNumber = root.xpath("//*[local-name()='serialNumber']")[0].text if root.xpath(
                "//*[local-name()='serialNumber']") else ""
            # dev_info['serialNumber']=serialNumber
            dev_info.update({"serialNumber": serialNumber})
            if '</model>' in str(page):
                dev_model = root.xpath("//*[local-name()='model']")[0].text
                if not dev_model:
                    dev_model = ''
                # dev_info['dev_model']=self.dev_model
                dev_info['dev_model'] = dev_model

            if '</deviceName>' in str(page):
                device_Name = root.xpath("//*[local-name()='deviceName']")[0].text
                if not device_Name:
                    device_Name = ''
                # dev_info['deviceName']=self.dev_model
                dev_info['deviceName'] = device_Name
            if '</deviceDescription>' in str(page):
                deviceDescription = root.xpath("//*[local-name()='deviceDescription']")[0].text
                if 'IPD' in deviceDescription:
                    # dev_info['dev_type'] = 'IPD'
                    dev_type = 'IPD'

                elif 'IPC' in deviceDescription:
                    # dev_info['dev_type'] = 'IPC'
                    dev_type = 'IPC'
                elif 'IPZ' in deviceDescription:
                    # dev_info['dev_type'] = 'IPZ'
                    dev_type = 'IPZ'
                else:
                    # dev_info['dev_type'] = deviceDescription
                    dev_type = deviceDescription
                dev_info['dev_type'] = dev_type
                dev_info['deviceDescription'] = deviceDescription
            elif '</deviceType>' in str(page):
                dev_type = root.xpath("//*[local-name()='deviceType']")[0].text
                # dev_info['dev_type']=self.dev_type
                dev_info['dev_type'] = dev_type
                dev_info['deviceType'] = dev_type

            if '</firmwareVersionInfo>' in str(page):
                firmwareVersionInfo = root.xpath("//*[local-name()='firmwareVersionInfo']")[0].text
                # print(firmwareVersionInfo)
                # print(firmwareVersionInfo.upper().split('-'))
                dev_platform = firmwareVersionInfo.upper().split('-')[2]
                # dev_info['dev_platform']=self.dev_platform
                dev_info['dev_platform'] = dev_platform
                dev_info.update({"firmwareVersionInfo": firmwareVersionInfo})
                # dev_info['firmwareVersionInfo']=firmwareVersionInfo
            else:
                dev_info['dev_platform'] = ''
            if '</firmwareVersion>' in str(page):
                firmwareVersion = root.xpath("//*[local-name()='firmwareVersion']")[0].text
                # dev_info['firmwareVersion'] = firmwareVersion
                dev_info['firmwareVersion'] = firmwareVersion
            else:
                dev_info['firmwareVersion'] = ''
            if '</firmwareReleasedDate>' in str(page):
                firmwareReleasedDate = root.xpath("//*[local-name()='firmwareReleasedDate']")[0].text
                dev_info['firmwareReleasedDate'] = firmwareReleasedDate
                # dev_info['firmwareReleasedDate'] = firmwareReleasedDate
            else:
                dev_info['firmwareReleasedDate'] = ''
            if dev_info['firmwareReleasedDate'] != '' and dev_info['firmwareVersion'] != '':
                dev_info['dev_softwareVersion'] = f"{dev_info['firmwareVersion']} {dev_info['firmwareReleasedDate']}"
                # dev_info['softwareVersion'] = dev_info['softwareVersion']
            # print(dev_info)
            return True, dev_info
        except Exception as e:
            return False, str(e)

    def getMsgResult_Test(self, dics):
        '''格式化信息和结果用于判断执行用例是否成功'''
        msg = ''
        result = PY_RUN_CORRECT
        list_dic = []
        islistFlag = 0
        for key, value in dics.items():
            list_dic.append(value["result"])
            if isinstance(value["msg"], list):
                islistFlag = 1
        # list_dic=[value for value in dics.values()]
        if not islistFlag:
            print(list_dic)
            if len(set(list_dic)) == 1:
                if list_dic[0] == "PY_RUN_CORRECT":

                    msg = ";".join([value["msg"] for key, value in dics.items()])
                    result = PY_RUN_CORRECT
                    return msg, result
                elif list_dic[0] == 'PY_RUN_NA':
                    msg = "设备所有通道功能不支持"
                    result = PY_RUN_NA
                    return msg, result
            for key, value in dics.items():
                msg += value["msg"] + ";"
                if value['result'] == 'PY_RUN_WRONG':
                    result = PY_RUN_WRONG
        return msg, result

    @staticmethod
    def VD_LOG_DEBUG(ip, *args, level = 'ERROR', sep=' ', end='\n', file=None):
        '''公共打印log'''

        # init_path = os.getcwd()
        # # 获取被调用函数在被调用时所处代码行数
        # line = sys._getframe().f_back.f_lineno
        # # 获取被调用函数所在模块文件名
        # file_name = sys._getframe(1).f_code.co_filename
        # file_name = file_name.replace("\\", "/").split("/")[-1]
        # # sys.stdout.write(f'"{__file__}:{sys._getframe().f_lineno}"  {x}\n')
        # args = (str(arg) for arg in args)  # REMIND 防止是数字不能被join
        # args_str = " ".join(args)
        # # 打印到标准输出，并设置文字和背景颜色 sys.stdout.write(f'"{file_name}:{line}" {time.strftime("%H:%M:%S")} \033[0;94m{"".join(
        # # args)}\033[0m\n') # 36 93 96 94
        #
        # # time_seq = time.strftime('[%Y-%m-%d %H:%M:%S]', time.localtime(time.time()))
        # # init_path=os.path.abspath(os.path.dirname(__file__))
        # # init_path=os.getcwd()
        # try:
        #     os.makedirs(f"log\\{ip}")
        # except FileExistsError:
        #     pass
        # log_path_name = f"%s\\log\\{ip}\\log_%s_%s.txt" % (init_path, ip, time_seq)
        # info = f'[{file_name}:{line}]{" " + args_str}\n'
        # info = info.replace("\r\n", "\n")
        # info_log = time_seq + info
        # info_log = info_log.encode("gbk", errors="ignore").decode("gbk")
        # with open(log_path_name, "a+") as f:
        #     f.write(info_log)
        #     sys.stdout.write(info)
        #     f.close()

        current_date = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        init_path = os.getcwd()
        LogPath = os.path.join(init_path, "Log", current_date, ip)
        if not os.path.exists(LogPath):
            os.makedirs(LogPath)
        # 执行日志文件
        testLogFile = os.path.join(LogPath, f"testLog_{ip}_{current_date}.txt")

        # 获取被调用函数在被调用时所处代码行数
        line = sys._getframe().f_back.f_lineno
        # 获取被调用函数所在模块文件名
        file_name = sys._getframe(1).f_code.co_filename
        file_name = file_name.replace("\\", "/").split("/")[-1]
        args_str = ' '.join(str(arg) for arg in args)

        try:
            testLogObj = open(testLogFile, "a+", encoding='utf-8')
            # info_log = f"[{str(datetime.datetime.now())[:19]} {ip} {file_name} {line}] [{level}]{args_str}"
            # info_log = info_log.encode("gbk", errors="ignore").decode("gbk")
            print(f"[{str(datetime.datetime.now())[:19]} {ip} {file_name} {line}] [{level}]{args_str}")
            print(f"[{str(datetime.datetime.now())[:19]} {ip} {file_name} {line}] [{level}]{args_str}", file=testLogObj)
            testLogObj.close()
        except Exception as e:
            testLogObj = open(testLogFile, "a+")
            print(traceback.format_exc(), file=testLogFile)
            print(f"[ERROR]将执行打印写入日志文件出现异常:{str(e)}!", file=testLogFile)
            testLogObj.close()

    def get_stream_supportChannelNum(self, a_list):
        '''获取码流ID'''
        one_list = 0
        two_list = 0
        three_list = 0
        four_list = 0
        five_list = 0
        # a_list = ["101", "102", "103", "104", "105", "201", "202", "203", "204"]
        for item in a_list:
            if "1" in item.split("0")[1]:
                one_list += 1
            if "2" in item.split("0")[1]:
                two_list += 1
            if "3" in item.split("0")[1]:
                three_list += 1
            if "4" in item.split("0")[1]:
                four_list += 1
            if "5" in item.split("0")[1]:
                five_list += 1
        return {
            "1": one_list,
            "2": two_list,
            "3": three_list,
            "4": four_list,
            "5": five_list
        }

    def getReceiveDicParas(self, evType, relationResuleList):
        """获取关联智能最多的组合
            找到每组中最多的指定智能的组合
        """
        a = {}
        for index, item in enumerate(relationResuleList):
            c = 0
            for b in item.values():
                if evType in b:
                    c += 1
            a[index] = c
        d = [int(i) for i in a.values()]
        maxInTypeValue = sorted(d)[-1]
        # print(sorted(d)[-1])
        for c, db in a.items():
            if int(db) == maxInTypeValue:
                return True, relationResuleList[int(c)]
            else:
                pass
        return False, "获取存在最多指定智能关联组合失败"

    def device_is_overseas_judge(self):
        '''设备是否为海外设备判断'''
        try:
            languageUrl = '/SDK/language'
            ret = self.isapi_get(languageUrl, path=[".//type"], type=["value"])
            if not ret['success']:
                return False, f"判断设备是否为海外设备协议执行失败:{ret['msg']}!"

            language = ret['data'][0][0]

            if language not in ['english', 'English']:
                return True, False

            return True, True
        except Exception as e:
            print(traceback.print_exc())
            return False, f"判断设备是否为海外设备出现异常:{str(e)}!"

    def channel_VCAResourse_OpenHEOP_cap_judge(self, channel=1, vcaType='facesnap'):
        """
            开放HEOP智能资源能力判断接口

            参数:
                channel:通道号,从1开始
                vcaType:智能资源名称
            返回:
                return 接口执行成功，[是否是开放HEOP架构，是否支持该智能，智能对应的id号]
                return 接口执行失败，详细报错信息

                成功:
                    return True, [True, True, 1]
                失败
                    return False, “详细失败报错信息”
        """
        try:
            # 开放HEOP能力判断
            openHeopCapUrl = "/ISAPI/Custom/OpenPlatform/capabilities"
            openHeopCap = self.get(openHeopCapUrl)
            print(openHeopCap)

            if "notSupport" in openHeopCap or "invalidOperation" in openHeopCap:
                return True, [False, False, 0]

            if "404 -- Not Found" in openHeopCap:
                return True, [False, False, 0]

            if "</OpenPlatformCap>" not in openHeopCap:
                return False, f"通道/ISAPI/Custom/OpenPlatform/capabilities接口判断设备是否为开放HEOP平台设备协议执行失败:{self.get_subStatusCode(openHeopCap)}!"

            if "</isSupportOpenPlatform>" not in openHeopCap:
                return True, [False, False, 0]

            isSupportOpenPlatform = re.findall(
                "<isSupportOpenPlatform>(.*?)</isSupportOpenPlatform>",
                openHeopCap, re.S)[0]

            if isSupportOpenPlatform == "false":
                return True, [False, False, 0]

            # 人脸比对和混合目标比对暂不测试
            if vcaType == "faceContrast" or vcaType == "faceHumanModelingContrast":
                return True, [True, False, 0]

            # 判断是否是多通道的开放HEOP架构，目前暂不测试多通道的设备
            if "</AppChannels>" in openHeopCap:
                return True, [True, False, 0]

            # 判断设备是否为海外设备
            isOverseasFlag, isOverseasMsg = self.device_is_overseas_judge()
            if not isOverseasFlag:
                return False, isOverseasMsg

            # 普通监控在开放HEIOP架构中不存在
            if vcaType == "close":
                return True, [True, False, 0]

            nameList = HEOP_APP_EN[
                vcaType] if isOverseasMsg else HEOP_APP_ZH[vcaType]

            # 获取设备当前已加载所有智能资源
            openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
            openHeopData = self.get(openHeopUrl)
            print(openHeopData)
            if "</AppList>" not in openHeopData:
                return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{self.get_subStatusCode(openHeopData)}!"

            if "</App>" not in openHeopData:
                return True, [True, False, 0]

            idList = re.findall("<id>(.*?)</id>", openHeopData, re.S)
            packageNameList = re.findall("<packageName>(.*?)</packageName>",
                                         openHeopData, re.S)

            print(f"packageNameList={packageNameList}")

            if len(idList) == 0 or len(packageNameList) == 0:
                return False, "通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表失败!"

            for i in nameList:
                if i not in packageNameList:
                    return True, [True, False, 0]

            packageName = nameList[0]
            packageName_id = {}
            for i in range(len(idList)):
                packageName_id[packageNameList[i]] = idList[i]

            print(packageName_id)

            index = packageName_id[packageName]

            return True, [True, True, index]

        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备通道{channel}判断智能资源能力出现异常:{str(e)}!"

    def channel_VCAResourse_cap_judge(self, channel=1, vcaType='facesnap'):
        """
            智能资源能力判断接口

            参数:
                channel:通道号,从1开始
                vcaType:智能资源名称
            返回:
                return 接口执行成功，[是否有智能资源概念(针对smart事件)，是否支持该智能，是否是关联智能关系, 是否是开放HEOP架构]
                return 接口执行失败，详细报错信息

                成功:
                    return True, [True, True, isRelative, True]
                失败
                    return False, “详细失败报错信息”
        """
        try:
            # 是否为关联智能资源标志
            isRelative = False

            # 是否为开放HEOP标准
            openHeopFlag, openHeopMsg = self.channel_VCAResourse_OpenHEOP_cap_judge(channel, vcaType)
            if not openHeopFlag:
                return False, openHeopMsg

            if openHeopMsg[0]:
                if not openHeopMsg[1]:
                    return True, [True, False, isRelative, True, 0]
                else:
                    return True, [True, True, isRelative, True, openHeopMsg[2]]

            # 关联智能资源能力
            relatedVCAResourceCapUrl = "/ISAPI/System/RelatedVCAResource/capabilities?format=json"
            relatedVCAResourceCap = self.get(relatedVCAResourceCapUrl)
            print(relatedVCAResourceCap)

            # 非关联智能资源能力
            vcaResourceCapUrl = f"/ISAPI/System/Video/inputs/channels/{channel}/VCAResource/capabilities"
            vcaResourceCap = self.get(vcaResourceCapUrl)
            print(vcaResourceCap)

            # 所有支持的智能资源能力列表
            resultCapList = []

            # # 设备为开放HEOP设备
            # if "<isSupportOpenPlatform>true</isSupportOpenPlatform>" in openHeopCap:
            #     return False, "当前设备为开放HEOP设备，目前工具暂不支持开放HEOP设备的APP有效性测试，请手动测试!"
            # 不存在智能应用概念
            if ("notSupport" in relatedVCAResourceCap
                or "invalidOperation" in relatedVCAResourceCap) and (
                    "notSupport" in vcaResourceCap
                    or "invalidOperation" in vcaResourceCap):
                if vcaType == "smart":
                    smartUrl = "/ISAPI/Smart/capabilities"
                    ret = self.get(smartUrl)
                    if "notSupport" in ret or "invalidOperation" in ret:
                        return True, [False, False, isRelative]
                    if "</SmartCap>" not in ret:
                        return False, f"设备通道{channel}获取Smart智能资源能力失败:{self.get_subStatusCode(ret)}!"
                    return True, [False, True, isRelative, False]
                else:
                    return True, [False, False, isRelative, False]
            # 关联智能模式
            elif ("notSupport" not in relatedVCAResourceCap
                  and "invalidOperation" not in relatedVCAResourceCap
            ) and "RelatedVCAResourceCap" in relatedVCAResourceCap:
                isRelative = True
                relationResuleList = []
                capList = json.loads(relatedVCAResourceCap
                                     )["RelatedVCAResourceCap"]["ResourceList"]
                for item in capList:
                    tempDict = {}
                    for key, value in item.items():
                        if key == "MainChannelResource":
                            MainChannelId = value['channelID']
                            MainChannelType = value['resourceType']
                            tempDict[str(MainChannelId)] = MainChannelType
                        elif key == "RelatedChannelsResource":
                            for sonChannel in value:
                                sonChannelId = sonChannel['channelID']
                                sonChannelType = sonChannel['resourceType']
                                tempDict[str(sonChannelId)] = sonChannelType
                    relationResuleList.append(tempDict)
                print(relationResuleList)
                for item in relationResuleList:
                    temp = item[str(channel)]
                    if isinstance(temp, list):
                        resultCapList += temp
                    elif temp not in capList:
                        resultCapList.append(temp)
                resultCapList = list(set(resultCapList))
            # 非关联智能模式
            elif ("notSupport" not in vcaResourceCap
                  and "invalidOperation" not in vcaResourceCap
            ) and "</VCAResource>" in vcaResourceCap:
                resultCapList = re.findall('<type opt="(.*?)">',
                                           vcaResourceCap, re.S)[0].split(',')
            else:
                return False, f"设备通道{channel}获取智能资源能力失败!"

            print(resultCapList)

            if len(resultCapList) != 0 and vcaType in resultCapList:
                return True, [True, True, isRelative, False]
            else:
                return True, [True, False, isRelative, False]
        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备通道{channel}判断智能资源能力出现异常:{str(e)}!"

    def channel_Intelligent_VCAResource_change_isRebootRequired(self, channel=1, vcaType='facesnap'):
        '''
            判断切换智能是不是需要重启操作接口

            参数:
                channel:通道号,从1开始
                vcaType:智能资源名称
            返回:
                return 接口执行成功，是否需要重启
                return 接口执行失败，详细报错信息

                成功:
                    return True, True
                    return True, False
                失败:
                    return False, “出现异常”
        '''
        try:
            rebootRequiredUrl = f"/ISAPI/System/Video/inputs/channels/{channel}/VCAResource/test"

            DATA = f'''
                <?xml version="1.0" encoding="UTF-8"?>
                <VCAResource version="2.0" xmlns="http://www.hikvision.com/ver20/XMLSchema">
                    <type>{vcaType}</type>
                </VCAResource>
                '''
            ret = self.post(rebootRequiredUrl, DATA)
            if "</VCAResourceResult>" not in ret:
                return False, f'设备通道{channel}判断切换智能{vcaType}是否需要重启协议执行失败:{self.get_subStatusCode(ret)}!'

            getReboot = re.findall('<reboot>(.*?)</reboot>', ret, re.S)[0]

            if getReboot == "false":
                return True, False

            return True, True
        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备通道{channel}判断切换智能{vcaType}是否需要重启出现异常:{str(e)}!"

    def integ_reboot_device(self, deviceInfo):
        '''判断设备重启和自检是否成功'''
        try:
            # # 重启设备
            ip = deviceInfo["IP"]
            username = deviceInfo["U_NAME"]
            password = deviceInfo["P_WORD"]
            dev_type = deviceInfo["dev_type"]
            url = "http://" + deviceInfo["IP"] + "/ISAPI/System/reboot/"  # 需要发起请求的URL
            url_1 = "http://" + deviceInfo["IP"] + "/ISAPI/System/deviceInfo/"
            auth = HTTPDigestAuth(username=username, password=password)
            res = requests.put(url=url, data=None, auth=auth, timeout=30)
            isRebootFlag = False
            if res.status_code == 200:
                time.sleep(10)
                # 判断是否重启开始
                t = 0
                while (t < 300):
                    try:
                        ret = requests.get(url=url_1, data=None, auth=auth, timeout=10)
                        ##保证设备进入重启状态，当下发重启协议的时候设备需要响应一定时间，否则协议是通的
                        if "</DeviceInfo>" not in ret.text:
                            self.VD_LOG_DEBUG(ip, "设备进入重启状态，下发重启成功，重启开始=======================")
                            isRebootFlag = True
                            break
                        else:
                            self.VD_LOG_DEBUG(ip, "设备还未进入重启状态，等待开始重启=======================")
                            time.sleep(10)
                            t += 10
                            continue
                    except:
                        self.VD_LOG_DEBUG(ip, "设备进入重启状态，下发重启成功，重启开始=======================")
                        isRebootFlag = True
                        break
            if isRebootFlag:
                t = 0
                while (t < 300):
                    try:
                        ret = requests.get(url=url_1, data=None, auth=auth, timeout=10)
                        print(ret)
                        if "</DeviceInfo>" not in ret.text:
                            self.VD_LOG_DEBUG(ip, "设备进入重启状态，等待重启成功=======================")
                            time.sleep(10)
                            t += 10
                        else:
                            self.VD_LOG_DEBUG(ip, "重启成功=======================")
                            return True, "重启成功"
                    except:
                        self.VD_LOG_DEBUG(ip, "设备进入重启状态，等待重启成功=======================")
                        time.sleep(10)
                        t += 10
                        continue
                # return False,"设备进入重启状态，但是等待五分钟没有重启成功，请查看设备是否有问题"


            else:
                return False, f"下发设备重启后五分钟设备依然未进入重启状态，出现异常!"

            return False, "设备进入重启状态，但是等待五分钟没有重启成功，请查看设备是否有问题"
        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备重启出现异常:{str(e)}!"

    def OpenPlatform_App_runStatus_judge(self, vcaType='facesnap'):
        """智能是否已经起来判断"""
        App_corresponding = {
            'basicBehavior':
                ['基本行为', 'basicBehavior', "行为分析", 'Behavior Analysis '],
            'fullBehavior':
                ['所有行为', 'fullBehavior', '行为分析', 'Behavior Analysis '],
            'facesnapBehavior': [
                '人脸抓拍行为', 'facesnapBehavior', '行为分析 + 人脸抓拍', "行为分析", "人脸抓拍",
                'Behavior Analysis  + Face Capture ',
                'Behavior Analysis  + Face Recognition', 'Behavior Analysis',
                'Face Capture', 'Face Recognition'
            ],
            'facesnap':
                ['人脸抓拍', 'facesnap', 'Face Capture', 'Face Recognition'],
            'TFS': ['交通取证检测', 'TFS', '违章事件', 'Traffic Violations '],
            'ITS': ['ITS', '智能交通', 'ITS(Intelligent Transportation System)'],
            'smartVehicleDetection': [
                '简易智能&&车辆检测', 'smartVehicleDetection', 'Smart事件 + 车辆检测',
                '车辆检测', 'SMART Event + Vehicle Detection', 'Vehicle Detection'
            ],
            'smartHVTDetection': [
                '简易智能&&混行检测', 'smartHVTDetection', 'Smart事件 + 混行检测',
                'SMART Event + Mixed-traffic Detection',
                'Mixed-traffic Detection', 'SMART Event'
            ],
            'smart': ['简易智能', 'smart', 'Smart Event', 'Smart事件'],
            'judicial': ['司法', 'judicial', '司法事件', 'Judicial Events'],
            'smart264AndRoadDetection': [
                'Smart264事件+道路监控', 'smart264AndRoadDetection',
                'Smart事件 + 道路监控', '道路监控', 'SMART Event + Road Traffic',
                'Road Traffic'
            ],
            'smart264AndFaceDetection': [
                'Smart264事件+人脸侦测', 'smart264AndFaceDetection',
                'Smart事件 + 人脸侦测', '人脸侦测', 'SMART Event + Face Detection',
                'Face Detection'
            ],
            'smart264AndHeatMap': [
                'Smart264事件+热度图', 'smart264AndHeatMap', 'Smart事件 + 热度图', '热度图',
                'SMART Event + Heat Map', 'Heat Map'
            ],
            'smartIntelligentMonitor': [
                '简易智能&&车辆违法停车检测', 'smartIntelligentMonitor', 'Smart事件 + 智慧监控',
                '智慧监控', 'SMART Event + Intelligent Monitoring',
                'Intelligent Monitoring'
            ],
            'smartVehicleIllegalParkingDetection': [
                'Smart事件 + 车辆检测 + 违停检测', 'smartVehicleIllegalParkingDetection',
                '违停检测', '车辆检测',
                'SMART Event + Vehicle Detection + Illegal Parking Detection',
                'Vehicle Detection', 'Illegal Parking Detection'
            ],
            'smartTrafficDataCollection': [
                'Smart事件+智慧监控', 'smartTrafficDataCollection',
                'Smart事件 + 交通数据采集', '交通数据采集',
                'SMART Event + Traffic Data Statistics',
                'Traffic Data Statistics'
            ],
            'roadDetection': ['道路检测', 'roadDetection', '道路监控', 'Road Traffic'],
            'humanRecognition': ['人体目标识别', 'humanRecognition'],
            'perimeterCapture':
                ['周界抓拍', 'perimeterCapture', 'Perimeter Capture'],
            'vehicleDetection': ['车辆检测', 'vehicleDetection'],
            'hvtVehicleDetection':
                ['混行检测', 'HVTDetection', 'hvtVehicleDetection'],
            'mixedTargetDetection': [
                '混合目标检测', 'mixedTargetDetection', '全结构化',
                'Capture Target With Feature'
            ],
            'trackingCaptureMode': ['全景抓拍模式', 'trackingCaptureMode'],
            'nonTrackingCaptureMode': ['单点抓拍模式', 'nonTrackingCaptureMode'],
            'close': ['关闭该通道下所有智能资源', 'close', '普通监控', 'Monitoring'],
            'faceHumanModelingContrast': [
                '混合目标', 'faceHumanModelingContrast', "混合目标比对", "比对模式", "人体",
                "人脸", "人体+人脸", 'Comparison Mode',
                'Multi-Target-Type Comparison', '全结构化', '比对建模'
            ],
            'cityManagement':
                ['智慧城管监测', 'cityManagement', '智慧城管', 'City Management'],
            'teacherBehavior': [
                '教师行为检测', 'teacherBehavior', '教师行为分析',
                "Teacher's Behavior Analysis"
            ],
            '12MPLiveView':
                ['主码流1200W', '12MPLiveView', '1200W预览模式', '12 MP Live Views'],
            'personQueueDetection':
                ['人员排队检测', 'personQueueDetection', '区域关注度', 'Queue Management'],
            'verticalPeopleCounting': ['垂直客流', 'verticalPeopleCounting'],
            'AIOpenPlatform':
                ['AI开放平台', 'AIOpenPlatform', '应用仓库', 'AI Open Platform'],
            'mixedTargetFaceHuman': [
                '混合目标抓拍人脸人体模式',
                'mixedTargetFaceHuman',
                '人脸 + 人体',
                '人脸',
                '人体',
                'Face + Human',
                'Face',
                'Human',
            ],
            'mixedTargetAll':
                ['混合目标抓拍全结构化模式', 'mixedTargetAll', 'Capture Target With Feature'],
            'safetyHelmet': ['安全帽检测', 'safetyHelmet'],
            'faceContrast': [
                '人脸比对', 'faceContrast', 'Face Picture Comparison', '人脸抓拍',
                'Face Capture', '比对建模'
            ],
            'faceMonitor': ['人脸布控', 'faceMonitor'],
            'vehicleMonitor': ['车辆布控', 'vehicleMonitor'],
            'faceCounting': ['人脸客流', 'faceCounting'],
            'heelPDC': ['倾斜客流', 'heelPDC', '倾斜客流统计', 'People Counting'],
            'personDensity':
                ['人员密度', 'personDensity', '人员密度检测', "People Density"],
            'behaviorMixedTargetAll':
                ['行为分析混合目标全结构化', 'behaviorMixedTargetAll', '全结构化'],
            'operationMonitor': ['作业监测', 'operationMonitor'],
            'fielddetection': ['区域入侵', 'fielddetection'],
            'personArming': ['人员布控', 'personArming', 'Person Arming'],
            'smokeDetectAlarm': ['抽烟检测', 'smokeDetectAlarm'],
            'smokeDetection': ['烟雾检测', 'smokeDetection'],
            'personDensityQueueLeavePosition': [
                '横幅检测', 'personDensityQueueLeavePosition'
            ],
            'waterQualityDetection': ['水质检测', 'waterQualityDetection'],
            'ATMSurround': ['ATM环境模式', 'ATMSurround'],
            'ATMPanel': ['ATM面板模式', 'ATMPanel'],
            'ATMSafetyCabin': ['ATM防护舱模式', 'ATMSafetyCabin'],
            'intelligentTraffic': ['智能交通', 'intelligentTraffic'],
            'objectsThrownDetection': [
                '高空抛物检测', '高空抛物', 'objectsThrownDetection', 'objectsThrown'
            ],
            'fullBehaviorAndMixedTargetDetection': [
                '行为分析 + 混合目标检测', 'fullBehaviorAndMixedTargetDetection',
                '混合目标检测', '行为分析', 'Multi-Target-Type Detection',
                'Behavior Analysis '
            ]
        }

        try:
            if vcaType == "close":
                time.sleep(10)
                return True, "普通监控无需判断!"

            # 判断设备是否为海外设备
            judgeFlag, isOverseas = self.device_is_overseas_judge()
            if not judgeFlag:
                return False, isOverseas

            url = "/ISAPI/Custom/OpenPlatform/App"
            errMsg = ""
            isUp = False
            for index in range(24):
                ret = self.get(url)

                print(ret)
                if "notSupport" in ret or "invalidOperation" in ret:
                    time.sleep(15)
                    isUp = True
                    break

                if "404 -- Not Found" in ret:
                    time.sleep(15)
                    isUp = True
                    break

                if "</packageName>" not in ret:
                    errMsg = f"设备等待两分钟依然未获取到{vcaType}智能是否起来的信息!"
                    time.sleep(5)
                    continue

                idList = re.findall("<id>(.*?)</id>", ret, re.S)
                packageNameList = re.findall(
                    "<packageName>(.*?)</packageName>", ret, re.S)
                runStatusList = re.findall("<runStatus>(.*?)</runStatus>", ret,
                                           re.S)

                id_status_dict = {}
                for i in range(len(idList)):
                    id_status_dict[idList[i]] = runStatusList[i]
                print(id_status_dict)

                name_status_dict = {}
                for i in range(len(packageNameList)):
                    name_status_dict[packageNameList[i]] = runStatusList[i]
                print(name_status_dict)

                if vcaType == "AIOpenPlatform":
                    pageItemUrl = "/ISAPI/Custom/OpenPlatform/pageItem"
                    ret = self.get(pageItemUrl)
                    if "notSupport" in ret or "invalidOperation" in ret:
                        time.sleep(15)
                        isUp = True
                        break
                    if "</pageItemList>" not in ret:
                        errMsg = "设备等待两分钟依然未获取到智能是否起来的信息!"
                        time.sleep(5)
                        continue

                    idListTemp = re.findall("<id>(.*?)</id>", ret, re.S)
                    appNameListTemp = re.findall("<appName>(.*?)</appName>",
                                                 ret, re.S)

                    print(
                        f"appNameListTemp={appNameListTemp}, idListTemp={idListTemp}"
                    )

                    isFind = False
                    for index in range(len(appNameListTemp)):
                        if appNameListTemp[index] == "aiOpenPlat":
                            status = runStatusList[index]
                            if status == "true":
                                isFind = True
                                break
                            else:
                                continue

                    if not isFind:
                        errMsg = "设备等待两分钟依然未获取到智能是否起来的信息!"
                        time.sleep(5)
                        continue

                    isUp = True
                    break

                nameList = HEOP_APP_EN[
                    vcaType] if isOverseas else HEOP_APP_ZH[vcaType]

                tempFlag = True
                for name in nameList:
                    if name not in packageNameList:
                        errMsg = f"设备等待两分钟依然未获取到{name}是否起来的信息!"
                        tempFlag = False
                        break
                    if name_status_dict[name] != "true":
                        errMsg = f"设备等待两分钟后，{name}的是否起来的状态依然是false!"
                        tempFlag = False
                        break

                if not tempFlag:
                    time.sleep(5)
                    continue

                isUp = True
                break

            if not isUp:
                return False, errMsg

            return True, "智能已经起来!"
        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备判断切换智能{vcaType}后智能是否起来出现异常:{str(e)}!"

    def channel_Intelligent_VCAResource_change(self, deviceInfo,
                                               channelList,
                                               vcaType='facesnap',
                                               isRelative=False,
                                               isOpenHeop=False,
                                               appId=0,
                                               sleepTime=60):
        """
            智能资源切换接口

            参数:
                channel:通道号,从1开始
                vcaType:智能资源名称
                isRelative:是否为关联智能资源
                isOpenHeop:是否为开放HEOP架构
                appId:开放HEOP架构智能对应id
                sleepTime:掉电记忆等待时间
            返回:
                return 接口执行成功，成功信息
                return 接口执行失败，详细报错信息

                成功:
                    return True, “智能资源切换成功”
                失败
                    return False, “详细失败报错信息”
        """
        try:
            ip = deviceInfo["IP"]
            username = deviceInfo["U_NAME"]
            password = deviceInfo["P_WORD"]
            dev_type = deviceInfo["dev_type"]
            needReboot = False
            needRebootFlag, needRebootMsg = self.channel_Intelligent_VCAResource_change_isRebootRequired(1, vcaType)
            if not needRebootFlag:
                print(needRebootMsg)
            else:
                needReboot = needRebootMsg

            if isOpenHeop:
                channel = 1
                # 获取设备当前已加载所有智能资源
                openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
                openHeopData = self.get(openHeopUrl)
                print(openHeopData)
                if "</AppList>" not in openHeopData:
                    return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{self.get_subStatusCode(openHeopData)}!"

                idList = re.findall("<id>(.*?)</id>", openHeopData, re.S)
                packageNameList = re.findall(
                    "<packageName>(.*?)</packageName>", openHeopData, re.S)
                runStatusList = re.findall("<runStatus>(.*?)</runStatus>",
                                           openHeopData, re.S)

                # 先停用所有已启用的智能
                for i in range(len(runStatusList)):
                    if runStatusList[i] == "true":
                        stopUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{idList[i]}/stop"
                        print(f"[PUT {stopUrl}]")
                        ret = requests.put(stopUrl,
                                           data=None,
                                           auth=HTTPDigestAuth(
                                               username, password),
                                           timeout=180)
                        status_code = ret.status_code
                        content = ret.text
                        ret.close()
                        print(content)
                        if "OK" not in content:
                            return False, f"在正式测试目标智能前关闭所有智能时出现关闭{packageNameList[i]}智能协议执行失败:{self.get_subStatusCode(content)}!"
                        time.sleep(5)
                    else:
                        continue

                # 掉电记忆
                # if dev_type == 'IPD' and channel == 1:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)
                # else:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)

                # 配置启用目标智能
                startUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{appId}/start"
                print(f"[PUT {startUrl}]")
                req = requests.put(startUrl,
                                   data=None,
                                   auth=HTTPDigestAuth(username, password),
                                   timeout=300)
                status_code = req.status_code
                content = req.text
                req.close()
                print(content)

                if status_code == 200 and ("rebootRequired" in content
                                           or needReboot):
                    rebootFlag, rebootMsg = self.integ_reboot_device(deviceInfo)
                    if not rebootFlag:
                        return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                    # if dev_type== 'IPD' and channel == 1:
                    #     moveFlag, moveMsg = self.put_device_position_cfg()
                    #     if not moveFlag:
                    #         return False, moveMsg
                    #     time.sleep(3)

                    isUpFlag, isUpMsg = self.OpenPlatform_App_runStatus_judge(vcaType)
                    if not isUpFlag:
                        return False, isUpMsg

                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
                elif status_code == 200 and "OK" in content:
                    isUpFlag, isUpMsg = self.OpenPlatform_App_runStatus_judge(vcaType)
                    if not isUpFlag:
                        return False, isUpMsg

                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
                else:
                    return False, f"设备通道{channel}切换智能资源{vcaType}协议执行失败:{self.get_subStatusCode(content)}!"
            elif not isRelative:
                # 非关联智能
                # 获取当前设备该通道智能资源
                # 可支持的通道
                a_dic = {}
                for channel in channelList:
                    vcaResourceUrl = f'/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
                    ret = self.get(vcaResourceUrl)
                    print(ret)
                    if "</VCAResource>" not in ret:
                        return False, f'设备通道{channel}获取当前智能资源协议执行失败:{self.get_subStatusCode(ret)}!'

                    currentType = re.findall("<type>(.*?)</type>", ret, re.S)[0]
                    print("currentType=", currentType)

                    # 判断当前智能是否为指定智能，如果是跳到下一个通道继续判断
                    if currentType == vcaType:
                        continue

                    # 当前智能资源与需要切换的智能资源比较
                    # if currentType == vcaType:
                    #     if dev_type == 'IPD' and channel == 1:
                    #         print("等待60s保证掉电记忆配置保存!")
                    #         time.sleep(sleepTime)
                    #         moveFlag, moveMsg = self.put_device_position_cfg()
                    #         if not moveFlag:
                    #             return False, moveMsg
                    #         time.sleep(3)
                    #     else:
                    #         print("等待60s保证掉电记忆配置保存!")
                    #         time.sleep(sleepTime)
                    #     return True, f"设备通道{channel}当前智能资源就是{vcaType},无需切换!"

                    # 切换智能资源操作
                    # if dev_type == 'IPD' and channel == 1:
                    #     print("等待60s保证掉电记忆配置保存!")
                    #     time.sleep(sleepTime)
                    #     # moveFlag, moveMsg = self.put_device_position_cfg()
                    #     # if not moveFlag:
                    #     #     return False, moveMsg
                    #     time.sleep(3)
                    # else:
                    #     print("等待60s保证掉电记忆配置保存!")
                    #     time.sleep(sleepTime)

                    vcaResourceUrl = f'http://{ip}/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
                    DATA = f"""
                        <VCAResource xmlns="http://www.hikvision.com/ver20/XMLSchema" version="2.0">
                            <type>{vcaType}</type>
                        </VCAResource>
                        """
                    print(DATA)
                    print(f"[PUT {vcaResourceUrl}]")
                    req = requests.put(vcaResourceUrl,
                                       data=DATA,
                                       auth=HTTPDigestAuth(username, password),
                                       timeout=300)
                    status_code = req.status_code
                    content = req.text
                    req.close()
                    print(content)
                    if status_code == 200:
                        a_dic[str(channel)] = content
                    else:
                        a_dic[str(channel)] = "协议失败"
                if not a_dic:
                    return True, f"设备所有可切换的通道当前智能资源就是{vcaType},无需切换!"
                elif "协议失败" not in a_dic.values():
                    rebootFlag, rebootMsg = self.integ_reboot_device(deviceInfo)
                    if not rebootFlag:
                        return False, f"所有支持的通道切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                    isUpFlag, isUpMsg = self.OpenPlatform_App_runStatus_judge(vcaType)
                    if not isUpFlag:
                        return False, isUpMsg

                    return True, f"所有支持的通道切换智能资源{vcaType}成功,已完成重启!"
                    # if  "rebootRequired" in a_dic.values() or needReboot:
                    #     rebootFlag, rebootMsg = integ_reboot_device(integ_isapi, deviceInfo)
                    #     if not rebootFlag:
                    #         return False, f"所有支持的通道切换智能资源{vcaType}后重启失败:{rebootMsg}!"
                    #
                    #     isUpFlag, isUpMsg = OpenPlatform_App_runStatus_judge(integ_isapi,
                    #         vcaType)
                    #     if not isUpFlag:
                    #         return False, isUpMsg
                    #
                    #     return True, f"所有支持的通道切换智能资源{vcaType}成功,已完成重启!"
                    # elif "OK" in a_dic.values():
                    #     isUpFlag, isUpMsg =OpenPlatform_App_runStatus_judge(integ_isapi,
                    #         vcaType)
                    #     if not isUpFlag:
                    #         return False, isUpMsg
                    #
                    #     return True, f"所有支持的通道切换智能资源{vcaType}成功,已完成重启!"
                    # else:
                    #     return False, f"设备通道切换智能资源{vcaType}协议执行失败!"

                else:
                    return False, f"设备通道切换智能资源{vcaType}协议执行失败!"
            else:
                # 获取当前设备该通道智能资源
                chan_current_type = {}

                vcaResourceUrl = '/ISAPI/System/RelatedVCAResource?format=json'
                vcaResource = self.get(vcaResourceUrl)
                print(vcaResource)
                if "RelatedVCAResource" not in vcaResource:
                    return False, f"设备通道获取关联智能资源能力失败:{self.get_subStatusCode(vcaResource)}!"

                channelsResource = json.loads(
                    vcaResource)['RelatedVCAResource']['ChannelsResource']
                print(channelsResource)

                for item in channelsResource:
                    chan_current_type[str(
                        item['channelID'])] = item['resourceType']

                print(chan_current_type)

                # 关联智能资源能力
                relatedVCAResourceCapUrl = "/ISAPI/System/RelatedVCAResource/capabilities?format=json"
                relatedVCAResourceCap = self.get(
                    relatedVCAResourceCapUrl)
                print(relatedVCAResourceCap)
                if "RelatedVCAResourceCap" not in relatedVCAResourceCap:
                    return False, f"设备通道获取关联智能资源能力失败:{self.get_subStatusCode(relatedVCAResourceCap)}"

                # 主通道号
                MainChannelId = ''
                # 关联关系表
                relationResuleList = []
                # 处理关联关系并生成关联关系表
                capList = json.loads(relatedVCAResourceCap
                                     )['RelatedVCAResourceCap']['ResourceList']
                for item in capList:
                    tempDict = {}
                    for key, value in item.items():
                        if key == "MainChannelResource":
                            MainChannelId = value['channelID']
                            MainChannelType = value['resourceType']
                            tempDict[str(MainChannelId)] = MainChannelType
                        elif key == "RelatedChannelsResource":
                            for sonChannel in value:
                                sonChannelId = sonChannel['channelID']
                                sonChannelType = sonChannel['resourceType']
                                tempDict[str(sonChannelId)] = sonChannelType
                    relationResuleList.append(tempDict)

                print(MainChannelId)
                print(relationResuleList)

                flag, paras = self.getReceiveDicParas(vcaType, relationResuleList)
                print(paras)

                putData = {"RelatedVCAResource": {"ChannelsResource": []}}

                for key, value in paras.items():
                    temp = {"channelID": "1", "resourceType": "smart"}
                    temp["channelID"] = int(key)
                    if isinstance(value, str):
                        temp["resourceType"] = value
                    elif isinstance(value, list):
                        flag_find = 0
                        for i in value:
                            if i == vcaType:
                                temp["resourceType"] = i
                                flag_find = 1
                                break
                        if not flag_find:
                            temp["resourceType"] = value[0]
                    else:
                        pass
                    putData['RelatedVCAResource']['ChannelsResource'].append(temp)

                DATA = json.dumps(putData)
                print(DATA)

                # 切换智能资源
                # if dev_type == 'IPD' and channel == 1:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)
                # else:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)

                vcaResourceUrl = f'http://{ip}/ISAPI/System/RelatedVCAResource?format=json'
                print(f"[PUT {vcaResourceUrl}]")
                req = requests.put(vcaResourceUrl,
                                   data=DATA,
                                   auth=HTTPDigestAuth(username, password),
                                   timeout=300)
                status_code = req.status_code
                content = req.text
                req.close()
                print(content)

                if status_code == 200 and ("rebootRequired" in content or needReboot):
                    rebootFlag, rebootMsg = self.integ_reboot_device(deviceInfo)
                    if not rebootFlag:
                        return False, f"设备切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                    # if dev_type == 'IPD' and channel == 1:
                    #     moveFlag, moveMsg = self.put_device_position_cfg()
                    #     if not moveFlag:
                    #         return False, moveMsg
                    #     time.sleep(3)
                    #
                    # isUpFlag, isUpMsg = self.OpenPlatform_App_runStatus_judge(
                    #     vcaType)
                    # if not isUpFlag:
                    #     return False, isUpMsg

                    return True, f"设备切换智能资源{vcaType}成功,已完成重启!"
                elif status_code == 200 and "OK" in content:
                    isUpFlag, isUpMsg = self.OpenPlatform_App_runStatus_judge(vcaType)
                    if not isUpFlag:
                        return False, isUpMsg
                    return True, f"设备切换智能资源{vcaType}成功,无需重启!"
                else:
                    return False, f"设备配置切换智能资源为{vcaType}协议执行失败:{self.get_subStatusCode(content)}!"
        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备切换智能资源为{vcaType}出现异常:{str(e)}!"

    def channel_Intelligent_VCAResource_change_without_judge_runStatus(self,
                                                                        deviceInfo,
                                                                        channel=1,
                                                                        vcaType='facesnap',
                                                                        isRelative=False,
                                                                        isOpenHeop=False,
                                                                        appId=0,
                                                                        sleepTime=60):
        """
            智能资源切换接口

            参数:
                channel:通道号,从1开始
                vcaType:智能资源名称
                isRelative:是否为关联智能资源
                isOpenHeop:是否为开放HEOP架构
                appId:开放HEOP架构智能对应id
            返回:
                return 接口执行成功，成功信息
                return 接口执行失败，详细报错信息

                成功:
                    return True, “智能资源切换成功”
                失败
                    return False, “详细失败报错信息”
        """

        try:
            ip = deviceInfo["IP"]
            username = deviceInfo["U_NAME"]
            password = deviceInfo["P_WORD"]
            needReboot = False
            needRebootFlag, needRebootMsg = self.channel_Intelligent_VCAResource_change_isRebootRequired(channel, vcaType)
            if not needRebootFlag:
                print(needRebootMsg)
            else:
                needReboot = needRebootMsg

            if isOpenHeop:
                # 获取设备当前已加载所有智能资源
                openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
                openHeopData = self.get(openHeopUrl)
                print(openHeopData)
                if "</AppList>" not in openHeopData:
                    return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{self.get_subStatusCode(openHeopData)}!"

                idList = re.findall("<id>(.*?)</id>", openHeopData, re.S)
                packageNameList = re.findall(
                    "<packageName>(.*?)</packageName>", openHeopData, re.S)
                runStatusList = re.findall("<runStatus>(.*?)</runStatus>",
                                           openHeopData, re.S)

                # 先停用所有已启用的智能
                for i in range(len(runStatusList)):
                    if runStatusList[i] == "true":
                        stopUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{idList[i]}/stop"
                        print(f"[PUT {stopUrl}]")
                        ret = requests.put(stopUrl,
                                           data=None,
                                           auth=HTTPDigestAuth(
                                               username, password),
                                           timeout=180)
                        status_code = ret.status_code
                        content = ret.text
                        ret.close()
                        print(content)
                        if "OK" not in content:
                            return False, f"在正式测试目标智能前关闭所有智能时出现关闭{packageNameList[i]}智能协议执行失败:{self.get_subStatusCode(content)}!"
                        time.sleep(5)
                    else:
                        continue

                # # 掉电记忆
                # if self.dict["dev_type"] == 'IPD' and channel == 1:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)
                # else:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)

                # 配置启用目标智能
                startUrl = f"http://{ip}/ISAPI/Custom/OpenPlatform/App/{appId}/start"
                print(f"[PUT {startUrl}]")
                req = requests.put(startUrl,
                                   data=None,
                                   auth=HTTPDigestAuth(username,
                                                       password),
                                   timeout=180)
                status_code = req.status_code
                content = req.text
                req.close()
                print(content)

                if status_code == 200 and ("rebootRequired" in content
                                           or needReboot):
                    rebootFlag, rebootMsg = self.integ_reboot_device(deviceInfo)
                    if not rebootFlag:
                        return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                    # if self.dict["dev_type"] == 'IPD' and channel == 1:
                    #     moveFlag, moveMsg = self.put_device_position_cfg()
                    #     if not moveFlag:
                    #         return False, moveMsg
                    #     time.sleep(3)

                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
                elif status_code == 200 and "OK" in content:
                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
                else:
                    return False, f"设备通道{channel}切换智能资源{vcaType}协议执行失败:{self.get_subStatusCode(content)}!"
            elif not isRelative:
                # 获取当前设备该通道智能资源
                vcaResourceUrl = f'/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
                ret = self.get(vcaResourceUrl)
                print(ret)
                if "</VCAResource>" not in ret:
                    return False, f'设备通道{channel}获取当前智能资源协议执行失败:{self.get_subStatusCode(ret)}!'

                currentType = re.findall("<type>(.*?)</type>", ret, re.S)[0]
                print("currentType=", currentType)

                # 当前智能资源与需要切换的智能资源比较
                # if currentType == vcaType:
                #     if self.dict["dev_type"] == 'IPD' and channel == 1:
                #         print("等待60s保证掉电记忆配置保存!")
                #         time.sleep(sleepTime)
                #         moveFlag, moveMsg = self.put_device_position_cfg()
                #         if not moveFlag:
                #             return False, moveMsg
                #         time.sleep(3)
                #     else:
                #         print("等待60s保证掉电记忆配置保存!")
                #         time.sleep(sleepTime)
                #     return True, f"设备通道{channel}当前智能资源就是{vcaType},无需切换!"

                # 切换智能资源操作
                # if self.dict["dev_type"] == 'IPD' and channel == 1:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)
                # else:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)

                vcaResourceUrl = f'http://{ip}/ISAPI/System/Video/inputs/channels/{channel}/VCAResource'
                DATA = f"""
                    <VCAResource xmlns="http://www.hikvision.com/ver20/XMLSchema" version="2.0">
                        <type>{vcaType}</type>
                    </VCAResource>
                    """
                print(DATA)
                print(f"[PUT {vcaResourceUrl}]")
                req = requests.put(vcaResourceUrl,
                                   data=DATA,
                                   auth=HTTPDigestAuth(username,
                                                       password),
                                   timeout=200)
                status_code = req.status_code
                content = req.text
                req.close()
                print(content)

                if status_code == 200 and ("rebootRequired" in content
                                           or needReboot):
                    rebootFlag, rebootMsg = self.integ_reboot_device(deviceInfo)
                    if not rebootFlag:
                        return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                    # if self.dict["dev_type"] == 'IPD' and channel == 1:
                    #     moveFlag, moveMsg = self.put_device_position_cfg()
                    #     if not moveFlag:
                    #         return False, moveMsg
                    #     time.sleep(3)

                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
                elif status_code == 200 and "OK" in content:
                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
                else:
                    return False, f"设备通道{channel}切换智能资源{vcaType}协议执行失败:{self.get_subStatusCode(content)}!"
            else:
                # 获取当前设备该通道智能资源
                chan_current_type = {}

                vcaResourceUrl = '/ISAPI/System/RelatedVCAResource?format=json'
                vcaResource = self.get(vcaResourceUrl)
                print(vcaResource)
                if "RelatedVCAResource" not in vcaResource:
                    return False, f"设备通道{channel}获取关联智能资源能力失败:{self.get_subStatusCode(vcaResource)}!"

                channelsResource = json.loads(
                    vcaResource)['RelatedVCAResource']['ChannelsResource']
                print(channelsResource)

                for item in channelsResource:
                    chan_current_type[str(
                        item['channelID'])] = item['resourceType']

                print(chan_current_type)

                # 当前智能资源与需要切换的智能资源比较
                if chan_current_type[str(channel)] == vcaType:
                    # if self.dict["dev_type"] == 'IPD' and channel == 1:
                    #     print("等待60s保证掉电记忆配置保存!")
                    #     time.sleep(sleepTime)
                    #     moveFlag, moveMsg = self.put_device_position_cfg()
                    #     if not moveFlag:
                    #         return False, moveMsg
                    #     time.sleep(3)
                    # else:
                    #     print("等待60s保证掉电记忆配置保存!")
                    #     time.sleep(sleepTime)
                    return True, f"设备通道{channel}当前智能资源就是{vcaType},无需切换!"

                # 关联智能资源能力
                relatedVCAResourceCapUrl = "/ISAPI/System/RelatedVCAResource/capabilities?format=json"
                relatedVCAResourceCap = self.get(
                    relatedVCAResourceCapUrl)
                print(relatedVCAResourceCap)
                if "RelatedVCAResourceCap" not in relatedVCAResourceCap:
                    return False, f"设备通道{channel}获取关联智能资源能力失败:{self.get_subStatusCode(relatedVCAResourceCap)}!"

                # 主通道号
                MainChannelId = ''
                # 关联关系表
                relationResuleList = []
                # 处理关联关系并生成关联关系表
                capList = json.loads(relatedVCAResourceCap
                                     )['RelatedVCAResourceCap']['ResourceList']
                for item in capList:
                    tempDict = {}
                    for key, value in item.items():
                        if key == "MainChannelResource":
                            MainChannelId = value['channelID']
                            MainChannelType = value['resourceType']
                            tempDict[str(MainChannelId)] = MainChannelType
                        elif key == "RelatedChannelsResource":
                            for sonChannel in value:
                                sonChannelId = sonChannel['channelID']
                                sonChannelType = sonChannel['resourceType']
                                tempDict[str(sonChannelId)] = sonChannelType
                    relationResuleList.append(tempDict)

                print(MainChannelId)
                print(relationResuleList)

                paras = {}
                for item in relationResuleList:
                    if vcaType in item[str(channel)]:
                        paras = item
                        break
                print(paras)

                putData = {"RelatedVCAResource": {"ChannelsResource": []}}

                for index in range(1, len(paras) + 1):
                    temp = {"channelID": "1", "resourceType": "smart"}
                    if index == channel:
                        temp["channelID"] = index
                        temp["resourceType"] = vcaType
                    else:
                        temp["channelID"] = index
                        if isinstance(paras[str(index)], list):
                            temp["resourceType"] = paras[str(index)][0]
                        else:
                            temp["resourceType"] = paras[str(index)]
                    putData['RelatedVCAResource']['ChannelsResource'].append(
                        temp)

                DATA = json.dumps(putData)
                self.VD_LOG_DEBUG(self.ip, DATA)

                # 切换智能资源
                # if self.dict["dev_type"] == 'IPD' and channel == 1:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)
                #     moveFlag, moveMsg = self.put_device_position_cfg()
                #     if not moveFlag:
                #         return False, moveMsg
                #     time.sleep(3)
                # else:
                #     print("等待60s保证掉电记忆配置保存!")
                #     time.sleep(sleepTime)

                vcaResourceUrl = f'http://{ip}/ISAPI/System/RelatedVCAResource?format=json'
                print(f"[PUT {vcaResourceUrl}]")
                req = requests.put(vcaResourceUrl,
                                   data=DATA,
                                   auth=HTTPDigestAuth(username,
                                                       password),
                                   timeout=200)
                status_code = req.status_code
                content = req.text
                req.close()
                print(content)

                if status_code == 200 and ("rebootRequired" in content
                                           or needReboot):
                    rebootFlag, rebootMsg = self.integ_reboot_device(deviceInfo)
                    if not rebootFlag:
                        return False, f"设备通道{channel}切换智能资源{vcaType}后重启失败:{rebootMsg}!"

                    # if self.dict["dev_type"] == 'IPD' and channel == 1:
                    #     moveFlag, moveMsg = self.put_device_position_cfg()
                    #     if not moveFlag:
                    #         return False, moveMsg
                    #     time.sleep(3)

                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,已完成重启!"
                elif status_code == 200 and "OK" in content:

                    return True, f"设备通道{channel}切换智能资源{vcaType}成功,无需重启!"
                else:
                    return False, f"设备通道{channel}配置切换智能资源为{vcaType}协议执行失败:{self.get_subStatusCode(content)}!"
        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备通道{channel}切换智能资源为{vcaType}出现异常:{str(e)}!"

    def Smart_isSupport_judge_with_Type(self, smartType=None):
        '''smart智能支持的事件能力判断'''
        try:
            smartCapUrl = "/ISAPI/Smart/capabilities"
            ret = self.get(smartCapUrl)
            if "</SmartCap>" not in ret:
                return False, f'设备获取Smart智能能力协议执行失败:{self.get_subStatusCode(ret)}!'

            supportCap = []
            for item in smartType:
                if f"<isSupport{item}>true</isSupport{item}>" in ret:
                    supportCap.append(item)

            return True, supportCap
        except Exception as e:
            print(str(e))
            return False, "获取smart事件能力失败"

    def integ_channel_Intelligent_VCAResource_judeg_and_change(self,
                                                            deviceInfo,
                                                            channel=1,
                                                            vcaType='facesnap',
                                                            isReturnErr=True,
                                                            sleepTime=60):
        """判断是否支持某智能、切换智能、判断智能是否起来"""
        try:
            # 智能资源切换能力判断
            capFlag, capMsg = self.channel_VCAResourse_cap_judge(channel, vcaType)
            if not capFlag:
                return False, capMsg

            if not capMsg[1]:
                return True, False

            # 设备智能资源之间是否存在关联
            isRelative = capMsg[2]

            # 设备是否为开放HEOP架构
            isOpenHeop = capMsg[3]

            appId = capMsg[4] if isOpenHeop else 0

            # 延时60s保证掉电记忆配置保存
            # 智能资源切换
            changeFlag, changeMsg = self.channel_Intelligent_VCAResource_change_without_judge_runStatus(deviceInfo,
                                                                                                       channel,
                                                                                                       vcaType,
                                                                                                       isRelative,
                                                                                                       isOpenHeop,
                                                                                                       appId,
                                                                                                       sleepTime=sleepTime)
            if not changeFlag:
                return False, changeMsg

            # 判断智能是否起来
            isUpFlag, isUpMsg = self.OpenPlatform_App_runStatus_judge(vcaType)
            if not isUpFlag and isReturnErr:
                return False, isUpMsg
            elif not isUpFlag and not isReturnErr:
                if "是否起来的状态依然是false" in isUpMsg:
                    return True, isUpMsg
                else:
                    return False, isUpMsg
            else:
                return True, True
        except Exception as e:
            print(traceback.print_exc())
            return False, f"设备通道{channel}判断智能{vcaType}是否支持且切换该智能判断智能是否起来的操作出现异常:{str(e)}!"

    def is_heop(self):
        '''开放HEOP能力判断'''
        try:
            openHeopCapUrl = "/ISAPI/Custom/OpenPlatform/capabilities"
            openHeopCap = self.get(openHeopCapUrl)
            print(openHeopCap)

            if "notSupport" in openHeopCap or "invalidOperation" in openHeopCap or "404 -- Not Found" in openHeopCap or "</isSupportOpenPlatform>" not in openHeopCap:
                return False, "", None

            if "</OpenPlatformCap>" not in openHeopCap:
                return False, f"通道/ISAPI/Custom/OpenPlatform/capabilities接口判断设备是否为开放HEOP平台设备协议执行失败:{self.get_xml_subStatusCode(openHeopCap)}!", None

            isSupportOpenPlatform = re.findall(
                "<isSupportOpenPlatform>(.*?)</isSupportOpenPlatform>",
                openHeopCap, re.S)[0]

            if isSupportOpenPlatform == "false":
                return False, "", None

            # 判断HEOP的版本
            openHeopVerUrl = '/ISAPI/Custom/OpenPlatform/algInfo?format=json'
            openHeopVerUrlData = self.get(openHeopVerUrl)
            print(openHeopVerUrlData)
            openHeopVerUrlData = json.loads(openHeopVerUrlData)
            openHeopVer = openHeopVerUrlData["HEOPVersion"] if "HEOPVersion" in openHeopVerUrlData else None

            openHeopUrl = "/ISAPI/Custom/OpenPlatform/App"
            openHeopData = self.get(openHeopUrl)
            print(openHeopData)
            if "</AppList>" not in openHeopData:
                return False, f"通过/ISAPI/Custom/OpenPlatform/App接口获取设备所有智能列表协议执行失败:{self.get_subStatusCode(openHeopData)}!", openHeopVer
            if "</App>" not in openHeopData:
                return False, "", openHeopVer
            packageNameList = re.findall("<packageName>(.*?)</packageName>",
                                         openHeopData, re.S)
            if len(packageNameList) == 0:
                return True, "", openHeopVer
            else:
                return True, packageNameList, openHeopVer
        except Exception as e:
            return False, str(e), None


    def Integ_support_list(self, chans_num, ip, is_AIOP, AIOP_list):
        '''
        获取智能资源【兼容新老架构】

        :params:
            chans_num: 设备通道数
            ip： 设备ip
            is_AIOP: 智能是否基于新框架
        :return:
            case1:智能资源能力集不为空，返回list格式的能力集（协议完全执行成功，能力集为全集；协议部分执行成功，能力集为子集）
            case2:智能资源能力集为空，返回str类型的信息（协议完全执行成功，返回“未获取到智能资源能力”； 协议执行失败， 返回errMsg）
        '''
        supportList = []
        # 1.能力判断
        # 1.1.AIOP设备
        errMsg = ''
        try:
            if is_AIOP:
                supportList = AIOP_list
            # 1.2.非AIOP设备
            else:
                # 1.2.1.通道循环
                for current_chan in range(1, chans_num + 1):
                    try:
                        # 每个通道智能资源的临时存放列表
                        resultCapList = []
                        # 关联智能资源能力
                        relatedVCAResourceCapUrl = "/ISAPI/System/RelatedVCAResource/capabilities?format=json"
                        relatedVCAResourceCap = self.get(relatedVCAResourceCapUrl)
                        self.VD_LOG_DEBUG(ip, relatedVCAResourceCap)

                        # 非关联智能资源能力
                        vcaResourceCapUrl = f"/ISAPI/System/Video/inputs/channels/{current_chan}/VCAResource/capabilities"
                        vcaResourceCap = self.get(vcaResourceCapUrl)
                        self.VD_LOG_DEBUG(ip, vcaResourceCap)

                        ##判断是否为非关联智能
                        if ("notSupport" not in vcaResourceCap and "invalidOperation" not in vcaResourceCap) and "</VCAResource>" in vcaResourceCap:
                            resultCapList = re.findall('<type opt="(.*?)">',
                                                       vcaResourceCap, re.S)[0].split(',')
                            supportList += resultCapList
                        elif ("notSupport" not in relatedVCAResourceCap and "invalidOperation" not in relatedVCAResourceCap
                        ) and "RelatedVCAResourceCap" in relatedVCAResourceCap:
                            capList = json.loads(relatedVCAResourceCap
                                                 )["RelatedVCAResourceCap"]["ResourceList"]
                            for item in capList:
                                tempDict = {}
                                for key, value in item.items():
                                    if key == "MainChannelResource":
                                        MainChannelId = value['channelID']
                                        MainChannelType = value['resourceType']
                                        tempDict[str(MainChannelId)] = MainChannelType
                                    elif key == "RelatedChannelsResource":
                                        for sonChannel in value:
                                            sonChannelId = sonChannel['channelID']
                                            sonChannelType = sonChannel['resourceType']
                                            tempDict[str(sonChannelId)] = sonChannelType
                                resultCapList.append(tempDict)
                            # 由于关联性智能与通道无关，所以执行了一次之后直接break
                            for item in resultCapList:
                                for k, v in item.items():
                                    if isinstance(v, list):
                                        supportList += v
                                    else:
                                        supportList.append(v)
                            break
                    except Exception as e:
                        errMsg += f'通道{current_chan}获取智能资源能力失败：{e}'
                        self.VD_LOG_DEBUG(ip, errMsg)
        except Exception as e:
            errMsg += f"获取人脸抓拍智能能力失败:{str(e)}"
            self.VD_LOG_DEBUG(ip, str(errMsg))
        finally:
            if supportList is not None:
                return supportList
            else:
                if not errMsg:
                    return "未获取到抓拍能力集"
                else:
                    return errMsg

    def get_deviceInfo(self, type = 'deviceDescription'):
        '''

        :param type:
            tyoe: 需要获取设备信息，默认获取设备类型，IPC/IPD
        :return:
            是否执行成功, 设备信息
        '''
        try:
            url = '/ISAPI/System/deviceInfo'
            devInfo = self.get(url)
            if "notSupport" in devInfo or "invalidOperation" in devInfo or "404 -- Not Found" in devInfo:
                return False, ""

            if "</DeviceInfo>" not in devInfo:
                return False, f"通过/ISAPI/System/deviceInfo接口获取设备信息协议执行失败:{self.get_xml_subStatusCode(devInfo)}!"

            res = re.findall(f"<{type}>(.*?)</{type}>",devInfo, re.S)[0]
            if res == 'IPCamera':
                return True, 'IPC'
            elif res == 'IPDome':
                return True, 'IPD'
        except Exception as e:
            print(traceback.print_exc())
            return False, f"获取设备信息操作失败:{str(e)}!"


    # 低功耗设备切换为全功耗模式
    def fullPowerConsumption(self, ip):
        '''切换为全功耗模式'''

        url = '/ISAPI/System/consumptionMode?format=json'

        # 1.默认功耗模式获取
        consumptionCap = self.get(url)
        consumptionCap = json.loads(consumptionCap)
        if "ConsumptionMode" in consumptionCap:
            if consumptionCap["ConsumptionMode"]["devWorkMode"] == "fullPowerConsumption":
                return True
            else:
                # 2.全功耗模式切换
                data = '''
                {
                    "ConsumptionMode": {
                        "enabled": true,
                        "devWorkMode": "fullPowerConsumption",
                        "Sleep": {
                            "enabled": true
                        },
                        "powerThreshold": 20,
                        "TimingSleep": {
                            "enabled": false,
                            "TimeRange": []
                        }
                    }
                }
                '''
                res = self.put(url, data)
                res = json.loads(res)
                if res['subStatusCode'] != 'ok':
                    return False

                # 3.全功耗模式校验
                consumptionCap = self.get(url)
                consumptionCap = json.loads(consumptionCap)
                if consumptionCap["ConsumptionMode"]["devWorkMode"] == "fullPowerConsumption":
                    return True
        else:
            return False

    # 低时延设备切换为非普通监控模式
    def lowLatencytoIntelVCA(self):
        for type in self.IntegSupportList:
            if type != 'close':
                capFlag, capMsg = self.integ_channel_Intelligent_VCAResource_judeg_and_change(self.deviceInfo,
                                                                            vcaType=type)
                if capMsg == True and capMsg == True:
                    self.VD_LOG_DEBUG(self.ip, f'当前设备启用的智能为【{type}】', level='INFO')
                    return capFlag, type
        return capFlag, capMsg

    # 获取设备全部通道和码流的分辨率
    def get_all_channel_stream(self):
        # 获取设备全部通道和码流的分辨率
        chans_num = self.chans_num
        ip = self.ip
        all_cap = []
        try:
            for current_chan in range(1, chans_num + 1):
                # 2.1通道制式能力判断
                chanlCapUrl = '/ISAPI/Image/channels/%d/capabilities' % current_chan
                cap = self.get(chanlCapUrl)
                if '</ImageChannel>' not in cap:
                    msg = f"通道{current_chan}获取图像参数能力协议执行失败:{self.get_subStatusCode(cap)}!"
                    self.VD_LOG_DEBUG(ip, msg)
                    return False, msg
                # 获取PN制，当不存在时，国内默认为P制50hz，海外默认为N制60hz

                if '</powerLineFrequency>' in cap:
                    ret = self.isapi_get(
                        chanlCapUrl,
                        path=[".//powerLineFrequency/powerLineFrequencyMode"],
                        type=["attrib"],
                        attrib=['opt'])
                    if not ret['success']:
                        msg = f"通道{current_chan}获取视频制式能力集协议执行失败:{ret['msg']}!"
                        self.VD_LOG_DEBUG(ip, msg)
                        return False, msg
                    FrequencyModeList = ret['data'][0][0].split(',')
                else:
                    # 如果未获取到PN制则切换场景尝试
                    # 获取场景切换
                    powerLineFrequencyFlag = 0
                    mountingScenarioUrl = '/ISAPI/Image/channels/%d/mountingScenario/capabilities' % current_chan
                    ret = self.isapi_get(mountingScenarioUrl, path=[".//mode"], type=["attrib"], attrib=['opt'])
                    if not ret['success']:
                        msg = f"获取通道{current_chan}场景信息协议执行失败:{ret['msg']}!"
                        self.VD_LOG_DEBUG(ip, msg)
                        return msg
                    mountingScenariolist = ret["data"][0][0].split(",")

                    for mountingScenario in mountingScenariolist:
                        res_change_url = f"/ISAPI/Image/channels/{current_chan}/mountingScenario"
                        data = f"<MountingScenario><mode>{mountingScenario}</mode></MountingScenario>"
                        res_change = self.put(res_change_url, data)
                        if "OK" not in res_change:
                            msg = f"通道{current_chan}切换图像参数能力协议执行失败:{self.get_subStatusCode(res_change)}!"
                            self.VD_LOG_DEBUG(ip, msg)
                            return False, msg
                        cap = self.get(chanlCapUrl)
                        if '</powerLineFrequency>' in cap:
                            powerLineFrequencyFlag = 1
                            break

                    if powerLineFrequencyFlag == 0:
                        if not self.isOverseas:
                            FrequencyModeList = ["50hz"]
                        else:
                            FrequencyModeList = ["60hz"]
                    else:
                        ret = self.isapi_get(
                            chanlCapUrl,
                            path=[".//powerLineFrequency/powerLineFrequencyMode"],
                            type=["attrib"],
                            attrib=['opt'])
                        if not ret['success']:
                            msg = f"通道{current_chan}获取视频制式能力集协议执行失败:{ret['msg']}!"
                            self.VD_LOG_DEBUG(ip, msg)
                            return False, msg
                        FrequencyModeList = ret['data'][0][0].split(',')

                FrequencyModeMsg = {}
                # 获取当前通道当前制式的帧率和码流分辨率
                for FrequencyMode in FrequencyModeList:
                    # 切换制式
                    FrequencyModeUrl = '/ISAPI/Image/channels/%d/powerLineFrequency' % current_chan
                    ret = self.set_by_xpath(
                        FrequencyModeUrl,
                        "//*[local-name()='powerLineFrequencyMode']", FrequencyMode)
                    if 'OK' not in ret and "rebootRequired" not in ret:
                        msg = f'通道{current_chan}配置视频制式为{FrequencyMode}协议执行失败:{self.get_subStatusCode(ret)}!'
                        self.VD_LOG_DEBUG(ip, msg)
                        return False, msg
                    time.sleep(6)

                    # 视频制式确认
                    ret = self.isapi_get(FrequencyModeUrl, path=[".//powerLineFrequencyMode"], type=["value"])
                    if not ret['success']:
                        msg = f'通道{current_chan}视频制式切换为{FrequencyMode}后重新获取视频制式参数协议执行失败:{ret["msg"]}!'
                        self.VD_LOG_DEBUG(ip, msg)
                        return False, msg
                    getFrequencyMode = ret["data"][0][0]
                    if getFrequencyMode != FrequencyMode:
                        msg = f'通道{current_chan}修改视频制式为{FrequencyMode}后重新获取视频制式校验失败，当前制式为{getFrequencyMode}!'
                        self.VD_LOG_DEBUG(ip, msg)
                        return False, msg
                    # 获取所有码率的能力集
                    stream_url = f"/ISAPI/Streaming/channels/{current_chan}01/capabilities"
                    stream_ret = self.isapi_get(stream_url, path=[".//id"], type=["attrib"], attrib=['opt'])
                    if not stream_ret['success']:
                        msg = f"通道{current_chan}获取码率能力集协议执行失败:{stream_ret['msg']}!"
                        self.VD_LOG_DEBUG(ip, msg)
                        return msg
                    streamlist = stream_ret['data'][0][0].split(",")
                    streamMsg = []
                    for stream in streamlist:
                        FrameUrl = f"/ISAPI/Streaming/channels/{current_chan}0{stream}/capabilities"
                        Frame_ret = self.isapi_get(FrameUrl,
                                                   path=[
                                                       ".//Video/videoResolutionWidth",
                                                       ".//Video/videoResolutionHeight",
                                                       ".//Video/maxFrameRate"
                                                   ],
                                                   type=["attrib", "attrib", "attrib"],
                                                   attrib=['opt', 'opt', 'opt'])
                        if not Frame_ret['success']:
                            msg = f"通道{current_chan}第{stream}码流获取视频分辨率能力集协议执行失败:{Frame_ret['msg']}!"
                            self.VD_LOG_DEBUG(ip, msg)
                            return msg
                        videoResolutionWidth, videoResolutionHeight, videoResolutionFrameRate = Frame_ret['data']
                        videoResolutionWidth = videoResolutionWidth[0].split(",")
                        videoResolutionHeight = videoResolutionHeight[0].split(",")
                        Max_FrameRate = videoResolutionFrameRate[0].split(",")
                        Max_FrameRate = [int(x) for x in Max_FrameRate]
                        FrameRate = max(Max_FrameRate) / 100.0
                        FrameRateMsg = ""
                        if FrameRate != 12.5:
                            FrameRateMsg = f"{int(FrameRate)} fps"
                        else:
                            FrameRateMsg = f"{FrameRate} fps"
                        FrameMsg = []
                        for i in range(len(videoResolutionWidth)):
                            FrameMsg.append(f"{videoResolutionWidth[i]} × {videoResolutionHeight[i]}")
                        FrameMsg = sorted(FrameMsg, key=lambda x: (int(x.split('×')[0].strip()), int(x.split('×')[1].strip())),reverse=True)
                        streamMsg.append(FrameRateMsg + "(" + ','.join(FrameMsg) + ")")
                    FrequencyModeMsg[FrequencyMode] = streamMsg
                all_cap.append(FrequencyModeMsg)
            return True, all_cap
        except Exception as e:
            self.VD_LOG_DEBUG(ip, str(e))
            return False, f"获取码流分辨率能力异常，异常原因：{str(e)}"

    # 获取曝光时间
    def getAllShutter(self):
        # 获取快门能力，如果有smart需要先切到smart
        # 不同情况下快门会变化，暂时返回默认情况下的快门数值
        errMsg = ""
        chans_num = self.chans_num
        ip = self.ip
        # 是否关联
        isRelative = False
        # 所支持的通道
        isSupportChal = []

        # 是否为开放HEOP
        isOpenHeop = 0
        ShutterLevelCap_min = 0
        ShutterLevelCap_max = 0
        allRes = {}
        for current_chan in range(1, chans_num + 1):
            try:
                # 2.1.智能资源切换Smart模式能力判断。
                capFlag, capMsg = self.channel_VCAResourse_cap_judge(current_chan, "smart")
                if capFlag:
                    # 设备智能资源之间是否存在关联
                    isRelative = capMsg[2]
                    if not isRelative and capMsg[1]:
                        isSupportChal.append(str(current_chan))

                    if isRelative:
                        changeFlag, changeMsg = self.channel_Intelligent_VCAResource_change(self.deviceInfo,
                                                                                            [], "smart", isRelative,
                                                                                            isOpenHeop, 1)
                        if not changeFlag:
                            self.VD_LOG_DEBUG(ip, changeMsg)
                            self.VD_LOG_DEBUG(ip, changeMsg)
                    else:
                        changeFlag, changeMsg = self.channel_Intelligent_VCAResource_change(self.deviceInfo,
                                                                                            isSupportChal, "smart",
                                                                                            isRelative,
                                                                                            isOpenHeop, 1)
                        if not changeFlag:
                            self.VD_LOG_DEBUG(ip, changeMsg)
                            self.VD_LOG_DEBUG(ip, changeMsg)

                ret = self.isapi_get('/ISAPI/Image/channels/%d/capabilities' % current_chan,
                                     path=[".//Shutter/ShutterLevel"],
                                     type=["attrib"],
                                     attrib=['opt'])
                if not ret["success"]:
                    msg = f"通道{current_chan}获取快门能力失败，失败原因:{ret['msg']}"
                    self.VD_LOG_DEBUG(ip, msg)
                    errMsg += msg
                    continue

                ShutterLevelCapList = ret["data"][0][0].split(",")
                allRes[current_chan] = ShutterLevelCapList
                self.VD_LOG_DEBUG(ip, "ShutterLevel ", ShutterLevelCapList)
                # if int(ShutterLevelCapList[-1]) > ShutterLevelCap_max:
                #     ShutterLevelCap_max = int(ShutterLevelCapList[-1])

            except Exception as e:
                self.VD_LOG_DEBUG(ip, traceback.print_exc())
                msg = f"设备通道{current_chan}获取快门能力出现异常:{str(e)}!"
                errMsg += msg

        if len(allRes) >= 1:
            return allRes[1]
        else:
            return errMsg


    # 调整名称校验报告的格式
    @staticmethod
    def checkItemNameExcel(excel):
        '''调整名称校验报告的格式'''
        # 报告格式
        custom_first_font = Font(name='Arial', size=16, bold=True, color=Color(rgb='000000'))
        custom_second_font = Font(name='微软雅黑', size=12, bold=True)

        cell = excel['A1']
        cell.value = "SPEC文件名称规范校验报告"
        cell.font = custom_first_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill('solid', fgColor='483D8B')
        excel.merge_cells('A1:B1')

        cell = excel['A2']
        cell.value = "本文件中使用的规格项名称"
        cell.font = custom_second_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill('solid', fgColor='D3D3D3')

        cell = excel['B2']
        cell.value = "最新版本SPEC数据库使用的规格项名称"
        cell.font = custom_second_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill('solid', fgColor='D3D3D3')

        cell = excel['C1']
        cell.value = '''备注：\n目前已针对已知的不符合【最新SPEC模板】的名称进行测试兼容，若测试过程中发现仍存在由于名称不规范导致出现漏测项，请反馈至软件组长及工具开发组！'''
        excel.merge_cells('C1:C2')
        cell.font = Font(name='Calibri', size=12, bold=True, color=Color(rgb='CD2626'))
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 调整报告宽度
        excel.column_dimensions['A'].width = 40
        excel.row_dimensions[1].height = 40
        excel.column_dimensions['B'].width = 40
        excel.row_dimensions[2].height = 30
        excel.column_dimensions['C'].width = 50

    # 根据【规格项类型】/【模块名】/【子项】对dataframe进行排序
    @staticmethod
    def sortExcelReport(save_path):
        df = pd.read_excel(save_path, engine='openpyxl', skiprows=3)

        ### 生成排序索引字典
        typeDict = {'产品参数': 0, '软件参数': 1}
        moduleDict = {}
        for i, module in enumerate(list(df['模块名'])):
            if module not in moduleDict:
                moduleDict[module] = i

        typeSortList = [typeDict.get(i, -1) for i in list(df['规格项类型'])]
        moduleSortList = [moduleDict.get(i, -1) for i in list(df['模块名'])]
        totalSortList = list(zip(typeSortList, moduleSortList, list(df.index)))
        newSortList = sorted(totalSortList, key=lambda x: (x[0], x[1], [2]))
        newIndex = [i[2] for i in newSortList]
        newdf = df.iloc[newIndex]
        return newdf

# if __name__ == '__main__':
#     test = MtdComFun()
#     a = '预置点，巡航扫描，自动扫描，垂直扫描，随机扫描，帧扫描，全景扫描'
#     b = '预置点，花样扫描，巡航扫描，自动扫描，垂直扫描，帧扫描，全景扫描'
#     print(test.resHandleGeneral_1(a, b, fun=test.resCompGeneral_1))